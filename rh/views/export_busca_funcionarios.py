"""
Exportação PDF e Excel dos resultados da busca avançada de funcionários.
Cabeçalho alinhado ao padrão Vale Transporte (logo + bloco de texto da empresa).

Parâmetro GET opcional ``relatorio`` (relatórios pré-definidos na página de Relatórios):
situação (ativos, experiencia, demitidos, aviso_previo, ferias) e saúde
(pcmso_vencimentos, aso_ultimo) — altera título, colunas e critério de listagem.
"""
import re
from datetime import date
from io import BytesIO
from typing import NamedTuple, Optional
from types import SimpleNamespace
from xml.sax.saxutils import escape as xml_escape

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from empresas.models import Empresa

from controles_rh.views.cesta_export import (
    PDF_LANDSCAPE_CONTENT_MM,
    _flowables_header_compact,
    _flowables_titulo_pdf_centro,
    _linha_extra_cabecalho_recibo,
    _nome_empresa_pdf,
    _unlink_temp_logo_paths,
)
from controles_rh.views.pdf_rodape import flowables_rodape_impressao

from ..models import Funcionario

from .base import _empresa_ativa_or_redirect
from .funcionarios import queryset_funcionarios_busca_avancada

RELATORIOS_SAUDE = frozenset({'pcmso_vencimentos', 'aso_ultimo'})


def _safe_filename_part(text):
    s = re.sub(r'[^\w\s\-]', '_', str(text), flags=re.UNICODE)
    s = re.sub(r'\s+', '_', s.strip())
    return (s[:60] or 'busca').strip('_')


def _header_text_html_busca_funcionarios(empresa, _comp, ctx):
    """Mesmo bloco textual que VT (nome fantasia, razão, linha de contexto, CNPJ/end/tel, e-mail)."""
    nome = _nome_empresa_pdf(empresa)
    bits = [f'<b>{xml_escape(nome)}</b>']
    razao = (empresa.razao_social or '').strip()
    if razao and razao.upper() != (nome or '').upper():
        bits.append(f'<font size="8" color="#64748b">{xml_escape(razao)}</font>')
    subtitulo = (getattr(ctx, 'subtitulo', None) or 'Busca de funcionários').strip()
    bits.append(xml_escape(subtitulo))
    extra = _linha_extra_cabecalho_recibo(empresa)
    if extra:
        bits.append(xml_escape(extra))
    email = (getattr(empresa, 'email', None) or '').strip()
    if email:
        bits.append(xml_escape(f'E-mail: {email}'))
    return '<br/>'.join(bits)


def _fmt_date(d):
    if d is None:
        return '—'
    if isinstance(d, date):
        return d.strftime('%d/%m/%Y')
    return str(d)


def _nome_maiusculo(f) -> str:
    n = (getattr(f, 'nome', None) or '').strip()
    return n.upper() if n else '—'


def _pcmso_vencimento_e_dias(f, hoje: date):
    """
    Próxima data de vencimento PCMSO (a menor futura); se só houver passadas, a mais recente vencida.
    """
    datas = [p.data_vencimento for p in f.pcmso_registros.all() if p.data_vencimento]
    if not datas:
        return None, '—'
    futuras = [d for d in datas if d >= hoje]
    if futuras:
        d = min(futuras)
        return d, str((d - hoje).days)
    d = max(d for d in datas if d < hoje)
    return d, str((d - hoje).days)


class _AsoUltimoResumo(NamedTuple):
    data: Optional[date]
    dias: str
    tipo: str


def _aso_ultimo_resumo(f, hoje: date) -> _AsoUltimoResumo:
    """ASO mais recente: data, dias até hoje e tipo (display)."""
    asos_com_data = [a for a in f.asos.all() if a.data]
    if not asos_com_data:
        return _AsoUltimoResumo(None, '—', '—')
    ult = max(asos_com_data, key=lambda a: (a.data, a.pk))
    d = ult.data
    tipo_lbl = ult.get_tipo_display() if getattr(ult, 'tipo', None) else '—'
    return _AsoUltimoResumo(d, str((hoje - d).days), tipo_lbl)


def _queryset_ativos_saude_export(empresa_ativa):
    return (
        Funcionario.objects.filter(empresa=empresa_ativa)
        .exclude(situacao_atual__in=['demitido', 'inativo'])
        .select_related('cargo', 'tipo_contrato')
        .prefetch_related('pcmso_registros', 'asos', 'ferias', 'afastamentos')
        .order_by('nome')
    )


def _faltam_dias_experiencia(fim: Optional[date], hoje: date) -> str:
    if fim is None:
        return '—'
    return str((fim - hoje).days)


def _ferias_gozo_atual(f, hoje: date):
    """Período de gozo que inclui hoje, se existir."""
    for fer in f.ferias.all():
        gi = fer.gozo_inicio
        gf = fer.gozo_fim
        if gi and gf and gi <= hoje <= gf:
            return gi, gf
    return None, None


def _relatorio_key(request) -> Optional[str]:
    key = (request.GET.get('relatorio') or '').strip().lower()
    allowed = frozenset(
        {
            'ativos',
            'experiencia',
            'demitidos',
            'aviso_previo',
            'ferias',
            'pcmso_vencimentos',
            'aso_ultimo',
        }
    )
    return key if key in allowed else None


def _pdf_widths_from_weights(weights):
    total = float(PDF_LANDSCAPE_CONTENT_MM)
    s = sum(weights)
    return [total * w / s for w in weights]


def build_export_table(request, funcionarios: list, hoje: Optional[date] = None):
    """
    Monta cabeçalhos, linhas e metadados para PDF e XLSX.
    ``funcionarios`` deve ser lista já avaliada do queryset (com prefetch esperado).
    """
    if hoje is None:
        hoje = timezone.localdate()

    rk = _relatorio_key(request)
    total = len(funcionarios)
    agora = timezone.localtime(timezone.now())
    emit = agora.strftime('%d/%m/%Y %H:%M')

    if rk is None:
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Situação',
        ]
        headers_pdf = ['#', 'NOME', 'MATRÍCULA', 'CPF', 'CARGO', 'SITUAÇÃO']
        weights = [0.35, 2.0, 0.95, 1.0, 1.35, 1.5]
        titulo = 'LISTAGEM — BUSCA DE FUNCIONÁRIOS'
        subtitulo_ctx = f'Busca de funcionários · Emitido em {emit}'
        xlsx_head = f'Busca de funcionários — Emitido em {emit}'
        filename_slug = 'Busca_funcionarios'
        sheet_title = 'Busca'
        pdf_doc_title = 'Busca de funcionários'
        body_font = 7

        rows = []
        for n, f in enumerate(funcionarios, start=1):
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    f.get_situacao_atual_display(),
                ]
            )
    elif rk == 'ativos':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Data de admissão',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'DATA ADMISSÃO',
        ]
        weights = [0.35, 2.0, 0.95, 1.0, 1.45, 1.2]
        titulo = 'LISTAGEM — FUNCIONÁRIOS ATIVOS'
        subtitulo_ctx = f'Funcionários ativos · Emitido em {emit}'
        xlsx_head = f'Funcionários ativos — Emitido em {emit}'
        filename_slug = 'Relatorio_funcionarios_ativos'
        sheet_title = 'Ativos'
        pdf_doc_title = 'Funcionários ativos'
        body_font = 7
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    _fmt_date(f.data_admissao),
                ]
            )
    elif rk == 'experiencia':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Data de admissão',
            'Início prorrogação',
            'Fim da experiência',
            'Faltam (dias)',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'DATA ADMISSÃO',
            'INÍCIO PRORROGAÇÃO',
            'FIM EXPERI\u00eaNCIA',
            'FALTAM DIAS',
        ]
        weights = [0.32, 1.55, 0.82, 0.88, 1.05, 0.95, 0.95, 0.95, 0.78]
        titulo = 'FUNCIONÁRIO EM EXPERI\u00caNCIA'
        subtitulo_ctx = f'Em experiência · Emitido em {emit}'
        xlsx_head = f'Em experiência · Emitido em {emit}'
        filename_slug = 'Relatorio_experiencia'
        sheet_title = 'Experiência'
        pdf_doc_title = 'Experiência probatória'
        body_font = 6
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    _fmt_date(f.data_admissao),
                    _fmt_date(f.inicio_prorrogacao),
                    _fmt_date(f.fim_prorrogacao),
                    _faltam_dias_experiencia(f.fim_prorrogacao, hoje),
                ]
            )
    elif rk == 'demitidos':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Data de admissão',
            'Data de demissão',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'DATA ADMISSÃO',
            'DATA DEMISSÃO',
        ]
        weights = [0.32, 1.75, 0.88, 0.92, 1.15, 1.05, 1.08]
        titulo = 'LISTAGEM — FUNCIONÁRIOS DEMITIDOS'
        subtitulo_ctx = f'Funcionários demitidos · Emitido em {emit}'
        xlsx_head = f'Funcionários demitidos — Emitido em {emit}'
        filename_slug = 'Relatorio_demitidos'
        sheet_title = 'Demitidos'
        pdf_doc_title = 'Funcionários demitidos'
        body_font = 6.5
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    _fmt_date(f.data_admissao),
                    _fmt_date(f.data_demissao),
                ]
            )
    elif rk == 'aviso_previo':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Início aviso',
            'Fim aviso',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'INÍCIO AVISO',
            'FIM AVISO',
        ]
        weights = [0.32, 1.75, 0.88, 0.92, 1.15, 1.1, 1.13]
        titulo = 'LISTAGEM — AVISO PRÉVIO ATIVO'
        subtitulo_ctx = f'Aviso prévio ativo · Emitido em {emit}'
        xlsx_head = f'Aviso prévio ativo — Emitido em {emit}'
        filename_slug = 'Relatorio_aviso_previo'
        sheet_title = 'Aviso prévio'
        pdf_doc_title = 'Aviso prévio ativo'
        body_font = 6.5
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    _fmt_date(f.data_inicio_aviso),
                    _fmt_date(f.data_fim_aviso),
                ]
            )
    elif rk == 'ferias':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Gozo início',
            'Gozo fim',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'GOZO INÍCIO',
            'GOZO FIM',
        ]
        weights = [0.32, 1.75, 0.88, 0.92, 1.15, 1.1, 1.13]
        titulo = 'LISTAGEM — FÉRIAS (GOZO EM CURSO)'
        subtitulo_ctx = f'Em férias · Emitido em {emit}'
        xlsx_head = f'Em férias · Emitido em {emit}'
        filename_slug = 'Relatorio_ferias'
        sheet_title = 'Férias'
        pdf_doc_title = 'Férias'
        body_font = 6.5
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            gi, gf = _ferias_gozo_atual(f, hoje)
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    _fmt_date(gi),
                    _fmt_date(gf),
                ]
            )
    elif rk == 'pcmso_vencimentos':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Vencimento PCMSO',
            'Dias restantes',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'VENCIMENTO PCMSO',
            'DIAS RESTANTES',
        ]
        weights = [0.3, 1.55, 0.82, 0.88, 1.05, 1.0, 0.95, 0.82]
        titulo = 'LISTAGEM — PCMSO (VENCIMENTOS)'
        subtitulo_ctx = f'Saúde · PCMSO · Emitido em {emit}'
        xlsx_head = f'PCMSO — vencimentos — Emitido em {emit}'
        filename_slug = 'Relatorio_pcmso_vencimentos'
        sheet_title = 'PCMSO'
        pdf_doc_title = 'PCMSO vencimentos'
        body_font = 6
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            dv, dias = _pcmso_vencimento_e_dias(f, hoje)
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    _fmt_date(dv),
                    dias,
                ]
            )
    elif rk == 'aso_ultimo':
        headers = [
            '#',
            'Nome',
            'Matrícula',
            'CPF',
            'Cargo',
            'Tipo ASO',
            'Último ASO',
            'Dias desde último',
        ]
        headers_pdf = [
            '#',
            'NOME',
            'MATRÍCULA',
            'CPF',
            'CARGO',
            'TIPO ASO',
            '\u00daLTIMO ASO',
            'DIAS DESDE \u00daLTIMO',
        ]
        weights = [0.28, 1.35, 0.78, 0.84, 0.98, 0.95, 0.92, 0.9, 0.82]
        titulo = 'LISTAGEM — ASO (ÚLTIMO REGISTRO)'
        subtitulo_ctx = f'Saúde · ASO · Emitido em {emit}'
        xlsx_head = f'ASO — último registro — Emitido em {emit}'
        filename_slug = 'Relatorio_aso_ultimo'
        sheet_title = 'ASO'
        pdf_doc_title = 'ASO último registro'
        body_font = 6
        rows = []
        for n, f in enumerate(funcionarios, start=1):
            aso_r = _aso_ultimo_resumo(f, hoje)
            rows.append(
                [
                    str(n),
                    _nome_maiusculo(f),
                    (f.matricula or '—').strip(),
                    (f.cpf or '—').strip(),
                    (f.cargo.nome if f.cargo else '—'),
                    aso_r.tipo,
                    _fmt_date(aso_r.data),
                    aso_r.dias,
                ]
            )
    else:
        raise ValueError(f'Layout de relatório não tratado: {rk!r}')

    data_rows_pdf = [headers_pdf] + rows
    cw_mm = _pdf_widths_from_weights(weights)

    return SimpleNamespace(
        relatorio=rk,
        headers=headers,
        headers_pdf=headers_pdf,
        data_rows_pdf=data_rows_pdf,
        col_widths_mm=cw_mm,
        titulo=titulo,
        subtitulo_ctx=subtitulo_ctx,
        xlsx_head=xlsx_head,
        filename_slug=filename_slug,
        sheet_title=sheet_title,
        pdf_doc_title=pdf_doc_title,
        body_font=body_font,
        total=total,
        rows_xlsx=rows,
    )


@login_required
def exportar_busca_funcionarios_pdf(request):
    empresa_ativa, redirect_response = _empresa_ativa_or_redirect(
        request,
        'Selecione uma empresa para exportar a busca.',
    )
    if redirect_response:
        return redirect_response

    empresa = Empresa.objects.get(pk=empresa_ativa.pk)
    hoje = timezone.localdate()
    rk = _relatorio_key(request)
    if rk in RELATORIOS_SAUDE:
        funcionarios = list(_queryset_ativos_saude_export(empresa_ativa))
    else:
        funcionarios = list(queryset_funcionarios_busca_avancada(request, empresa_ativa))
    cfg = build_export_table(request, funcionarios, hoje=hoje)

    agora = timezone.localtime(timezone.now())
    ctx_header = SimpleNamespace(
        subtitulo=cfg.subtitulo_ctx,
    )
    comp_dummy = SimpleNamespace(referencia='')

    styles = getSampleStyleSheet()
    data_rows = cfg.data_rows_pdf

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=cfg.pdf_doc_title,
    )
    story = []
    temp_logo_paths = []
    story.extend(
        _flowables_header_compact(
            empresa,
            comp_dummy,
            ctx_header,
            styles,
            _header_text_html_busca_funcionarios,
            temp_logo_paths,
        )
    )
    story.extend(_flowables_titulo_pdf_centro(cfg.titulo, styles))

    meta_style = ParagraphStyle(
        'busca_meta',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#334155'),
        spaceAfter=6,
    )
    story.append(
        Paragraph(
            xml_escape(f'Total de registros: {cfg.total}'),
            meta_style,
        )
    )
    story.append(Spacer(1, 2 * mm))

    last_idx = len(data_rows) - 1
    table = Table(
        data_rows,
        colWidths=[w * mm for w in cfg.col_widths_mm],
        repeatRows=1,
    )
    pdf_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), cfg.body_font),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
    ]
    if last_idx >= 1:
        pdf_style.append(
            (
                'ROWBACKGROUNDS',
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor('#f8fafc')],
            )
        )
    table.setStyle(TableStyle(pdf_style))
    story.append(table)
    story.extend(flowables_rodape_impressao(request, styles, space_before_mm=3))

    doc.build(story)
    _unlink_temp_logo_paths(temp_logo_paths)
    buf.seek(0)

    stamp = agora.strftime('%Y%m%d_%H%M')
    filename = f'{cfg.filename_slug}_{_safe_filename_part(empresa_ativa)}_{stamp}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def exportar_busca_funcionarios_xlsx(request):
    empresa_ativa, redirect_response = _empresa_ativa_or_redirect(
        request,
        'Selecione uma empresa para exportar a busca.',
    )
    if redirect_response:
        return redirect_response

    empresa = Empresa.objects.get(pk=empresa_ativa.pk)
    hoje = timezone.localdate()
    rk = _relatorio_key(request)
    if rk in RELATORIOS_SAUDE:
        funcionarios = list(_queryset_ativos_saude_export(empresa_ativa))
    else:
        funcionarios = list(queryset_funcionarios_busca_avancada(request, empresa_ativa))
    cfg = build_export_table(request, funcionarios, hoje=hoje)

    headers = cfg.headers
    ncols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = cfg.sheet_title[:31]

    row = 1
    if getattr(empresa, 'logo', None) and empresa.logo:
        try:
            with empresa.logo.open('rb') as f:
                xl_img = XLImage(BytesIO(f.read()))
            max_w = 220
            if xl_img.width > max_w:
                ratio = max_w / xl_img.width
                xl_img.width = max_w
                xl_img.height = int(xl_img.height * ratio)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            ws.add_image(xl_img, 'A1')
            ws.row_dimensions[row].height = max(50, min(xl_img.height * 0.78, 130))
            row += 1
        except Exception:
            pass

    agora = timezone.localtime(timezone.now())
    nome_pdf = _nome_empresa_pdf(empresa)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(
        row=row,
        column=1,
        value=f'{nome_pdf} — {cfg.xlsx_head}',
    )
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

    extra = _linha_extra_cabecalho_recibo(empresa)
    if extra:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c2 = ws.cell(row=row, column=1, value=extra)
        c2.font = Font(size=9)
        c2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c3 = ws.cell(row=row, column=1, value=f'Total de registros: {cfg.total}')
    c3.font = Font(bold=True, size=10)
    row += 1

    for col, h in enumerate(headers, 1):
        hc = ws.cell(row=row, column=col, value=h)
        hc.font = Font(bold=True)
        hc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row += 1

    for valores in cfg.rows_xlsx:
        for col, val in enumerate(valores, 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    stamp = agora.strftime('%Y%m%d_%H%M')
    filename = f'{cfg.filename_slug}_{_safe_filename_part(empresa_ativa)}_{stamp}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
