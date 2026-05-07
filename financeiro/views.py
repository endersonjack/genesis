"""Views do módulo financeiro."""
from __future__ import annotations

import calendar
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
import re
from urllib.parse import urlencode
from xml.sax.saxutils import escape as xml_escape

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Count,
    DecimalField,
    ExpressionWrapper,
    F,
    OuterRef,
    Q,
    Sum,
    Subquery,
    Value,
)
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from auditoria.models import RegistroAuditoria
from auditoria.registry import registrar_auditoria
from core.moeda_fmt import parse_valor_moeda_br
from core.moeda_fmt import format_decimal_br_moeda
from core.urlutils import redirect_empresa, reverse_empresa
from fornecedores.models import Fornecedor
from obras.models import Obra
from rh.models import Funcionario

from .forms import (
    BoletoPagamentoForm,
    CaixaEditForm,
    CaixaNovoForm,
    ContaBancariaForm,
    CategoriaFinanceiraForm,
    BoletoRascunhoFormSet,
    PagamentoImpostoForm,
    PagamentoImpostoItemFormSet,
    PagamentoPessoalForm,
    PagamentoPessoalItemFormSet,
    PagamentoNotaFiscalItemEditFormSet,
    PagamentoNotaFiscalForm,
    PagamentoNotaFiscalItemFormSet,
    PagamentoNotaFiscalPagamentoEditFormSet,
    PagamentoNotaFiscalPagamentoFormSet,
    PagamentoBancarioPagarForm,
    PagamentoBancarioParcelaValorForm,
    PagamentoBancarioAvulsoForm,
    PagamentoBancarioRecorrenteForm,
    nf_numero_exige_unicidade_para_fornecedor,
    RecebimentoAvulsoEditForm,
    RecebimentoAvulsoForm,
    RecebimentoLiquidacaoForm,
    RecebimentoMedicaoEditForm,
    RecebimentoMedicaoForm,
)
from .models import (
    BoletoPagamento,
    AutoridadeTributaria,
    Caixa,
    ContaBancaria,
    CategoriaFinanceira,
    MovimentoCaixa,
    PagamentoNotaFiscal,
    PagamentoNotaFiscalItem,
    PagamentoNotaFiscalPagamento,
    PagamentoImposto,
    PagamentoImpostoItem,
    PagamentoBancarioParcela,
    PagamentoBancarioAvulso,
    PagamentoBancarioRecorrente,
    PagamentoPessoal,
    PagamentoPessoalItem,
    RecebimentoAvulso,
    RecebimentoMedicao,
)

ULTIMOS_RECEBIMENTOS_LISTA_LIMITE = 50

MESES_FILTRO_CAIXA = (
    (1, 'Janeiro'),
    (2, 'Fevereiro'),
    (3, 'Março'),
    (4, 'Abril'),
    (5, 'Maio'),
    (6, 'Junho'),
    (7, 'Julho'),
    (8, 'Agosto'),
    (9, 'Setembro'),
    (10, 'Outubro'),
    (11, 'Novembro'),
    (12, 'Dezembro'),
)


def _empresa(request):
    return getattr(request, 'empresa_ativa', None)


def _is_htmx(request) -> bool:
    return str(request.headers.get('HX-Request', '')).lower() == 'true'


def _periodo_caixa_params(request) -> dict:
    hoje = timezone.localdate()
    mes_raw = str(request.GET.get('mes') or '').strip().lower()
    todos_os_meses = mes_raw in {'todos', 'all', '0'}
    try:
        mes = int(mes_raw or hoje.month)
    except (TypeError, ValueError):
        mes = hoje.month
    try:
        ano = int(str(request.GET.get('ano') or hoje.year).replace('.', '').replace(',', ''))
    except (TypeError, ValueError):
        ano = hoje.year
    if todos_os_meses:
        mes = 0
    elif mes < 1 or mes > 12:
        mes = hoje.month
    if ano < 2000 or ano > hoje.year + 5:
        ano = hoje.year

    if mes == 0:
        inicio = date(ano, 1, 1)
        fim = date(ano + 1, 1, 1)
    else:
        inicio = date(ano, mes, 1)
        fim = date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1)
    return {
        'mes': mes,
        'ano': ano,
        'inicio': inicio,
        'fim': fim,
        'meses': MESES_FILTRO_CAIXA,
        'todos_os_meses': mes == 0,
    }


def _filtrar_linhas_periodo(linhas: list[dict], inicio, fim) -> list[dict]:
    return [
        linha
        for linha in linhas
        if linha.get('data') and inicio <= linha['data'] < fim
    ]


def _linha_extrato_movimento_efetivo(linha: dict) -> bool:
    """Inclui o lançamento em totais do período e no saldo após cada linha.

    Parcelas de pagamento bancário **em aberto** aparecem no extrato detalhado apenas
    como compromisso; o efeito no caixa conta quando estiver **paga** (consolidado).
    """

    return not linha.get('extrato_apenas_comprometido')


def _recalcular_saldo_linhas(linhas: list[dict], campo_valor: str) -> list[dict]:
    linhas.sort(key=lambda item: item['sort_key'])
    saldo = Decimal('0')
    for linha in linhas:
        if not _linha_extrato_movimento_efetivo(linha):
            linha['saldo_apos_lancamento'] = saldo
            continue
        valor = linha.get(campo_valor) or Decimal('0')
        saldo = saldo + valor if linha['entrada'] else saldo - valor
        linha['saldo_apos_lancamento'] = saldo
    linhas = list(reversed(linhas))
    for idx, linha in enumerate(linhas, start=1):
        linha['numero'] = idx
    return linhas


def _query_string(base: dict) -> str:
    return '&'.join(
        f'{chave}={valor}'
        for chave, valor in base.items()
        if valor not in (None, '')
    )


def _int_param(request, nome: str) -> int | None:
    try:
        valor = int(request.GET.get(nome) or '')
    except (TypeError, ValueError):
        return None
    return valor if valor > 0 else None


def _somar_meses(data_ref: date, quantidade: int) -> date:
    mes = data_ref.month - 1 + quantidade
    ano = data_ref.year + mes // 12
    mes = mes % 12 + 1
    dia = min(data_ref.day, calendar.monthrange(ano, mes)[1])
    return date(ano, mes, dia)


def _data_pagamento_bancario(data_inicio: date, dia_pagamento: int, indice_zero: int) -> date:
    mes_ref = _somar_meses(data_inicio.replace(day=1), indice_zero)
    dia = min(dia_pagamento, calendar.monthrange(mes_ref.year, mes_ref.month)[1])
    return date(mes_ref.year, mes_ref.month, dia)


def _meses_entre(inicio: date, fim: date) -> int:
    return (fim.year - inicio.year) * 12 + (fim.month - inicio.month)


def _periodo_mensal_atual():
    hoje = timezone.localdate()
    inicio = hoje.replace(day=1)
    fim = date(hoje.year + 1, 1, 1) if hoje.month == 12 else date(hoje.year, hoje.month + 1, 1)
    return inicio, fim


def _formset_tem_erros(formset) -> bool:
    return any(bool(errors) for errors in formset.errors) or bool(
        formset.non_form_errors()
    )


def _pagamento_nf_tem_erros_validacao(form, itens_fs, pagamentos_fs, boletos_fs) -> bool:
    return (
        bool(form.errors)
        or bool(form.non_field_errors())
        or _formset_tem_erros(itens_fs)
        or _formset_tem_erros(pagamentos_fs)
        or _formset_tem_erros(boletos_fs)
    )


def _active_tab_pagamento_nf(request, form, itens_fs, pagamentos_fs, boletos_fs) -> str:
    requested = (request.GET.get('tab') or 'descricao').strip().lower()
    if request.method != 'POST':
        return requested
    if form.errors or form.non_field_errors():
        return 'descricao'
    if _formset_tem_erros(itens_fs):
        return 'itens'
    if _formset_tem_erros(pagamentos_fs) or _formset_tem_erros(boletos_fs):
        return 'pagamento'
    return requested


def _recebimento_para_linha(recebimento, tipo: str) -> dict:
    movimento = getattr(recebimento, 'movimento', None)
    return {
        'pk': recebimento.pk,
        'tipo': tipo,
        'tipo_label': 'Avulso' if tipo == 'avulso' else 'Medição',
        'data': recebimento.data,
        'data_pagamento': recebimento.data_pagamento,
        'caixa': recebimento.caixa,
        'conta_bancaria': recebimento.conta_bancaria,
        'cliente': recebimento.cliente,
        'medicao': getattr(recebimento, 'medicao_numero', ''),
        'nf': getattr(recebimento, 'nota_fiscal_numero', ''),
        'obra': getattr(recebimento, 'obra', None),
        'valor': recebimento.valor,
        'impostos': recebimento.impostos,
        'valor_liquido': recebimento.valor_liquido,
        'descricao': recebimento.descricao,
        'status': recebimento.status,
        'movimento': movimento,
    }


def _totais_recebimentos(recebimentos: list[dict]) -> dict:
    return {
        'valor': sum((r['valor'] for r in recebimentos), Decimal('0')),
        'impostos': sum((r['impostos'] for r in recebimentos), Decimal('0')),
        'valor_liquido': sum((r['valor_liquido'] for r in recebimentos), Decimal('0')),
    }


def _criar_movimento_recebimento(recebimento, categoria_origem, data_liquidacao=None):
    data_liquidacao = data_liquidacao or timezone.localdate()
    mov = MovimentoCaixa(
        empresa=recebimento.empresa,
        caixa=recebimento.caixa,
        natureza=MovimentoCaixa.Natureza.ENTRADA,
        categoria_origem=categoria_origem,
        valor=recebimento.valor_liquido,
        data=data_liquidacao,
        descricao=recebimento.descricao,
        observacao=recebimento.observacao,
    )
    mov.full_clean()
    mov.save()
    recebimento.movimento = mov
    recebimento.status = recebimento.Status.PAGO
    recebimento.data_pagamento = mov.data
    recebimento.save(
        update_fields=(
            'movimento',
            'status',
            'data_pagamento',
            'conta_bancaria',
            'valor',
            'impostos',
            'valor_liquido',
            'atualizado_em',
        )
    )
    return mov


def _aplicar_situacao_boleto(boleto: BoletoPagamento, hoje=None) -> BoletoPagamento:
    hoje = hoje or timezone.localdate()
    if boleto.status == BoletoPagamento.Status.PAGO:
        boleto.situacao_label = 'Pago'
        boleto.situacao_badge_class = 'text-bg-success'
    elif boleto.status == BoletoPagamento.Status.CANCELADO:
        boleto.situacao_label = 'Cancelado'
        boleto.situacao_badge_class = 'text-bg-dark'
    elif boleto.vencimento < hoje:
        boleto.situacao_label = 'Vencido'
        boleto.situacao_badge_class = 'text-bg-danger'
    elif boleto.vencimento == hoje:
        boleto.situacao_label = 'Vence Hoje'
        boleto.situacao_badge_class = 'text-bg-warning'
    else:
        boleto.situacao_label = 'À vencer'
        boleto.situacao_badge_class = 'text-bg-secondary'
    return boleto


def _resumo_pagamento_nf(pagamentos, boletos, total_itens: Decimal, hoje=None) -> dict:
    hoje = hoje or timezone.localdate()
    pagamentos = list(pagamentos or [])
    boletos_validos = [
        boleto for boleto in boletos
        if boleto.status != BoletoPagamento.Status.CANCELADO
    ]
    pagamentos_diretos = [
        pagamento for pagamento in pagamentos
        if pagamento.tipo
        in (
            PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA,
            PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO,
        )
    ]
    tem_config_boletos = any(
        pagamento.tipo == PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS
        for pagamento in pagamentos
    ) or bool(boletos_validos)
    total_pagamentos_diretos = sum(
        (pagamento.total_a_pagar() for pagamento in pagamentos_diretos),
        Decimal('0'),
    )
    total_boletos_pago = sum(
        (boleto.valor_pago or Decimal('0') for boleto in boletos_validos),
        Decimal('0'),
    )
    total_pago = total_pagamentos_diretos + total_boletos_pago
    valor_em_aberto = max(total_itens - total_pago, Decimal('0'))

    labels = []
    if pagamentos_diretos:
        tipos = []
        if any(p.tipo == PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA for p in pagamentos_diretos):
            tipos.append('À vista')
        if any(p.tipo == PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO for p in pagamentos_diretos):
            tipos.append('Crédito')
        labels.append(' + '.join(tipos))
    if tem_config_boletos:
        parcelas = len(boletos_validos)
        labels.append(f'Boletos ({parcelas} parcela{"s" if parcelas != 1 else ""})')
    pagamento_label = ' + '.join(labels) if labels else 'Sem pagamento'

    boletos_abertos = [
        boleto for boleto in boletos_validos
        if boleto.status != BoletoPagamento.Status.PAGO
    ]
    tem_boleto_vencido = any(boleto.vencimento < hoje for boleto in boletos_abertos)

    if total_pago >= total_itens and total_itens > 0:
        situacao_label = 'Pago'
        situacao_badge_class = 'text-bg-success'
    elif tem_boleto_vencido:
        situacao_label = 'Vencido'
        situacao_badge_class = 'text-bg-danger'
    elif total_pago > 0:
        situacao_label = 'Pago Parcial'
        situacao_badge_class = 'text-bg-primary'
    else:
        situacao_label = 'Não pago'
        situacao_badge_class = 'text-bg-secondary'

    return {
        'pagamento_label': pagamento_label,
        'situacao_label': situacao_label,
        'situacao_badge_class': situacao_badge_class,
        'valor_pago': total_pago,
        'valor_em_aberto': valor_em_aberto,
        'total_pagamentos_diretos': total_pagamentos_diretos,
        'total_boletos_pago': total_boletos_pago,
    }


def _data_ultimo_pagamento_nf(pagamentos, boletos) -> 'date | None':
    datas = []
    for p in pagamentos or []:
        if getattr(p, 'data', None) and p.total_a_pagar() > 0:
            datas.append(p.data)
    for b in boletos or []:
        if getattr(b, 'data_pagamento', None):
            datas.append(b.data_pagamento)
    return max(datas) if datas else None


def _busca_pagamentos_nf_corresponde_status(
    resumo: dict,
    *,
    status_vencido: bool,
    status_aberto: bool,
    status_pago: bool,
) -> bool:
    situacao = resumo['situacao_label']
    valor_pago = resumo.get('valor_pago') or Decimal('0')
    return (
        (status_vencido and situacao == 'Vencido')
        or (status_aberto and situacao in ('Não pago', 'Pago Parcial'))
        or (status_pago and valor_pago > 0)
    )


def _status_busca_pagamento_nf(resumo: dict, boletos, total_itens: Decimal, hoje=None) -> str:
    hoje = hoje or timezone.localdate()
    valor_pago = resumo.get('valor_pago') or Decimal('0')
    boletos = list(boletos or [])
    boletos_abertos = [
        boleto for boleto in boletos
        if boleto.status not in (BoletoPagamento.Status.PAGO, BoletoPagamento.Status.CANCELADO)
    ]
    if valor_pago >= total_itens and total_itens > 0:
        return 'Pago Completo'
    if any(boleto.vencimento < hoje for boleto in boletos_abertos):
        return 'Vencido'
    if any(boleto.vencimento == hoje for boleto in boletos_abertos):
        return 'Vence Hoje'
    if valor_pago > 0:
        return 'Pago Parcial'
    if boletos_abertos:
        return 'Em Aberto'
    return 'Sem Pagamento'


def _badge_status_busca_pagamento(status: str) -> str:
    return {
        'Pago Completo': 'text-bg-success',
        'Pago Parcial': 'text-bg-primary',
        'Vencido': 'text-bg-danger',
        'Vence Hoje': 'text-bg-warning',
        'Em Aberto': 'text-bg-secondary',
        'Sem Pagamento': 'text-bg-light text-dark border',
        'Pago': 'text-bg-success',
    }.get(status, 'text-bg-secondary')


def _parcela_dashboard_boleto(boleto: BoletoPagamento) -> str:
    """Ex.: parcela 1 de 4 boletos ativos (exclui cancelados)."""
    nf = boleto.pagamento_nf
    ativos = [
        x
        for x in nf.boletos.all()
        if x.status != BoletoPagamento.Status.CANCELADO
    ]
    if not ativos:
        return f'{boleto.parcela}/1'
    max_p = max(x.parcela for x in ativos)
    return f'{boleto.parcela}/{max_p}'


def _subquery_total_itens_nf():
    return (
        PagamentoNotaFiscalItem.objects.filter(pagamento_nf=OuterRef('pk'))
        .values('pagamento_nf')
        .annotate(total=Sum('valor_total'))
        .values('total')[:1]
    )


def _decimal_audit(valor) -> str:
    return format_decimal_br_moeda(valor or Decimal('0'))


def _date_audit(valor) -> str:
    return valor.isoformat() if valor else ''


def _snapshot_pagamento_nf(nf: PagamentoNotaFiscal) -> dict:
    nf = (
        PagamentoNotaFiscal.objects.filter(pk=nf.pk)
        .select_related('fornecedor', 'caixa')
        .prefetch_related('itens__categoria', 'pagamentos', 'boletos')
        .first()
    )
    if not nf:
        return {}
    return {
        'fornecedor': nf.fornecedor.nome if nf.fornecedor_id else '',
        'numero_nf': nf.numero_nf,
        'data_emissao': _date_audit(nf.data_emissao),
        'caixa': nf.caixa.nome if nf.caixa_id else '',
        'descricao': nf.descricao or '',
        'total_itens': _decimal_audit(nf.total_itens()),
        'itens': [
            {
                'descricao': item.descricao,
                'categoria': item.categoria.nome if item.categoria_id else '',
                'quantidade': _decimal_audit(item.quantidade),
                'unidade': item.unidade or '',
                'valor_total': _decimal_audit(item.valor_total),
            }
            for item in nf.itens.all().order_by('pk')
        ],
        'pagamentos': [
            {
                'tipo': pagamento.get_tipo_display(),
                'data': _date_audit(pagamento.data),
                'valor': _decimal_audit(pagamento.valor),
                'acrescimos': _decimal_audit(pagamento.acrescimos),
                'descontos': _decimal_audit(pagamento.descontos),
            }
            for pagamento in nf.pagamentos.all().order_by('data', 'pk')
        ],
        'boletos': [
            {
                'numero_doc': boleto.numero_doc,
                'parcela': boleto.parcela,
                'vencimento': _date_audit(boleto.vencimento),
                'status': boleto.get_status_display(),
                'valor': _decimal_audit(boleto.valor),
                'valor_pago': _decimal_audit(boleto.valor_pago or Decimal('0')),
                'data_pagamento': _date_audit(boleto.data_pagamento),
            }
            for boleto in nf.boletos.all().order_by('vencimento', 'parcela', 'pk')
        ],
    }


def _diff_snapshot_nf(antes: dict, depois: dict) -> list[str]:
    labels = {
        'fornecedor': 'Fornecedor',
        'numero_nf': 'Nº NF',
        'data_emissao': 'Data de emissão',
        'caixa': 'Caixa',
        'descricao': 'Descrição',
        'total_itens': 'Total dos itens',
    }
    alteracoes = []
    for key, label in labels.items():
        if antes.get(key) != depois.get(key):
            alteracoes.append(f'{label}: {antes.get(key) or "—"} → {depois.get(key) or "—"}')
    for key, label in (
        ('itens', 'Itens'),
        ('pagamentos', 'Pagamentos'),
        ('boletos', 'Boletos'),
    ):
        if antes.get(key) != depois.get(key):
            alteracoes.append(f'{label} atualizados')
    return alteracoes


def _audit_nf(
    request,
    nf: PagamentoNotaFiscal,
    *,
    acao: str,
    resumo: str,
    alteracoes: list[str] | None = None,
    extra: dict | None = None,
) -> None:
    detalhes = {
        'nf_pk': nf.pk,
        'numero_nf': nf.numero_nf,
        'fornecedor': nf.fornecedor.nome if nf.fornecedor_id else '',
        'alteracoes': alteracoes or [],
    }
    if extra:
        detalhes.update(extra)
    registrar_auditoria(
        request,
        acao=acao,
        modulo='financeiro',
        resumo=resumo,
        detalhes=detalhes,
        empresa=nf.empresa,
    )


def _initial_pagamentos_nf(nf):
    return [
        {
            'tipo': pagamento.tipo,
            'data': pagamento.data.isoformat() if pagamento.data else '',
            'valor': format_decimal_br_moeda(pagamento.valor),
            'acrescimos': format_decimal_br_moeda(pagamento.acrescimos),
            'descontos': format_decimal_br_moeda(pagamento.descontos),
            'conta_bancaria': pagamento.conta_bancaria_id or '',
            'observacao': pagamento.observacao,
        }
        for pagamento in nf.pagamentos.all().order_by('data', 'pk')
    ]


def _salvar_pagamentos_nf(nf, pagamentos_fs, boletos_fs, total_itens: Decimal) -> None:
    pagamentos = []
    tem_pagamento_boletos = False

    for form in pagamentos_fs:
        if not getattr(form, 'cleaned_data', None):
            continue
        if form.cleaned_data.get('DELETE'):
            continue
        tipo = form.cleaned_data.get('tipo')
        valor = form.cleaned_data.get('valor') or Decimal('0')
        observacao = (form.cleaned_data.get('observacao') or '').strip()
        acrescimos = form.cleaned_data.get('acrescimos') or Decimal('0')
        descontos = form.cleaned_data.get('descontos') or Decimal('0')
        data = form.cleaned_data.get('data') or nf.data_emissao
        if not tipo and valor == 0 and not observacao and acrescimos == 0 and descontos == 0:
            continue
        if not tipo:
            raise ValidationError('Informe o tipo de cada pagamento preenchido.')

        pagamento = form.save(commit=False)
        pagamento.pagamento_nf = nf
        pagamento.data = data
        if pagamento.tipo == PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS:
            tem_pagamento_boletos = True
            if not pagamento.valor or pagamento.valor == 0:
                pagamento.valor = total_itens
        pagamento.full_clean()
        pagamentos.append(pagamento)

    PagamentoNotaFiscalPagamento.objects.filter(pagamento_nf=nf).delete()
    nf.boletos.all().delete()

    for pagamento in pagamentos:
        pagamento.pk = None
        pagamento.save()

    if not tem_pagamento_boletos:
        return

    boletos = []
    for boleto_form in boletos_fs:
        if not getattr(boleto_form, 'cleaned_data', None):
            continue
        if boleto_form.cleaned_data.get('DELETE'):
            continue
        boletos.append(
            BoletoPagamento(
                pagamento_nf=nf,
                numero_doc=boleto_form.cleaned_data['numero_doc'],
                parcela=boleto_form.cleaned_data['parcela'],
                vencimento=boleto_form.cleaned_data['vencimento'],
                valor=boleto_form.cleaned_data['valor'],
                status=BoletoPagamento.Status.RASCUNHO,
            )
        )
    if not boletos:
        raise ValidationError('Pagamento por boletos exige gerar e salvar ao menos 1 boleto.')
    for boleto in boletos:
        boleto.full_clean()
    BoletoPagamento.objects.bulk_create(boletos)


def _annotate_saldo(qs):
    dec = DecimalField(max_digits=16, decimal_places=2)
    return (
        qs.order_by('tipo', 'nome')
        .annotate(
            _entradas=Coalesce(
                Sum(
                    'movimentos__valor',
                    filter=Q(movimentos__natureza='entrada'),
                ),
                Value(Decimal('0')),
                output_field=dec,
            ),
            _saidas=Coalesce(
                Sum(
                    'movimentos__valor',
                    filter=Q(movimentos__natureza='saida'),
                ),
                Value(Decimal('0')),
                output_field=dec,
            ),
        )
        .annotate(
            saldo=ExpressionWrapper(
                F('_entradas') - F('_saidas'),
                output_field=dec,
            ),
        )
    )


def _format_quantidade_extrato(valor: Decimal) -> str:
    valor = valor or Decimal('0')
    texto = format(valor.normalize(), 'f')
    return texto.rstrip('0').rstrip('.') if '.' in texto else texto


def _somar_por_caixa(saldos: dict[int, Decimal], rows, campo_caixa: str = 'caixa') -> None:
    for row in rows:
        caixa_id = row.get(campo_caixa)
        if not caixa_id:
            continue
        saldos[caixa_id] = saldos.get(caixa_id, Decimal('0')) + (row.get('total') or Decimal('0'))


def _subtrair_por_caixa(saldos: dict[int, Decimal], rows, campo_caixa: str = 'caixa') -> None:
    for row in rows:
        caixa_id = row.get(campo_caixa)
        if not caixa_id:
            continue
        saldos[caixa_id] = saldos.get(caixa_id, Decimal('0')) - (row.get('total') or Decimal('0'))


def _saldos_caixas_por_id(empresa) -> dict[int, Decimal]:
    dec = DecimalField(max_digits=16, decimal_places=2)
    zero = Value(Decimal('0'), output_field=dec)
    total_pagamento_direto = ExpressionWrapper(
        F('valor') + F('acrescimos') - F('descontos'),
        output_field=dec,
    )
    total_boleto_pago = Coalesce(
        'valor_pago',
        ExpressionWrapper(
            F('valor') + F('acrescimos') - F('descontos'),
            output_field=dec,
        ),
        output_field=dec,
    )
    saldos: dict[int, Decimal] = {}

    movimentos = (
        MovimentoCaixa.objects.filter(empresa=empresa)
        .values('caixa')
        .annotate(
            entradas=Coalesce(
                Sum('valor', filter=Q(natureza=MovimentoCaixa.Natureza.ENTRADA)),
                zero,
                output_field=dec,
            ),
            saidas=Coalesce(
                Sum('valor', filter=Q(natureza=MovimentoCaixa.Natureza.SAIDA)),
                zero,
                output_field=dec,
            ),
        )
    )
    for row in movimentos:
        caixa_id = row.get('caixa')
        if caixa_id:
            saldos[caixa_id] = (row.get('entradas') or Decimal('0')) - (
                row.get('saidas') or Decimal('0')
            )

    recebimentos_avulsos_sem_movimento = (
        RecebimentoAvulso.objects.filter(
            empresa=empresa,
            status=RecebimentoAvulso.Status.PAGO,
            movimento__isnull=True,
        )
        .values('caixa')
        .annotate(total=Coalesce(Sum('valor_liquido'), zero, output_field=dec))
    )
    _somar_por_caixa(saldos, recebimentos_avulsos_sem_movimento)

    recebimentos_medicao_sem_movimento = (
        RecebimentoMedicao.objects.filter(
            empresa=empresa,
            status=RecebimentoMedicao.Status.PAGO,
            movimento__isnull=True,
        )
        .values('caixa')
        .annotate(total=Coalesce(Sum('valor_liquido'), zero, output_field=dec))
    )
    _somar_por_caixa(saldos, recebimentos_medicao_sem_movimento)

    pagamentos_diretos = (
        PagamentoNotaFiscalPagamento.objects.filter(
            pagamento_nf__empresa=empresa,
            tipo__in=(
                PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA,
                PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO,
            ),
        )
        .values('pagamento_nf__caixa')
        .annotate(total=Coalesce(Sum(total_pagamento_direto), zero, output_field=dec))
    )
    _subtrair_por_caixa(saldos, pagamentos_diretos, campo_caixa='pagamento_nf__caixa')

    boletos_pagos = (
        BoletoPagamento.objects.filter(
            pagamento_nf__empresa=empresa,
            status=BoletoPagamento.Status.PAGO,
        )
        .values('pagamento_nf__caixa')
        .annotate(total=Coalesce(Sum(total_boleto_pago), zero, output_field=dec))
    )
    _subtrair_por_caixa(saldos, boletos_pagos, campo_caixa='pagamento_nf__caixa')

    pagamentos_pessoal = (
        PagamentoPessoalItem.objects.filter(pagamento__empresa=empresa)
        .values('pagamento__caixa')
        .annotate(total=Coalesce(Sum('valor_total'), zero, output_field=dec))
    )
    _subtrair_por_caixa(saldos, pagamentos_pessoal, campo_caixa='pagamento__caixa')
    pagamentos_impostos = (
        PagamentoImpostoItem.objects.filter(pagamento__empresa=empresa)
        .values('pagamento__caixa')
        .annotate(total=Coalesce(Sum('valor_total'), zero, output_field=dec))
    )
    _subtrair_por_caixa(saldos, pagamentos_impostos, campo_caixa='pagamento__caixa')
    pagamentos_bancarios = (
        PagamentoBancarioParcela.objects.filter(
            recorrencia__empresa=empresa,
            status=PagamentoBancarioParcela.Status.PAGO,
        )
        .values('recorrencia__caixa')
        .annotate(total=Coalesce(Sum('valor'), zero, output_field=dec))
    )
    _subtrair_por_caixa(saldos, pagamentos_bancarios, campo_caixa='recorrencia__caixa')
    pagamentos_bancarios_avulsos = (
        PagamentoBancarioAvulso.objects.filter(empresa=empresa)
        .values('caixa')
        .annotate(total=Coalesce(Sum('valor'), zero, output_field=dec))
    )
    _subtrair_por_caixa(saldos, pagamentos_bancarios_avulsos, campo_caixa='caixa')
    return saldos


def _caixas_com_saldo(empresa, *, somente_ativos: bool = True):
    qs = Caixa.objects.filter(empresa=empresa)
    if somente_ativos:
        qs = qs.filter(ativo=True)
    caixas = list(qs.order_by('tipo', 'nome'))
    saldos = _saldos_caixas_por_id(empresa)
    for caixa in caixas:
        caixa.saldo = saldos.get(caixa.pk, Decimal('0'))
    return caixas


def _aplicar_filtros_extrato(
    linhas: list[dict],
    *,
    fornecedor_id: int | None = None,
    categoria_id: int | None = None,
    caixa_id: int | None = None,
) -> list[dict]:
    filtradas = linhas
    if caixa_id:
        filtradas = [
            linha
            for linha in filtradas
            if linha.get('caixa_id') == caixa_id
        ]
    if fornecedor_id:
        filtradas = [
            linha
            for linha in filtradas
            if not linha.get('entrada') and linha.get('fornecedor_id') == fornecedor_id
        ]
    if categoria_id:
        filtradas = [
            linha
            for linha in filtradas
            if linha.get('categoria_id') == categoria_id
            or categoria_id in (linha.get('categoria_ids') or set())
        ]
    return filtradas


def _extrato_caixa(request, caixa: Caixa, inicio=None, fim=None) -> list[dict]:
    linhas = []

    movimentos = (
        MovimentoCaixa.objects.filter(caixa=caixa)
        .select_related(
            'recebimento_avulso',
            'recebimento_avulso__cliente',
            'recebimento_medicao',
            'recebimento_medicao__cliente',
            'recebimento_medicao__obra',
        )
        .order_by('data', 'pk')
    )
    for movimento in movimentos:
        natureza = movimento.natureza
        linhas.append(
            {
                'data': movimento.data,
                'sort_key': (movimento.data, movimento.pk, 10),
                'natureza': natureza,
                'origem': movimento.get_categoria_origem_display(),
                'descricao': movimento.descricao,
                'valor': movimento.valor,
                'entrada': natureza == MovimentoCaixa.Natureza.ENTRADA,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': None,
                'categoria_ids': set(),
                'referencia': f'MOV #{movimento.pk}',
                'url': '',
            }
        )

    recebimentos_avulsos = (
        RecebimentoAvulso.objects.filter(
            caixa=caixa,
            status=RecebimentoAvulso.Status.PAGO,
            movimento__isnull=True,
        )
        .select_related('cliente', 'categoria')
        .order_by('data_pagamento', 'data', 'pk')
    )
    for recebimento in recebimentos_avulsos:
        data = recebimento.data_pagamento or recebimento.data or recebimento.criado_em.date()
        linhas.append(
            {
                'data': data,
                'sort_key': (data, recebimento.pk, 20),
                'natureza': MovimentoCaixa.Natureza.ENTRADA,
                'origem': 'Recebimento avulso',
                'descricao': recebimento.descricao or f'Recebimento de {recebimento.cliente}',
                'valor': recebimento.valor_liquido,
                'entrada': True,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': recebimento.categoria_id,
                'categoria_ids': {recebimento.categoria_id} if recebimento.categoria_id else set(),
                'referencia': f'REC #{recebimento.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:recebimento_editar',
                    kwargs={'tipo': 'avulso', 'pk': recebimento.pk},
                ),
            }
        )

    recebimentos_medicao = (
        RecebimentoMedicao.objects.filter(
            caixa=caixa,
            status=RecebimentoMedicao.Status.PAGO,
            movimento__isnull=True,
        )
        .select_related('cliente', 'obra', 'categoria')
        .order_by('data_pagamento', 'data', 'pk')
    )
    for recebimento in recebimentos_medicao:
        data = recebimento.data_pagamento or recebimento.data or recebimento.criado_em.date()
        descricao = recebimento.descricao or f'Recebimento de {recebimento.cliente}'
        if recebimento.medicao_numero:
            descricao = f'{descricao} - Medição {recebimento.medicao_numero}'
        linhas.append(
            {
                'data': data,
                'sort_key': (data, recebimento.pk, 30),
                'natureza': MovimentoCaixa.Natureza.ENTRADA,
                'origem': 'Recebimento por medição',
                'descricao': descricao,
                'valor': recebimento.valor_liquido,
                'entrada': True,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': recebimento.categoria_id,
                'categoria_ids': {recebimento.categoria_id} if recebimento.categoria_id else set(),
                'referencia': f'REC #{recebimento.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:recebimento_editar',
                    kwargs={'tipo': 'medicao', 'pk': recebimento.pk},
                ),
            }
        )

    pagamentos_diretos = (
        PagamentoNotaFiscalPagamento.objects.filter(
            pagamento_nf__caixa=caixa,
            tipo__in=(
                PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA,
                PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO,
            ),
        )
        .select_related('pagamento_nf', 'pagamento_nf__fornecedor')
        .prefetch_related('pagamento_nf__itens')
        .order_by('data', 'pk')
    )
    for pagamento in pagamentos_diretos:
        nf = pagamento.pagamento_nf
        categoria_ids = set(
            nf.itens.exclude(categoria_id__isnull=True).values_list('categoria_id', flat=True)
        )
        linhas.append(
            {
                'data': pagamento.data,
                'sort_key': (pagamento.data, pagamento.pk, 40),
                'natureza': MovimentoCaixa.Natureza.SAIDA,
                'origem': pagamento.get_tipo_display(),
                'descricao': f'{nf.fornecedor} - NF {nf.numero_nf}',
                'valor': pagamento.total_a_pagar(),
                'entrada': False,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': nf.fornecedor_id,
                'categoria_id': None,
                'categoria_ids': categoria_ids,
                'referencia': f'NF #{nf.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_nf_detalhe',
                    kwargs={'pk': nf.pk},
                ),
            }
        )

    boletos_pagos = (
        BoletoPagamento.objects.filter(
            pagamento_nf__caixa=caixa,
            status=BoletoPagamento.Status.PAGO,
        )
        .select_related('pagamento_nf', 'pagamento_nf__fornecedor')
        .prefetch_related('pagamento_nf__itens')
        .order_by('data_pagamento', 'vencimento', 'pk')
    )
    for boleto in boletos_pagos:
        nf = boleto.pagamento_nf
        categoria_ids = set(
            nf.itens.exclude(categoria_id__isnull=True).values_list('categoria_id', flat=True)
        )
        data = boleto.data_pagamento or boleto.vencimento
        valor = boleto.valor_pago
        if valor is None:
            valor = (boleto.valor + boleto.acrescimos - boleto.descontos).quantize(
                Decimal('0.01')
            )
        linhas.append(
            {
                'data': data,
                'sort_key': (data, boleto.pk, 50),
                'natureza': MovimentoCaixa.Natureza.SAIDA,
                'origem': 'Boleto pago',
                'descricao': f'{nf.fornecedor} - NF {nf.numero_nf} - {boleto.numero_doc}',
                'valor': valor,
                'entrada': False,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': nf.fornecedor_id,
                'categoria_id': None,
                'categoria_ids': categoria_ids,
                'referencia': f'BOL #{boleto.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_nf_detalhe',
                    kwargs={'pk': nf.pk},
                ),
            }
        )

    pagamentos_pessoal = (
        PagamentoPessoal.objects.filter(caixa=caixa)
        .select_related('funcionario')
        .prefetch_related('itens', 'itens__categoria')
        .order_by('data_pagamento', 'data_emissao', 'pk')
    )
    for pagamento in pagamentos_pessoal:
        categoria_ids = {
            item.categoria_id
            for item in pagamento.itens.all()
            if item.categoria_id
        }
        descricao_item = ', '.join(
            item.descricao for item in pagamento.itens.all() if item.descricao
        )
        linhas.append(
            {
                'data': pagamento.data_pagamento or pagamento.data_emissao,
                'sort_key': (pagamento.data_pagamento or pagamento.data_emissao, pagamento.pk, 60),
                'natureza': MovimentoCaixa.Natureza.SAIDA,
                'origem': 'À vista',
                'descricao': (
                    f'{pagamento.funcionario.nome} - {descricao_item or "Pagamento pessoal"}'
                    if pagamento.funcionario_id
                    else descricao_item or 'Pagamento pessoal geral'
                ),
                'valor': pagamento.total_itens(),
                'entrada': False,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': None,
                'categoria_ids': categoria_ids,
                'referencia': f'PES #{pagamento.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_pessoal_detalhe',
                    kwargs={'pk': pagamento.pk},
                ),
            }
        )

    pagamentos_impostos = (
        PagamentoImposto.objects.filter(caixa=caixa)
        .select_related('autoridade')
        .prefetch_related('itens', 'itens__categoria')
        .order_by('data_pagamento', 'data_emissao', 'pk')
    )
    for pagamento in pagamentos_impostos:
        categoria_ids = {item.categoria_id for item in pagamento.itens.all() if item.categoria_id}
        descricao_item = ', '.join(item.descricao for item in pagamento.itens.all() if item.descricao)
        linhas.append(
            {
                'data': pagamento.data_pagamento or pagamento.data_emissao,
                'sort_key': (pagamento.data_pagamento or pagamento.data_emissao, pagamento.pk, 70),
                'natureza': MovimentoCaixa.Natureza.SAIDA,
                'origem': 'À vista',
                'descricao': f'{pagamento.autoridade.nome} - {descricao_item or "Pagamento de imposto"}',
                'valor': pagamento.total_itens(),
                'entrada': False,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': None,
                'categoria_ids': categoria_ids,
                'referencia': f'IMP #{pagamento.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_imposto_detalhe',
                    kwargs={'pk': pagamento.pk},
                ),
            }
        )

    pagamentos_bancarios = (
        PagamentoBancarioParcela.objects.filter(
            recorrencia__caixa=caixa,
            status=PagamentoBancarioParcela.Status.PAGO,
        )
        .select_related('recorrencia', 'recorrencia__categoria')
        .order_by('data_pagamento', 'data_vencimento', 'pk')
    )
    for parcela in pagamentos_bancarios:
        recorrencia = parcela.recorrencia
        data = parcela.data_pagamento or parcela.data_vencimento
        categoria_ids = {recorrencia.categoria_id} if recorrencia.categoria_id else set()
        linhas.append(
            {
                'data': data,
                'sort_key': (data, parcela.pk, 80),
                'natureza': MovimentoCaixa.Natureza.SAIDA,
                'origem': 'Bancário',
                'descricao': f'{recorrencia.descricao} - Parcela {parcela.numero_parcela}',
                'valor': parcela.valor,
                'entrada': False,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': recorrencia.categoria_id,
                'categoria_ids': categoria_ids,
                'referencia': f'BAN #{parcela.pk}',
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_bancario_detalhe',
                    kwargs={'pk': recorrencia.pk},
                ),
            }
        )

    pagamentos_avulsos = (
        PagamentoBancarioAvulso.objects.filter(caixa=caixa)
        .select_related('categoria', 'conta_bancaria')
        .order_by('data_pagamento', 'pk')
    )
    for avulso in pagamentos_avulsos:
        data = avulso.data_pagamento
        categoria_ids = {avulso.categoria_id} if avulso.categoria_id else set()
        linhas.append(
            {
                'data': data,
                'sort_key': (data, avulso.pk, 81),
                'natureza': MovimentoCaixa.Natureza.SAIDA,
                'origem': 'Bancário avulso',
                'descricao': avulso.descricao or 'Pagamento bancário avulso',
                'valor': avulso.valor,
                'entrada': False,
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'fornecedor_id': None,
                'categoria_id': avulso.categoria_id,
                'categoria_ids': categoria_ids,
                'referencia': f'BVA #{avulso.pk}',
                'url': reverse_empresa(request, 'financeiro:pagamento_bancario_lista'),
            }
        )

    if inicio and fim:
        linhas = _filtrar_linhas_periodo(linhas, inicio, fim)
    return _recalcular_saldo_linhas(linhas, 'valor')


def _extrato_caixa_detalhado(
    request,
    caixa: Caixa,
    inicio=None,
    fim=None,
    *,
    incluir_recebimentos_abertos: bool = False,
) -> list[dict]:
    linhas = []

    status_recebimentos = [RecebimentoAvulso.Status.PAGO]
    if incluir_recebimentos_abertos:
        status_recebimentos.append(RecebimentoAvulso.Status.ABERTO)

    recebimentos_avulsos = (
        RecebimentoAvulso.objects.filter(caixa=caixa, status__in=status_recebimentos)
        .select_related('cliente', 'categoria')
        .order_by('data_pagamento', 'data', 'pk')
    )
    for recebimento in recebimentos_avulsos:
        data = recebimento.data_pagamento or recebimento.data or recebimento.criado_em.date()
        linhas.append(
            {
                'data': data,
                'sort_key': (data, recebimento.pk, 20),
                'entrada': True,
                'natureza': 'Entrada',
                'categoria': 'Avulso',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': recebimento.categoria_id,
                'categoria_ids': {recebimento.categoria_id} if recebimento.categoria_id else set(),
                'nf': '-',
                'pessoa': recebimento.cliente,
                'fornecedor_id': None,
                'descricao': (
                    f'{recebimento.descricao or "Recebimento avulso"}'
                    f'{" - Em aberto" if recebimento.status == RecebimentoAvulso.Status.ABERTO else ""}'
                ),
                'valor_bruto': recebimento.valor,
                'descontos': recebimento.impostos,
                'valor_liquido': recebimento.valor_liquido,
                'url': reverse_empresa(
                    request,
                    'financeiro:recebimento_editar',
                    kwargs={'tipo': 'avulso', 'pk': recebimento.pk},
                ),
            }
        )

    status_recebimentos = [RecebimentoMedicao.Status.PAGO]
    if incluir_recebimentos_abertos:
        status_recebimentos.append(RecebimentoMedicao.Status.ABERTO)

    recebimentos_medicao = (
        RecebimentoMedicao.objects.filter(caixa=caixa, status__in=status_recebimentos)
        .select_related('cliente', 'obra', 'categoria')
        .order_by('data_pagamento', 'data', 'pk')
    )
    for recebimento in recebimentos_medicao:
        data = recebimento.data_pagamento or recebimento.data or recebimento.criado_em.date()
        descricao = f'Medição {recebimento.medicao_numero} - {recebimento.obra.nome}'
        if recebimento.status == RecebimentoMedicao.Status.ABERTO:
            descricao = f'{descricao} - Em aberto'
        linhas.append(
            {
                'data': data,
                'sort_key': (data, recebimento.pk, 30),
                'entrada': True,
                'natureza': 'Entrada',
                'categoria': 'Medição',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': recebimento.categoria_id,
                'categoria_ids': {recebimento.categoria_id} if recebimento.categoria_id else set(),
                'nf': recebimento.nota_fiscal_numero or '-',
                'pessoa': recebimento.cliente,
                'fornecedor_id': None,
                'descricao': descricao,
                'valor_bruto': recebimento.valor,
                'descontos': recebimento.impostos,
                'valor_liquido': recebimento.valor_liquido,
                'url': reverse_empresa(
                    request,
                    'financeiro:recebimento_editar',
                    kwargs={'tipo': 'medicao', 'pk': recebimento.pk},
                ),
            }
        )

    itens_nf = (
        PagamentoNotaFiscalItem.objects.filter(caixa=caixa)
        .select_related('categoria', 'pagamento_nf', 'pagamento_nf__fornecedor')
        .order_by('pagamento_nf__data_emissao', 'pagamento_nf_id', 'pk')
    )
    for item in itens_nf:
        nf = item.pagamento_nf
        descricao = (
            f'{item.descricao} - {_format_quantidade_extrato(item.quantidade)} '
            f'{item.unidade or "-"} - R$ {format_decimal_br_moeda(item.valor_unitario)}'
        )
        linhas.append(
            {
                'data': nf.data_emissao,
                'sort_key': (nf.data_emissao, item.pk, 40),
                'entrada': False,
                'natureza': 'Saída',
                'categoria': item.categoria.nome if item.categoria_id else '-',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': item.categoria_id,
                'categoria_ids': {item.categoria_id} if item.categoria_id else set(),
                'nf': nf.numero_nf,
                'pessoa': nf.fornecedor,
                'fornecedor_id': nf.fornecedor_id,
                'descricao': descricao,
                'valor_bruto': item.valor_total,
                'descontos': Decimal('0'),
                'valor_liquido': item.valor_total,
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_nf_detalhe',
                    kwargs={'pk': nf.pk},
                ),
            }
        )

    itens_pessoal = (
        PagamentoPessoalItem.objects.filter(pagamento__caixa=caixa)
        .select_related('categoria', 'pagamento', 'pagamento__funcionario')
        .order_by('pagamento__data_pagamento', 'pagamento__data_emissao', 'pagamento_id', 'pk')
    )
    for item in itens_pessoal:
        pagamento = item.pagamento
        data = pagamento.data_pagamento or pagamento.data_emissao
        linhas.append(
            {
                'data': data,
                'sort_key': (data, item.pk, 50),
                'entrada': False,
                'natureza': 'Saída',
                'categoria': item.categoria.nome if item.categoria_id else '-',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': item.categoria_id,
                'categoria_ids': {item.categoria_id} if item.categoria_id else set(),
                'nf': '-',
                'pessoa': pagamento.funcionario or 'Geral',
                'fornecedor_id': None,
                'descricao': item.descricao,
                'valor_bruto': item.valor_total,
                'descontos': Decimal('0'),
                'valor_liquido': item.valor_total,
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_pessoal_detalhe',
                    kwargs={'pk': pagamento.pk},
                ),
            }
        )

    itens_impostos = (
        PagamentoImpostoItem.objects.filter(pagamento__caixa=caixa)
        .select_related('categoria', 'pagamento', 'pagamento__autoridade')
        .order_by('pagamento__data_pagamento', 'pagamento__data_emissao', 'pagamento_id', 'pk')
    )
    for item in itens_impostos:
        pagamento = item.pagamento
        data = pagamento.data_pagamento or pagamento.data_emissao
        linhas.append(
            {
                'data': data,
                'sort_key': (data, item.pk, 60),
                'entrada': False,
                'natureza': 'Saída',
                'categoria': item.categoria.nome if item.categoria_id else '-',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': item.categoria_id,
                'categoria_ids': {item.categoria_id} if item.categoria_id else set(),
                'nf': '-',
                'pessoa': pagamento.autoridade,
                'fornecedor_id': None,
                'descricao': item.descricao,
                'valor_bruto': item.valor_total,
                'descontos': Decimal('0'),
                'valor_liquido': item.valor_total,
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_imposto_detalhe',
                    kwargs={'pk': pagamento.pk},
                ),
            }
        )

    parcelas_bancarias = (
        PagamentoBancarioParcela.objects.filter(
            recorrencia__caixa=caixa,
            status__in=(
                PagamentoBancarioParcela.Status.ABERTO,
                PagamentoBancarioParcela.Status.PAGO,
            ),
        )
        .select_related(
            'recorrencia',
            'recorrencia__categoria',
            'recorrencia__conta_bancaria',
            'conta_bancaria',
        )
        .order_by('data_vencimento', 'data_pagamento', 'pk')
    )
    for parcela in parcelas_bancarias:
        recorrencia = parcela.recorrencia
        if parcela.status == PagamentoBancarioParcela.Status.PAGO:
            data = parcela.data_pagamento or parcela.data_vencimento
            extrato_apenas_comprometido = False
            sufixo = ''
        else:
            data = parcela.data_vencimento
            extrato_apenas_comprometido = True
            sufixo = ' - Em aberto'
        conta = parcela.conta_bancaria or recorrencia.conta_bancaria
        linhas.append(
            {
                'data': data,
                'sort_key': (data, parcela.pk, 70),
                'entrada': False,
                'natureza': 'Saída',
                'categoria': recorrencia.categoria.nome if recorrencia.categoria_id else '-',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': recorrencia.categoria_id,
                'categoria_ids': {recorrencia.categoria_id} if recorrencia.categoria_id else set(),
                'nf': '-',
                'pessoa': conta,
                'fornecedor_id': None,
                'descricao': (
                    f'{recorrencia.descricao} - Parcela {parcela.numero_parcela}{sufixo}'
                ),
                'valor_bruto': parcela.valor,
                'descontos': Decimal('0'),
                'valor_liquido': parcela.valor,
                'extrato_apenas_comprometido': extrato_apenas_comprometido,
                'url': reverse_empresa(
                    request,
                    'financeiro:pagamento_bancario_detalhe',
                    kwargs={'pk': recorrencia.pk},
                ),
            }
        )

    pagamentos_avulsos = (
        PagamentoBancarioAvulso.objects.filter(caixa=caixa)
        .select_related('categoria', 'conta_bancaria')
        .order_by('data_pagamento', 'pk')
    )
    for avulso in pagamentos_avulsos:
        linhas.append(
            {
                'data': avulso.data_pagamento,
                'sort_key': (avulso.data_pagamento, avulso.pk, 71),
                'entrada': False,
                'natureza': 'Saída',
                'categoria': avulso.categoria.nome if avulso.categoria_id else '-',
                'caixa_id': caixa.pk,
                'caixa_nome': caixa.nome,
                'categoria_id': avulso.categoria_id,
                'categoria_ids': {avulso.categoria_id} if avulso.categoria_id else set(),
                'nf': '-',
                'pessoa': avulso.conta_bancaria,
                'fornecedor_id': None,
                'descricao': avulso.descricao or 'Pagamento bancário avulso',
                'valor_bruto': avulso.valor,
                'descontos': Decimal('0'),
                'valor_liquido': avulso.valor,
                'url': reverse_empresa(request, 'financeiro:pagamento_bancario_lista'),
            }
        )

    if inicio and fim:
        linhas = _filtrar_linhas_periodo(linhas, inicio, fim)
    return _recalcular_saldo_linhas(linhas, 'valor_liquido')


def _sort_value_extrato_detalhado(linha: dict, campo: str):
    if campo == 'numero':
        return linha.get('numero') or 0
    if campo == 'data':
        return linha.get('data')
    if campo == 'natureza':
        return linha.get('natureza') or ''
    if campo == 'caixa':
        return str(linha.get('caixa_nome') or '').lower()
    if campo == 'categoria':
        return str(linha.get('categoria') or '').lower()
    if campo == 'nf':
        return str(linha.get('nf') or '').lower()
    if campo == 'pessoa':
        return str(linha.get('pessoa') or '').lower()
    if campo == 'descricao':
        return str(linha.get('descricao') or '').lower()
    if campo == 'valor_bruto':
        return linha.get('valor_bruto') or Decimal('0')
    if campo == 'descontos':
        return linha.get('descontos') or Decimal('0')
    if campo == 'valor_liquido':
        return linha.get('valor_liquido') or Decimal('0')
    if campo == 'saldo':
        return linha.get('saldo_apos_lancamento') or Decimal('0')
    return linha.get('data')


def _ordenar_extrato_detalhado(linhas: list[dict], campo: str, direcao: str) -> list[dict]:
    campos_validos = {
        'numero',
        'data',
        'natureza',
        'caixa',
        'categoria',
        'nf',
        'pessoa',
        'descricao',
        'valor_bruto',
        'descontos',
        'valor_liquido',
        'saldo',
    }
    campo = campo if campo in campos_validos else 'data'
    direcao = 'asc' if direcao == 'asc' else 'desc'
    linhas = sorted(
        linhas,
        key=lambda linha: (_sort_value_extrato_detalhado(linha, campo), linha.get('numero') or 0),
        reverse=direcao == 'desc',
    )
    for idx, linha in enumerate(linhas, start=1):
        linha['numero'] = idx
    return linhas


def _sort_links_extrato_detalhado(
    sort_atual: str,
    direcao_atual: str,
    query_base: dict | None = None,
) -> dict:
    campos = (
        'numero',
        'data',
        'natureza',
        'caixa',
        'categoria',
        'nf',
        'pessoa',
        'descricao',
        'valor_bruto',
        'descontos',
        'valor_liquido',
        'saldo',
    )
    links = {}
    for campo in campos:
        proxima_direcao = 'desc'
        if sort_atual == campo and direcao_atual == 'desc':
            proxima_direcao = 'asc'
        query = dict(query_base or {})
        query.update({'sort': campo, 'dir': proxima_direcao})
        links[campo] = {
            'url': f'?{_query_string(query)}',
            'active': sort_atual == campo,
            'indicator': '↓' if direcao_atual == 'desc' else '↑',
        }
    return links


@login_required
def dashboard(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    return render(
        request,
        'financeiro/dashboard.html',
        {'page_title': 'Financeiro'},
    )


@login_required
def relatorios(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    hoje = timezone.localdate()
    dec = DecimalField(max_digits=16, decimal_places=2)

    def inicio_mes(data_ref):
        return data_ref.replace(day=1)

    def somar_meses(data_ref, quantidade):
        mes = data_ref.month - 1 + quantidade
        ano = data_ref.year + mes // 12
        mes = mes % 12 + 1
        return date(ano, mes, 1)

    mes_atual = inicio_mes(hoje)
    meses_faturamento = [somar_meses(mes_atual, deslocamento) for deslocamento in range(-11, 1)]
    inicio_periodo = meses_faturamento[0]
    fim_periodo = somar_meses(mes_atual, 1)

    def faturamento_por_mes(modelo, status_pago):
        return (
            modelo.objects.filter(
                empresa=empresa,
                status=status_pago,
                data_pagamento__isnull=False,
                data_pagamento__gte=inicio_periodo,
                data_pagamento__lt=fim_periodo,
            )
            .annotate(mes=TruncMonth('data_pagamento'))
            .values('mes', 'caixa_id')
            .annotate(total=Coalesce(Sum('valor_liquido'), Value(Decimal('0')), output_field=dec))
            .order_by('mes')
        )

    caixas = list(Caixa.objects.filter(empresa=empresa).order_by('tipo', 'nome'))
    faturamento_mensal = {mes: Decimal('0') for mes in meses_faturamento}
    faturamento_por_caixa = {
        str(caixa.pk): {mes: Decimal('0') for mes in meses_faturamento}
        for caixa in caixas
    }

    def aplicar_faturamento(item):
        mes = item['mes'].date() if hasattr(item['mes'], 'date') else item['mes']
        caixa_id = str(item['caixa_id']) if item['caixa_id'] else ''
        faturamento_mensal[mes] += item['total']
        if caixa_id in faturamento_por_caixa:
            faturamento_por_caixa[caixa_id][mes] += item['total']

    for item in faturamento_por_mes(RecebimentoAvulso, RecebimentoAvulso.Status.PAGO):
        aplicar_faturamento(item)
    for item in faturamento_por_mes(RecebimentoMedicao, RecebimentoMedicao.Status.PAGO):
        aplicar_faturamento(item)

    nomes_meses = {
        1: 'Jan',
        2: 'Fev',
        3: 'Mar',
        4: 'Abr',
        5: 'Mai',
        6: 'Jun',
        7: 'Jul',
        8: 'Ago',
        9: 'Set',
        10: 'Out',
        11: 'Nov',
        12: 'Dez',
    }
    faturamento_chart = {
        'labels': [f'{nomes_meses[mes.month]}/{str(mes.year)[-2:]}' for mes in meses_faturamento],
        'series': {
            'todos': {
                'label': 'Todos',
                'values': [float(faturamento_mensal[mes]) for mes in meses_faturamento],
                'total': float(sum(faturamento_mensal.values(), Decimal('0'))),
            },
            **{
                str(caixa.pk): {
                    'label': caixa.nome,
                    'values': [float(faturamento_por_caixa[str(caixa.pk)][mes]) for mes in meses_faturamento],
                    'total': float(
                        sum(faturamento_por_caixa[str(caixa.pk)].values(), Decimal('0'))
                    ),
                }
                for caixa in caixas
            },
        },
        'caixas': [
            {
                'id': str(caixa.pk),
                'nome': caixa.nome,
            }
            for caixa in caixas
        ],
    }
    faturamento_total_periodo = sum(
        faturamento_mensal.values(),
        Decimal('0'),
    )

    return render(
        request,
        'financeiro/relatorios.html',
        {
            'page_title': 'Relatórios Financeiros',
            'faturamento_chart': faturamento_chart,
            'faturamento_total_periodo': faturamento_total_periodo,
        },
    )


def _buscar_pagamentos_context(request, empresa) -> dict:
    query = str(request.GET.get('q', '')).strip()
    fornecedor = str(request.GET.get('fornecedor', '')).strip()
    categoria_id = _int_param(request, 'categoria')
    valor_raw = str(request.GET.get('valor', '')).strip()
    data_inicio = str(request.GET.get('data_inicio', '')).strip()
    data_fim = str(request.GET.get('data_fim', '')).strip()
    status_vencido_raw = str(request.GET.get('status_vencido', '')).strip()
    status_aberto_raw = str(request.GET.get('status_aberto', '')).strip()
    status_pago_raw = str(request.GET.get('status_pago', '')).strip()
    origem = str(request.GET.get('origem', '')).strip()
    tem_filtro_principal = any(
        [
            query,
            fornecedor,
            valor_raw,
            data_inicio,
            data_fim,
            categoria_id,
        ]
    )
    tem_status_expresso = (
        'status_vencido' in request.GET
        or 'status_aberto' in request.GET
        or 'status_pago' in request.GET
    )
    status_vencido = (
        status_vencido_raw in ('1', 'on', 'true')
        if tem_status_expresso
        else False
    )
    status_aberto = (
        status_aberto_raw in ('1', 'on', 'true')
        if tem_status_expresso
        else False
    )
    status_pago = (
        status_pago_raw in ('1', 'on', 'true')
        if tem_status_expresso
        else False
    )
    if origem == 'dashboard' and not tem_filtro_principal:
        status_vencido = False
        status_aberto = False
        status_pago = False
    resultados = []
    erro_valor = ''
    filtros_ativos = tem_filtro_principal or any(
        [
            status_vencido,
            status_aberto,
            status_pago,
        ]
    )

    if filtros_ativos:
        hoje = timezone.localdate()
        valor = None
        if valor_raw:
            try:
                valor = parse_valor_moeda_br(valor_raw)
            except Exception:
                erro_valor = 'Informe um valor válido no padrão 0,00.'
        data_inicio_dt = parse_date(data_inicio) if data_inicio else None
        data_fim_dt = parse_date(data_fim) if data_fim else None

        def data_no_periodo(data_ref) -> bool:
            if not data_ref:
                return False
            if data_inicio_dt and data_ref < data_inicio_dt:
                return False
            if data_fim_dt and data_ref > data_fim_dt:
                return False
            return True

        def boleto_corresponde_status(boleto) -> bool:
            if not (status_vencido or status_aberto or status_pago):
                return True
            boleto_pago = (
                boleto.status == BoletoPagamento.Status.PAGO
                or (boleto.valor_pago or Decimal('0')) > 0
            )
            boleto_aberto = boleto.status not in (
                BoletoPagamento.Status.PAGO,
                BoletoPagamento.Status.CANCELADO,
            )
            return (
                (status_pago and boleto_pago)
                or (status_vencido and boleto_aberto and boleto.vencimento < hoje)
                or (status_aberto and boleto_aberto and boleto.vencimento >= hoje)
            )

        nfs_qs = (
            PagamentoNotaFiscal.objects.filter(empresa=empresa)
            .select_related('fornecedor', 'caixa')
            .prefetch_related('boletos', 'pagamentos', 'itens__categoria')
            .annotate(
                total_itens_calc=Coalesce(
                    Subquery(
                        _subquery_total_itens_nf(),
                        output_field=DecimalField(max_digits=16, decimal_places=2),
                    ),
                    Value(Decimal('0')),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                )
            )
            .order_by('-criado_em', '-pk')
        )
        if fornecedor:
            nfs_qs = nfs_qs.filter(fornecedor__nome__icontains=fornecedor)
        if categoria_id:
            nfs_qs = nfs_qs.filter(itens__categoria_id=categoria_id)
        if query:
            nfs_qs = nfs_qs.filter(
                Q(numero_nf__icontains=query) | Q(boletos__numero_doc__icontains=query)
            )
        if valor_raw and not erro_valor:
            if valor is None:
                resultados = []
                nfs_qs = PagamentoNotaFiscal.objects.none()
            else:
                nfs_qs = nfs_qs.filter(
                    Q(total_itens_calc=valor) | Q(boletos__valor=valor)
                )

        for nf in nfs_qs.distinct():
            origens = set()
            pagamentos = list(nf.pagamentos.all().order_by('data', 'pk'))
            boletos_all = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
            boletos_relacionados = list(boletos_all)
            boletos_vencidos = [
                boleto for boleto in boletos_all
                if (
                    boleto.status
                    in (BoletoPagamento.Status.RASCUNHO, BoletoPagamento.Status.EMITIDO)
                    and boleto.vencimento < hoje
                )
            ]
            resumo = _resumo_pagamento_nf(
                pagamentos,
                boletos_all,
                nf.total_itens_calc,
                hoje,
            )
            status_labels = [resumo['situacao_label']]
            if status_vencido or status_aberto or status_pago:
                corresponde_status = _busca_pagamentos_nf_corresponde_status(
                    resumo,
                    status_vencido=status_vencido,
                    status_aberto=status_aberto,
                    status_pago=status_pago,
                )
                if not corresponde_status:
                    continue
            if query:
                if query.lower() in (nf.numero_nf or '').lower():
                    origens.add('nota_fiscal')
                boletos_relacionados = [
                    boleto for boleto in boletos_relacionados
                    if query.lower() in (boleto.numero_doc or '').lower()
                ]
                if boletos_relacionados:
                    origens.add('numero_doc')
            elif status_vencido and not status_aberto and not status_pago:
                boletos_relacionados = boletos_vencidos
            if valor_raw and not erro_valor and valor is not None:
                if nf.total_itens_calc == valor:
                    origens.add('valor_nf')
                boletos_por_valor = [
                    boleto for boleto in nf.boletos.all().order_by('vencimento', 'parcela', 'pk')
                    if boleto.valor == valor
                ]
                if boletos_por_valor:
                    origens.add('valor_doc')
                    if not query:
                        boletos_relacionados = boletos_por_valor

            url_nf = reverse_empresa(
                request,
                'financeiro:pagamento_nf_detalhe',
                kwargs={'pk': nf.pk},
            )
            categoria_label = ', '.join(
                sorted(
                    {
                        item.categoria.nome
                        for item in nf.itens.all()
                        if item.categoria_id
                    }
                )
            ) or '-'
            status_linha = _status_busca_pagamento_nf(
                resumo,
                boletos_all,
                nf.total_itens_calc,
                hoje,
            )
            boletos_validos = [
                boleto for boleto in boletos_all
                if boleto.status != BoletoPagamento.Status.CANCELADO
            ]
            total_parcelas = len(boletos_validos)
            pagamentos_diretos = [
                pagamento for pagamento in pagamentos
                if pagamento.tipo
                in (
                    PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA,
                    PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO,
                )
            ]
            for pagamento in pagamentos_diretos:
                data_pagamento_linha = pagamento.data or nf.data_emissao
                if (status_vencido or status_aberto or status_pago) and not status_pago:
                    continue
                if (data_inicio_dt or data_fim_dt) and not data_no_periodo(data_pagamento_linha):
                    continue
                resultados.append(
                    {
                        'kind': 'nf',
                        'tipo_registro': 'Nota Fiscal',
                        'forma_pagamento': pagamento.get_tipo_display(),
                        'forma_total_key': pagamento.tipo,
                        'forma_detalhe': '',
                        'url': url_nf,
                        'entidade': nf.fornecedor.nome,
                        'documento': nf.numero_nf,
                        'data': nf.data_emissao,
                        'categoria_label': categoria_label,
                        'nf': nf,
                        'origens': origens,
                        'numero_doc': nf.numero_nf,
                        'parcela_label': '',
                        'valor_linha': pagamento.total_a_pagar(),
                        'valor_total': pagamento.total_a_pagar(),
                        'vencimento': data_pagamento_linha,
                        'data_pagamento': data_pagamento_linha,
                        'status_label': status_linha,
                        'status_badge_class': _badge_status_busca_pagamento(status_linha),
                    }
                )
            for boleto in boletos_relacionados:
                if not boleto_corresponde_status(boleto):
                    continue
                data_filtro_boleto = boleto.data_pagamento if status_pago else boleto.vencimento
                if not data_filtro_boleto:
                    data_filtro_boleto = boleto.vencimento
                if (data_inicio_dt or data_fim_dt) and not data_no_periodo(data_filtro_boleto):
                    continue
                resultados.append(
                    {
                        'kind': 'nf',
                        'tipo_registro': 'Nota Fiscal',
                        'forma_pagamento': 'Boletos',
                        'forma_total_key': 'boletos',
                        'forma_detalhe': f'Qtd de parcelas: {total_parcelas}',
                        'url': url_nf,
                        'entidade': nf.fornecedor.nome,
                        'documento': nf.numero_nf,
                        'data': nf.data_emissao,
                        'categoria_label': categoria_label,
                        'nf': nf,
                        'origens': origens,
                        'numero_doc': boleto.numero_doc or nf.numero_nf,
                        'parcela_label': f'{boleto.parcela}/{total_parcelas or boleto.parcela}',
                        'valor_linha': boleto.valor,
                        'valor_total': boleto.valor,
                        'vencimento': boleto.vencimento,
                        'data_pagamento': boleto.data_pagamento,
                        'status_label': status_linha,
                        'status_badge_class': _badge_status_busca_pagamento(status_linha),
                    }
                )
            if not pagamentos_diretos and not boletos_relacionados:
                if (data_inicio_dt or data_fim_dt) and not data_no_periodo(nf.data_emissao):
                    continue
                resultados.append(
                    {
                        'kind': 'nf',
                        'tipo_registro': 'Nota Fiscal',
                        'forma_pagamento': 'Sem pagamento',
                        'forma_total_key': '',
                        'forma_detalhe': '',
                        'url': url_nf,
                        'entidade': nf.fornecedor.nome,
                        'documento': nf.numero_nf,
                        'data': nf.data_emissao,
                        'categoria_label': categoria_label,
                        'nf': nf,
                        'origens': origens,
                        'numero_doc': nf.numero_nf,
                        'parcela_label': '',
                        'valor_linha': nf.total_itens_calc,
                        'valor_total': nf.total_itens_calc,
                        'vencimento': nf.data_emissao,
                        'data_pagamento': None,
                        'status_label': status_linha,
                        'status_badge_class': _badge_status_busca_pagamento(status_linha),
                    }
                )

        pessoal_qs = (
            PagamentoPessoal.objects.filter(empresa=empresa)
            .select_related('funcionario', 'caixa')
            .prefetch_related('itens__categoria')
            .order_by('-criado_em', '-pk')
        )
        if fornecedor:
            pessoal_qs = pessoal_qs.filter(
                Q(funcionario__nome__icontains=fornecedor)
                | Q(descricao__icontains=fornecedor)
                | (
                    Q(tipo_destino=PagamentoPessoal.TipoDestino.GERAL)
                    if 'geral'.startswith(fornecedor.lower()) or fornecedor.lower() in 'geral'
                    else Q(pk__isnull=True)
                )
            )
        if categoria_id:
            pessoal_qs = pessoal_qs.filter(itens__categoria_id=categoria_id)
        if data_inicio:
            pessoal_qs = pessoal_qs.filter(data_pagamento__gte=data_inicio)
        if data_fim:
            pessoal_qs = pessoal_qs.filter(data_pagamento__lte=data_fim)
        if query:
            pessoal_qs = pessoal_qs.filter(
                Q(funcionario__nome__icontains=query)
                | Q(descricao__icontains=query)
                | Q(itens__descricao__icontains=query)
                | Q(itens__categoria__nome__icontains=query)
            )
        if valor_raw and not erro_valor:
            if valor is None:
                pessoal_qs = PagamentoPessoal.objects.none()

        for pagamento in pessoal_qs.distinct():
            total = pagamento.total_itens()
            if valor_raw and not erro_valor and valor is not None and total != valor:
                continue
            if not status_pago:
                continue
            categorias = sorted(
                {item.categoria.nome for item in pagamento.itens.all() if item.categoria_id}
            )
            entidade = (
                pagamento.funcionario.nome
                if pagamento.funcionario_id
                else 'Geral (Pessoal)'
            )
            resultados.append(
                {
                    'kind': 'pessoal',
                    'url': reverse_empresa(
                        request,
                        'financeiro:pagamento_pessoal_detalhe',
                        kwargs={'pk': pagamento.pk},
                    ),
                    'entidade': entidade,
                    'documento': pagamento.descricao or 'Pagamento pessoal',
                    'data': pagamento.data_pagamento,
                    'categoria_label': ', '.join(categorias) or '-',
                    'tipo_registro': 'Pessoal',
                    'forma_pagamento': 'À vista',
                    'forma_total_key': '',
                    'forma_detalhe': '',
                    'numero_doc': pagamento.descricao or 'Pagamento pessoal',
                    'parcela_label': '',
                    'valor_linha': total,
                    'valor_total': total,
                    'vencimento': pagamento.data_pagamento,
                    'data_pagamento': pagamento.data_pagamento,
                    'status_label': 'Pago Completo',
                    'status_badge_class': _badge_status_busca_pagamento('Pago Completo'),
                }
            )

        impostos_qs = (
            PagamentoImposto.objects.filter(empresa=empresa)
            .select_related('autoridade', 'caixa')
            .prefetch_related('itens__categoria')
            .order_by('-criado_em', '-pk')
        )
        if fornecedor:
            impostos_qs = impostos_qs.filter(autoridade__nome__icontains=fornecedor)
        if categoria_id:
            impostos_qs = impostos_qs.filter(itens__categoria_id=categoria_id)
        if data_inicio:
            impostos_qs = impostos_qs.filter(data_pagamento__gte=data_inicio)
        if data_fim:
            impostos_qs = impostos_qs.filter(data_pagamento__lte=data_fim)
        if query:
            impostos_qs = impostos_qs.filter(
                Q(autoridade__nome__icontains=query)
                | Q(itens__descricao__icontains=query)
                | Q(itens__categoria__nome__icontains=query)
            )
        if valor_raw and not erro_valor:
            if valor is None:
                impostos_qs = PagamentoImposto.objects.none()

        for pagamento in impostos_qs.distinct():
            total = pagamento.total_itens()
            if valor_raw and not erro_valor and valor is not None and total != valor:
                continue
            if not status_pago:
                continue
            categorias = sorted(
                {item.categoria.nome for item in pagamento.itens.all() if item.categoria_id}
            )
            resultados.append(
                {
                    'kind': 'imposto',
                    'url': reverse_empresa(
                        request,
                        'financeiro:pagamento_imposto_detalhe',
                        kwargs={'pk': pagamento.pk},
                    ),
                    'entidade': pagamento.autoridade.nome,
                    'documento': 'Pagamento de imposto',
                    'data': pagamento.data_pagamento,
                    'categoria_label': ', '.join(categorias) or '-',
                    'tipo_registro': 'Imposto',
                    'forma_pagamento': 'À vista',
                    'forma_total_key': '',
                    'forma_detalhe': '',
                    'numero_doc': 'Pagamento de imposto',
                    'parcela_label': '',
                    'valor_linha': total,
                    'valor_total': total,
                    'vencimento': pagamento.data_pagamento,
                    'data_pagamento': pagamento.data_pagamento,
                    'status_label': 'Pago Completo',
                    'status_badge_class': _badge_status_busca_pagamento('Pago Completo'),
                }
            )

        resultados.sort(
            key=lambda item: (
                item.get('vencimento') or item.get('data') or hoje,
                item.get('documento') or '',
            ),
            reverse=True,
        )

    totais_busca = {
        'quantidade': len(resultados),
        'valor_total': sum((item.get('valor_total') or Decimal('0') for item in resultados), Decimal('0')),
        'boletos_qtd': sum(1 for item in resultados if item.get('forma_total_key') == 'boletos'),
        'boletos_total': sum((item.get('valor_total') or Decimal('0') for item in resultados if item.get('forma_total_key') == 'boletos'), Decimal('0')),
        'avista_qtd': sum(1 for item in resultados if item.get('forma_total_key') == PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA),
        'avista_total': sum((item.get('valor_total') or Decimal('0') for item in resultados if item.get('forma_total_key') == PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA), Decimal('0')),
        'credito_qtd': sum(1 for item in resultados if item.get('forma_total_key') == PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO),
        'credito_total': sum((item.get('valor_total') or Decimal('0') for item in resultados if item.get('forma_total_key') == PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO), Decimal('0')),
        'valor_nf': sum((item.get('valor_total') or Decimal('0') for item in resultados if item.get('tipo_registro') == 'Nota Fiscal'), Decimal('0')),
        'valor_pessoal': sum((item.get('valor_total') or Decimal('0') for item in resultados if item.get('kind') == 'pessoal'), Decimal('0')),
        'valor_impostos': sum((item.get('valor_total') or Decimal('0') for item in resultados if item.get('kind') == 'imposto'), Decimal('0')),
    }

    return {
        'page_title': 'Buscar Pagamentos',
        'query': query,
        'fornecedor': fornecedor,
        'categoria_id': categoria_id,
        'valor': valor_raw,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'status_vencido': status_vencido,
        'status_aberto': status_aberto,
        'status_pago': status_pago,
        'erro_valor': erro_valor,
        'filtros_ativos': filtros_ativos,
        'resultados': resultados,
        'totais_busca': totais_busca,
        'categorias_filtro': CategoriaFinanceira.objects.filter(
            empresa=empresa,
            ativo=True,
            tipo=CategoriaFinanceira.Tipo.SAIDA,
        ).order_by('movimentacao_tipo', 'nome'),
    }


@login_required
def buscar_pagamentos(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    return render(
        request,
        'financeiro/busca_pagamentos.html',
        _buscar_pagamentos_context(request, empresa),
    )


def _linhas_export_busca_pagamentos(context: dict) -> list[list]:
    linhas = []
    for item in context['resultados']:
        linhas.append(
            [
                item.get('status_label') or '',
                item.get('data'),
                item.get('entidade') or '',
                item.get('tipo_registro') or '',
                item.get('categoria_label') or '',
                item.get('numero_doc') or '',
                item.get('forma_pagamento') or '',
                item.get('forma_detalhe') or '',
                item.get('vencimento'),
                item.get('valor_linha') or Decimal('0'),
                item.get('parcela_label') or '',
                item.get('data_pagamento'),
            ]
        )
    return linhas


def _filtros_export_busca_pagamentos(context: dict) -> str:
    filtros = []
    if context.get('query'):
        filtros.append(f'Doc/NF: {context["query"]}')
    if context.get('fornecedor'):
        filtros.append(f'Pessoa/Entidade: {context["fornecedor"]}')
    if context.get('valor'):
        filtros.append(f'Valor: R$ {context["valor"]}')
    if context.get('data_inicio') or context.get('data_fim'):
        inicio = context.get('data_inicio') or '...'
        fim = context.get('data_fim') or '...'
        filtros.append(f'Período: {inicio} a {fim}')
    status = []
    if context.get('status_vencido'):
        status.append('Vencidos')
    if context.get('status_aberto'):
        status.append('Em aberto')
    if context.get('status_pago'):
        status.append('Pagos')
    if status:
        filtros.append(f'Status: {", ".join(status)}')
    return ' | '.join(filtros) or 'Sem filtros'


@login_required
def buscar_pagamentos_xlsx(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    context = _buscar_pagamentos_context(request, empresa)
    headers = [
        'Status',
        'Data Emissão',
        'Fornecedor / Funcionário / Entidade',
        'Tipo de Saída',
        'Categoria',
        'Nº Doc',
        'Tipo Pagamento',
        'Detalhe Tipo',
        'Vencimento',
        'Valor',
        'Parcela',
        'Data de Pagamento',
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Buscar Pagamentos'

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    title = ws.cell(row=row, column=1, value=f'Buscar Pagamentos - {_nome_empresa_pdf(empresa)}')
    title.font = Font(bold=True, size=13)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    ws.cell(row=row, column=1, value=_filtros_export_busca_pagamentos(context))
    row += 2

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row += 1

    for valores in _linhas_export_busca_pagamentos(context):
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col == 10:
                cell.number_format = 'R$ #,##0.00'
        row += 1

    totais = context['totais_busca']
    row += 1
    ws.cell(row=row, column=1, value='Total geral').font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(totais['valor_total'])).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'{totais["quantidade"]} resultado(s)')
    row += 1
    for label, qtd_key, total_key in (
        ('Boletos', 'boletos_qtd', 'boletos_total'),
        ('À Vista', 'avista_qtd', 'avista_total'),
        ('Crédito', 'credito_qtd', 'credito_total'),
    ):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(totais[total_key]))
        ws.cell(row=row, column=3, value=f'{totais[qtd_key]} resultado(s)')
        row += 1

    widths = [16, 13, 34, 16, 22, 16, 18, 22, 13, 14, 12, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    ref = timezone.localdate().strftime('%Y%m%d')
    filename = f'Buscar_Pagamentos_{ref}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def buscar_pagamentos_pdf(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    context = _buscar_pagamentos_context(request, empresa)
    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=9 * mm,
        bottomMargin=9 * mm,
    )
    story = []
    story.extend(_flowables_titulo_pdf_centro('Buscar Pagamentos', styles))
    meta_style = ParagraphStyle(
        'bp_meta_pdf',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#475569'),
        alignment=1,
    )
    story.append(Paragraph(xml_escape(_nome_empresa_pdf(empresa)), meta_style))
    story.append(Paragraph(xml_escape(_filtros_export_busca_pagamentos(context)), meta_style))
    story.append(Spacer(1, 3 * mm))

    cell = ParagraphStyle(
        'bp_pdf_cell',
        parent=styles['Normal'],
        fontSize=5.7,
        leading=7,
        textColor=colors.HexColor('#0f172a'),
    )
    cell_right = ParagraphStyle('bp_pdf_cell_right', parent=cell, alignment=2)
    data = [[
        'Status',
        'Emissão',
        'Entidade',
        'Tipo Saída',
        'Categoria',
        'Nº Doc',
        'Tipo Pgto',
        'Venc.',
        'Valor',
        'Pgto',
    ]]
    for item in context['resultados']:
        tipo_pgto = item.get('forma_pagamento') or ''
        if item.get('forma_detalhe'):
            tipo_pgto = f'{tipo_pgto}\n{item["forma_detalhe"]}'
        if item.get('parcela_label'):
            tipo_pgto = f'{tipo_pgto}\nParcela {item["parcela_label"]}'
        data.append(
            [
                _pdf_cell(item.get('status_label') or '', cell),
                _pdf_cell(_format_data_pdf(item.get('data')), cell),
                _pdf_cell(item.get('entidade') or '', cell),
                _pdf_cell(item.get('tipo_registro') or '', cell),
                _pdf_cell(item.get('categoria_label') or '', cell),
                _pdf_cell(item.get('numero_doc') or '', cell),
                _pdf_cell(tipo_pgto, cell),
                _pdf_cell(_format_data_pdf(item.get('vencimento')), cell),
                _pdf_cell(_format_moeda_pdf(item.get('valor_linha')), cell_right),
                _pdf_cell(_format_data_pdf(item.get('data_pagamento')), cell),
            ]
        )
    table = Table(
        data,
        repeatRows=1,
        colWidths=[18 * mm, 16 * mm, 48 * mm, 20 * mm, 27 * mm, 20 * mm, 27 * mm, 16 * mm, 21 * mm, 18 * mm],
    )
    table.setStyle(_pdf_table_style())
    story.append(table)

    totais = context['totais_busca']
    story.append(Spacer(1, 3 * mm))
    totais_table = Table(
        [[
            'Boletos',
            'À Vista',
            'Crédito',
            'Total Geral',
        ], [
            f'{totais["boletos_qtd"]} | {_format_moeda_pdf(totais["boletos_total"])}',
            f'{totais["avista_qtd"]} | {_format_moeda_pdf(totais["avista_total"])}',
            f'{totais["credito_qtd"]} | {_format_moeda_pdf(totais["credito_total"])}',
            f'{totais["quantidade"]} | {_format_moeda_pdf(totais["valor_total"])}',
        ]],
        colWidths=[43 * mm, 43 * mm, 43 * mm, 50 * mm],
    )
    totais_table.setStyle(_pdf_table_style())
    story.append(totais_table)
    story.extend(_flowables_rodape_impressao(request, styles))
    doc.build(story)
    ref = timezone.localdate().strftime('%Y%m%d')
    filename = f'Buscar_Pagamentos_{ref}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _dashboard_alertas_financeiro_data(empresa):
    hoje = timezone.localdate()
    boleto_status_aberto = (
        BoletoPagamento.Status.RASCUNHO,
        BoletoPagamento.Status.EMITIDO,
    )
    boletos_venc_hoje = list(
        BoletoPagamento.objects.filter(
            pagamento_nf__empresa=empresa,
            vencimento=hoje,
            status__in=boleto_status_aberto,
        )
        .select_related('pagamento_nf', 'pagamento_nf__fornecedor')
        .prefetch_related('pagamento_nf__boletos')
        .order_by('vencimento', 'parcela', 'pk')[:50]
    )
    boletos_vencidos = list(
        BoletoPagamento.objects.filter(
            pagamento_nf__empresa=empresa,
            vencimento__lt=hoje,
            status__in=boleto_status_aberto,
        )
        .select_related('pagamento_nf', 'pagamento_nf__fornecedor')
        .prefetch_related('pagamento_nf__boletos', 'pagamento_nf__pagamentos', 'pagamento_nf__itens')
        .order_by('vencimento', 'parcela', 'pk')[:100]
    )
    notas_sem_pagamento = list(
        PagamentoNotaFiscal.objects.filter(
            empresa=empresa,
            data_emissao__lte=hoje,
            pagamentos__isnull=True,
        )
        .select_related('fornecedor', 'caixa')
        .annotate(
            total_itens_calc=Coalesce(
                Subquery(
                    _subquery_total_itens_nf(),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=16, decimal_places=2),
            ),
        )
        .order_by('-data_emissao', '-pk')[:50]
    )
    notas_com_pagamento = (
        PagamentoNotaFiscal.objects.filter(
            empresa=empresa,
            pagamentos__isnull=False,
        )
        .exclude(pagamentos__tipo=PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS)
        .select_related('fornecedor', 'caixa')
        .prefetch_related('pagamentos', 'boletos')
        .annotate(
            total_itens_calc=Coalesce(
                Subquery(
                    _subquery_total_itens_nf(),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=16, decimal_places=2),
            )
        )
        .order_by('-data_emissao', '-pk')
        .distinct()
    )
    notas_pagamento_parcial = []
    for nf in notas_com_pagamento:
        boletos_nf = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
        pagamentos_nf = list(nf.pagamentos.all().order_by('data', 'pk'))
        resumo_nf = _resumo_pagamento_nf(
            pagamentos_nf,
            boletos_nf,
            nf.total_itens_calc,
            hoje,
        )
        if resumo_nf['situacao_label'] != 'Pago Parcial':
            continue
        if resumo_nf['valor_em_aberto'] <= 0:
            continue
        nf.valor_em_aberto_dashboard = resumo_nf['valor_em_aberto']
        nf.pagamento_label_dashboard = resumo_nf['pagamento_label']
        notas_pagamento_parcial.append(nf)
        if len(notas_pagamento_parcial) >= 50:
            break

    def nf_total_dashboard(nf: PagamentoNotaFiscal) -> Decimal:
        tc = getattr(nf, 'total_itens_calc', None)
        if tc is not None:
            return tc
        items = nf.itens.all()
        return sum((i.valor_total for i in items), Decimal('0')).quantize(Decimal('0.01'))

    dashboard_boletos_vencidos_linhas = []
    for b in boletos_vencidos:
        nf = b.pagamento_nf
        boletos_nf = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
        pags_nf = list(nf.pagamentos.all().order_by('data', 'pk'))
        total_nf = nf_total_dashboard(nf)
        resumo = _resumo_pagamento_nf(pags_nf, boletos_nf, total_nf, hoje)
        fornecedor = nf.fornecedor
        dashboard_boletos_vencidos_linhas.append(
            {
                'index': len(dashboard_boletos_vencidos_linhas) + 1,
                'vencimento': b.vencimento,
                'dias_atrasados': (hoje - b.vencimento).days,
                'numero_doc': b.numero_doc or '—',
                'parcela_label': _parcela_dashboard_boleto(b),
                'valor': b.valor,
                'fornecedor_nome': fornecedor.nome,
                'fornecedor_doc': fornecedor.cpf_cnpj_formatado,
                'numero_nf': nf.numero_nf,
                'valor_total': total_nf,
                'valor_pago': resumo['valor_pago'],
                'pagamento_nf_pk': nf.pk,
            }
        )

    dashboard_pagamentos_aberto_linhas = []
    idx_linha = 0
    for nf in notas_sem_pagamento:
        idx_linha += 1
        vt = nf.total_itens_calc
        dashboard_pagamentos_aberto_linhas.append(
            {
                'index': idx_linha,
                'nf_pk': nf.pk,
                'situacao': 'Sem pagamento',
                'badge_class': 'text-bg-secondary',
                'ref_principal': nf.fornecedor.nome,
                'data_emissao': nf.data_emissao,
                'valor_total': vt,
                'valor_pago': Decimal('0'),
                'ultimo_pagamento': None,
                'valor_a_pagar': vt,
            }
        )

    for nf in notas_pagamento_parcial:
        idx_linha += 1
        boletos_nf = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
        pags_nf = list(nf.pagamentos.all().order_by('data', 'pk'))
        total_nf = nf.total_itens_calc
        resumo = _resumo_pagamento_nf(pags_nf, boletos_nf, total_nf, hoje)
        ultimo = _data_ultimo_pagamento_nf(pags_nf, boletos_nf)
        dashboard_pagamentos_aberto_linhas.append(
            {
                'index': idx_linha,
                'nf_pk': nf.pk,
                'situacao': 'Pago parcial',
                'badge_class': 'text-bg-primary',
                'ref_principal': nf.fornecedor.nome,
                'data_emissao': nf.data_emissao,
                'valor_total': total_nf,
                'valor_pago': resumo['valor_pago'],
                'ultimo_pagamento': ultimo,
                'valor_a_pagar': nf.valor_em_aberto_dashboard,
            }
        )

    dashboard_boletos_venc_hoje_linhas = []
    for ii, b in enumerate(boletos_venc_hoje, start=1):
        nf = b.pagamento_nf
        fornecedor = nf.fornecedor
        dashboard_boletos_venc_hoje_linhas.append(
            {
                'index': ii,
                'numero_doc': b.numero_doc or '—',
                'parcela_label': _parcela_dashboard_boleto(b),
                'valor': b.valor,
                'fornecedor_nome': fornecedor.nome,
                'fornecedor_doc': fornecedor.cpf_cnpj_formatado,
                'numero_nf': nf.numero_nf,
                'pagamento_nf_pk': nf.pk,
            }
        )

    total_boletos_venc_hoje = sum((b.valor for b in boletos_venc_hoje), Decimal('0'))
    total_boletos_vencidos = sum((b.valor for b in boletos_vencidos), Decimal('0'))
    total_notas_sem_pagamento = sum((nf.total_itens_calc for nf in notas_sem_pagamento), Decimal('0'))
    total_notas_pagamento_parcial = sum(
        (nf.valor_em_aberto_dashboard for nf in notas_pagamento_parcial),
        Decimal('0'),
    )
    total_pagamentos_em_aberto = total_notas_sem_pagamento + total_notas_pagamento_parcial
    return {
        'hoje': hoje,
        'dashboard_boletos_venc_hoje_linhas': dashboard_boletos_venc_hoje_linhas,
        'total_boletos_venc_hoje': total_boletos_venc_hoje,
        'dashboard_boletos_vencidos_linhas': dashboard_boletos_vencidos_linhas,
        'total_boletos_vencidos': total_boletos_vencidos,
        'dashboard_pagamentos_aberto_linhas': dashboard_pagamentos_aberto_linhas,
        'total_pagamentos_em_aberto': total_pagamentos_em_aberto,
    }


@login_required
def dashboard_card_pdf(request, tipo: str):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    data = _dashboard_alertas_financeiro_data(empresa)
    configs = {
        'boletos-vencendo-hoje': {
            'titulo': 'Boletos vencendo hoje',
            'total': data['total_boletos_venc_hoje'],
            'linhas': data['dashboard_boletos_venc_hoje_linhas'],
            'headers': ['#', 'Nº Doc', 'Parcela', 'Valor', 'Fornecedor', 'CPF/CNPJ', 'Nº NF'],
            'row': lambda r: [
                r['index'],
                r['numero_doc'],
                r['parcela_label'],
                _format_moeda_pdf(r['valor']),
                r['fornecedor_nome'],
                r['fornecedor_doc'],
                r['numero_nf'],
            ],
            'widths': [9, 28, 22, 25, 55, 32, 25],
        },
        'boletos-vencidos': {
            'titulo': 'Boletos vencidos',
            'total': data['total_boletos_vencidos'],
            'linhas': data['dashboard_boletos_vencidos_linhas'],
            'headers': ['#', 'Vencimento', 'Dias', 'Nº Doc', 'Parcela', 'Valor', 'Fornecedor', 'Nº NF', 'Valor total', 'Valor pago'],
            'row': lambda r: [
                r['index'],
                _format_data_pdf(r['vencimento']),
                r['dias_atrasados'],
                r['numero_doc'],
                r['parcela_label'],
                _format_moeda_pdf(r['valor']),
                r['fornecedor_nome'],
                r['numero_nf'],
                _format_moeda_pdf(r['valor_total']),
                _format_moeda_pdf(r['valor_pago']),
            ],
            'widths': [8, 23, 13, 25, 19, 23, 44, 20, 25, 25],
        },
        'pagamentos-em-aberto': {
            'titulo': 'Pagamentos em aberto',
            'total': data['total_pagamentos_em_aberto'],
            'linhas': data['dashboard_pagamentos_aberto_linhas'],
            'headers': ['#', 'Situação', 'Referência', 'Emissão', 'Valor total', 'Valor pago', 'Último pagamento', 'Valor à pagar'],
            'row': lambda r: [
                r['index'],
                r['situacao'],
                r['ref_principal'],
                _format_data_pdf(r['data_emissao']),
                _format_moeda_pdf(r['valor_total']),
                _format_moeda_pdf(r['valor_pago']),
                _format_data_pdf(r['ultimo_pagamento']),
                _format_moeda_pdf(r['valor_a_pagar']),
            ],
            'widths': [8, 26, 62, 24, 28, 28, 28, 28],
        },
    }
    config = configs.get(tipo)
    if not config:
        return redirect_empresa(request, 'financeiro:dashboard')

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=config['titulo'],
    )
    styles = getSampleStyleSheet()
    story = []
    story.extend(_flowables_titulo_pdf_centro(config['titulo'], styles))
    meta = ParagraphStyle(
        'dash_card_pdf_meta',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=1,
        textColor=colors.HexColor('#475569'),
    )
    story.append(
        Paragraph(
            xml_escape(
                f'{_nome_empresa_pdf(empresa)} · {data["hoje"].strftime("%d/%m/%Y")} · Total: {_format_moeda_pdf(config["total"])}'
            ),
            meta,
        )
    )
    story.append(Spacer(1, 5 * mm))

    cell = ParagraphStyle('dash_card_pdf_cell', parent=styles['Normal'], fontSize=6.5, leading=8)
    table_data = [config['headers']]
    for linha in config['linhas']:
        table_data.append([_pdf_cell(value, cell) for value in config['row'](linha)])
    if len(table_data) == 1:
        table_data.append(['-', 'Nenhuma informação encontrada.'] + [''] * (len(config['headers']) - 2))
    total_row = [''] * len(config['headers'])
    total_row[-2] = 'Total'
    total_row[-1] = _format_moeda_pdf(config['total'])
    table_data.append(total_row)

    table = Table(
        table_data,
        colWidths=[w * mm for w in config['widths']],
        repeatRows=1,
    )
    table.setStyle(_pdf_table_style())
    last_row = len(table_data) - 1
    table.setStyle(
        TableStyle(
            [
                ('SPAN', (0, last_row), (-3, last_row)),
                ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#e0f2fe')),
                ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
                ('ALIGN', (-2, last_row), (-1, last_row), 'RIGHT'),
            ]
        )
    )
    story.append(table)
    story.extend(_flowables_rodape_impressao(request, styles))
    doc.build(story)

    ref = data['hoje'].strftime('%Y%m%d')
    filename = f'{_safe_filename_part(config["titulo"])}_{ref}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def partial_dashboard_cards(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    if not _is_htmx(request):
        return redirect_empresa(request, 'financeiro:dashboard')

    hoje = timezone.localdate()
    boleto_status_aberto = (
        BoletoPagamento.Status.RASCUNHO,
        BoletoPagamento.Status.EMITIDO,
    )
    boletos_venc_hoje = list(
        BoletoPagamento.objects.filter(
            pagamento_nf__empresa=empresa,
            vencimento=hoje,
            status__in=boleto_status_aberto,
        )
        .select_related('pagamento_nf', 'pagamento_nf__fornecedor')
        .prefetch_related('pagamento_nf__boletos')
        .order_by('vencimento', 'parcela', 'pk')[:50]
    )
    boletos_vencidos = list(
        BoletoPagamento.objects.filter(
            pagamento_nf__empresa=empresa,
            vencimento__lt=hoje,
            status__in=boleto_status_aberto,
        )
        .select_related('pagamento_nf', 'pagamento_nf__fornecedor')
        .prefetch_related('pagamento_nf__boletos', 'pagamento_nf__pagamentos', 'pagamento_nf__itens')
        .order_by('vencimento', 'parcela', 'pk')[:100]
    )
    notas_sem_pagamento = list(
        PagamentoNotaFiscal.objects.filter(
            empresa=empresa,
            data_emissao__lte=hoje,
            pagamentos__isnull=True,
        )
        .select_related('fornecedor', 'caixa')
        .annotate(
            total_itens_calc=Coalesce(
                Subquery(
                    _subquery_total_itens_nf(),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=16, decimal_places=2),
            ),
        )
        .order_by('-data_emissao', '-pk')[:50]
    )
    notas_com_pagamento = (
        PagamentoNotaFiscal.objects.filter(
            empresa=empresa,
            pagamentos__isnull=False,
        )
        .exclude(pagamentos__tipo=PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS)
        .select_related('fornecedor', 'caixa')
        .prefetch_related('pagamentos', 'boletos')
        .annotate(
            total_itens_calc=Coalesce(
                Subquery(
                    _subquery_total_itens_nf(),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=16, decimal_places=2),
            )
        )
        .order_by('-data_emissao', '-pk')
        .distinct()
    )
    notas_pagamento_parcial = []
    for nf in notas_com_pagamento:
        boletos_nf = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
        pagamentos_nf = list(nf.pagamentos.all().order_by('data', 'pk'))
        resumo_nf = _resumo_pagamento_nf(
            pagamentos_nf,
            boletos_nf,
            nf.total_itens_calc,
            hoje,
        )
        if resumo_nf['situacao_label'] != 'Pago Parcial':
            continue
        if resumo_nf['valor_em_aberto'] <= 0:
            continue
        nf.valor_em_aberto_dashboard = resumo_nf['valor_em_aberto']
        nf.pagamento_label_dashboard = resumo_nf['pagamento_label']
        notas_pagamento_parcial.append(nf)
        if len(notas_pagamento_parcial) >= 50:
            break

    def nf_total_dashboard(nf: PagamentoNotaFiscal) -> Decimal:
        tc = getattr(nf, 'total_itens_calc', None)
        if tc is not None:
            return tc
        items = nf.itens.all()
        return sum((i.valor_total for i in items), Decimal('0')).quantize(Decimal('0.01'))

    dashboard_boletos_vencidos_linhas = []
    for b in boletos_vencidos:
        nf = b.pagamento_nf
        boletos_nf = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
        pags_nf = list(nf.pagamentos.all().order_by('data', 'pk'))
        total_nf = nf_total_dashboard(nf)
        resumo = _resumo_pagamento_nf(pags_nf, boletos_nf, total_nf, hoje)
        fornecedor = nf.fornecedor
        dashboard_boletos_vencidos_linhas.append(
            {
                'index': len(dashboard_boletos_vencidos_linhas) + 1,
                'vencimento': b.vencimento,
                'dias_atrasados': (hoje - b.vencimento).days,
                'numero_doc': b.numero_doc or 'â€”',
                'parcela_label': _parcela_dashboard_boleto(b),
                'valor': b.valor,
                'fornecedor_nome': fornecedor.nome,
                'fornecedor_doc': fornecedor.cpf_cnpj_formatado,
                'numero_nf': nf.numero_nf,
                'valor_total': total_nf,
                'valor_pago': resumo['valor_pago'],
                'pagamento_nf_pk': nf.pk,
            }
        )

    dashboard_pagamentos_aberto_linhas = []
    idx_linha = 0
    for nf in notas_sem_pagamento:
        idx_linha += 1
        vt = nf.total_itens_calc
        dashboard_pagamentos_aberto_linhas.append(
            {
                'index': idx_linha,
                'nf_pk': nf.pk,
                'situacao': 'Sem pagamento',
                'badge_class': 'text-bg-secondary',
                'ref_principal': nf.fornecedor.nome,
                'data_emissao': nf.data_emissao,
                'valor_total': vt,
                'valor_pago': Decimal('0'),
                'ultimo_pagamento': None,
                'valor_a_pagar': vt,
            }
        )

    for nf in notas_pagamento_parcial:
        idx_linha += 1
        boletos_nf = list(nf.boletos.all().order_by('vencimento', 'parcela', 'pk'))
        pags_nf = list(nf.pagamentos.all().order_by('data', 'pk'))
        total_nf = nf.total_itens_calc
        resumo = _resumo_pagamento_nf(pags_nf, boletos_nf, total_nf, hoje)
        ultimo = _data_ultimo_pagamento_nf(pags_nf, boletos_nf)
        dashboard_pagamentos_aberto_linhas.append(
            {
                'index': idx_linha,
                'nf_pk': nf.pk,
                'situacao': 'Pago parcial',
                'badge_class': 'text-bg-primary',
                'ref_principal': nf.fornecedor.nome,
                'data_emissao': nf.data_emissao,
                'valor_total': total_nf,
                'valor_pago': resumo['valor_pago'],
                'ultimo_pagamento': ultimo,
                'valor_a_pagar': nf.valor_em_aberto_dashboard,
            }
        )

    dashboard_boletos_venc_hoje_linhas = []
    for ii, b in enumerate(boletos_venc_hoje, start=1):
        nf = b.pagamento_nf
        fornecedor = nf.fornecedor
        dashboard_boletos_venc_hoje_linhas.append(
            {
                'index': ii,
                'numero_doc': b.numero_doc or '—',
                'parcela_label': _parcela_dashboard_boleto(b),
                'valor': b.valor,
                'fornecedor_nome': fornecedor.nome,
                'fornecedor_doc': fornecedor.cpf_cnpj_formatado,
                'numero_nf': nf.numero_nf,
                'pagamento_nf_pk': nf.pk,
            }
        )

    total_boletos_venc_hoje = sum((b.valor for b in boletos_venc_hoje), Decimal('0'))
    total_boletos_vencidos = sum((b.valor for b in boletos_vencidos), Decimal('0'))
    total_notas_sem_pagamento = sum((nf.total_itens_calc for nf in notas_sem_pagamento), Decimal('0'))
    total_notas_pagamento_parcial = sum(
        (nf.valor_em_aberto_dashboard for nf in notas_pagamento_parcial),
        Decimal('0'),
    )
    total_pagamentos_em_aberto = total_notas_sem_pagamento + total_notas_pagamento_parcial
    ultimos_lancamentos = list(
        MovimentoCaixa.objects.filter(empresa=empresa)
        .select_related('caixa')
        .order_by('-data', '-pk')[:10]
    )

    # Feed unificado (últimas movimentações): entradas/saídas + NFs (com/sem pagamento)
    nf_qs = (
        PagamentoNotaFiscal.objects.filter(empresa=empresa)
        .select_related('fornecedor', 'caixa')
        .prefetch_related('pagamentos')
        .annotate(
            total_itens=Coalesce(
                Subquery(
                    _subquery_total_itens_nf(),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=16, decimal_places=2),
            )
        )
        .order_by('-criado_em', '-pk')[:25]
    )
    pessoal_qs = (
        PagamentoPessoal.objects.filter(empresa=empresa)
        .select_related('funcionario', 'caixa')
        .prefetch_related('itens')
        .order_by('-criado_em', '-pk')[:25]
    )
    impostos_qs = (
        PagamentoImposto.objects.filter(empresa=empresa)
        .select_related('autoridade', 'caixa')
        .prefetch_related('itens')
        .order_by('-criado_em', '-pk')[:25]
    )
    eventos = []
    for m in MovimentoCaixa.objects.filter(empresa=empresa).select_related('caixa').order_by(
        '-criado_em', '-pk'
    )[:25]:
        eventos.append(
            {
                'kind': 'movimento',
                'sort_dt': m.criado_em,
                'mov_pk': m.pk,
                'data': m.data,
                'caixa_nome': m.caixa.nome if m.caixa_id else '—',
                'caixa_pk': m.caixa_id,
                'descricao': m.descricao,
                'valor': m.valor,
                'valor_sinal': '+' if m.natureza == MovimentoCaixa.Natureza.ENTRADA else '-',
                'badge_text': 'Entrada' if m.natureza == MovimentoCaixa.Natureza.ENTRADA else 'Saída',
                'badge_class': 'text-bg-success'
                if m.natureza == MovimentoCaixa.Natureza.ENTRADA
                else 'text-bg-danger',
            }
        )
    for nf in nf_qs:
        pagamentos_nf = list(nf.pagamentos.all())
        tipos_pagamento = {pg.tipo for pg in pagamentos_nf}
        if len(tipos_pagamento) > 1:
            badge_text = 'Misto'
            badge_class = 'text-bg-danger'
        elif PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS in tipos_pagamento:
            badge_text = 'Boleto'
            badge_class = 'text-bg-danger'
        elif PagamentoNotaFiscalPagamento.TipoPagamento.AVISTA in tipos_pagamento:
            badge_text = 'À vista'
            badge_class = 'text-bg-danger'
        elif PagamentoNotaFiscalPagamento.TipoPagamento.CREDITO in tipos_pagamento:
            badge_text = 'Crédito'
            badge_class = 'text-bg-danger'
        else:
            badge_text = 'Sem pagamento'
            badge_class = 'text-bg-secondary'

        eventos.append(
            {
                'kind': 'nf',
                'sort_dt': nf.criado_em,
                'data': nf.data_emissao,
                'caixa_nome': nf.caixa.nome if nf.caixa_id else '—',
                'descricao': f'{nf.fornecedor.nome} — NF {nf.numero_nf}',
                'valor': nf.total_itens,
                'valor_sinal': '-',
                'badge_text': badge_text,
                'badge_class': badge_class,
                'nf_pk': nf.pk,
            }
        )
    for pagamento in pessoal_qs:
        descricao_item = ', '.join(
            item.descricao for item in pagamento.itens.all() if item.descricao
        )
        eventos.append(
            {
                'kind': 'pessoal',
                'sort_dt': pagamento.criado_em,
                'data': pagamento.data_pagamento or pagamento.data_emissao,
                'caixa_nome': pagamento.caixa.nome if pagamento.caixa_id else '-',
                'descricao': (
                    f'{pagamento.funcionario.nome} - {descricao_item or "Pagamento pessoal"}'
                    if pagamento.funcionario_id
                    else descricao_item or 'Pagamento pessoal geral'
                ),
                'valor': pagamento.total_itens(),
                'valor_sinal': '-',
                'badge_text': 'À vista',
                'badge_class': 'text-bg-danger',
                'pessoal_pk': pagamento.pk,
            }
        )
    for pagamento in impostos_qs:
        descricao_item = ', '.join(item.descricao for item in pagamento.itens.all() if item.descricao)
        eventos.append(
            {
                'kind': 'imposto',
                'sort_dt': pagamento.criado_em,
                'data': pagamento.data_pagamento or pagamento.data_emissao,
                'caixa_nome': pagamento.caixa.nome if pagamento.caixa_id else '-',
                'descricao': f'{pagamento.autoridade.nome} - {descricao_item or "Pagamento de imposto"}',
                'valor': pagamento.total_itens(),
                'valor_sinal': '-',
                'badge_text': 'À vista',
                'badge_class': 'text-bg-danger',
                'imposto_pk': pagamento.pk,
            }
        )
    eventos.sort(key=lambda e: (e['sort_dt'] or hoje), reverse=True)
    eventos = eventos[:10]

    return render(
        request,
        'financeiro/partials/dashboard_cards_loaded.html',
        {
            'hoje': hoje,
            'dashboard_boletos_venc_hoje_linhas': dashboard_boletos_venc_hoje_linhas,
            'total_boletos_venc_hoje': total_boletos_venc_hoje,
            'dashboard_boletos_vencidos_linhas': dashboard_boletos_vencidos_linhas,
            'total_boletos_vencidos': total_boletos_vencidos,
            'dashboard_pagamentos_aberto_linhas': dashboard_pagamentos_aberto_linhas,
            'total_pagamentos_em_aberto': total_pagamentos_em_aberto,
            'ultimos_lancamentos': ultimos_lancamentos,
            'ultimos_eventos': eventos,
        },
    )


@login_required
def caixa_lista(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    caixas = list(_caixas_com_saldo(empresa, somente_ativos=False))
    hoje = timezone.localdate()
    inicio_mes = date(hoje.year, hoje.month, 1)
    fim_mes = date(hoje.year + 1, 1, 1) if hoje.month == 12 else date(hoje.year, hoje.month + 1, 1)
    for caixa in caixas:
        linhas_mes = _extrato_caixa_detalhado(
            request,
            caixa,
            inicio_mes,
            fim_mes,
            incluir_recebimentos_abertos=True,
        )
        entradas_mes = sum(
            (
                l['valor_liquido']
                for l in linhas_mes
                if l['entrada'] and _linha_extrato_movimento_efetivo(l)
            ),
            Decimal('0'),
        )
        saidas_mes = sum(
            (
                l['valor_liquido']
                for l in linhas_mes
                if not l['entrada'] and _linha_extrato_movimento_efetivo(l)
            ),
            Decimal('0'),
        )
        caixa.saldo_exibicao = entradas_mes - saidas_mes
        caixa.saldo_exibicao_label = 'Saldo Mês Corrente'
        ultimo = next((linha for linha in _extrato_caixa(request, caixa) if linha.get('data')), None)
        caixa.ultimo_lancamento = ultimo
    ativos = [c for c in caixas if c.ativo]
    saldo_unificado_mes = Decimal('0')
    for caixa in ativos:
        linhas_mes = _extrato_caixa_detalhado(
            request,
            caixa,
            inicio_mes,
            fim_mes,
            incluir_recebimentos_abertos=True,
        )
        entradas_mes = sum(
            (
                l['valor_liquido']
                for l in linhas_mes
                if l['entrada'] and _linha_extrato_movimento_efetivo(l)
            ),
            Decimal('0'),
        )
        saidas_mes = sum(
            (
                l['valor_liquido']
                for l in linhas_mes
                if not l['entrada'] and _linha_extrato_movimento_efetivo(l)
            ),
            Decimal('0'),
        )
        saldo_unificado_mes += entradas_mes - saidas_mes
    saldo_consolidado = saldo_unificado_mes
    saldo_consolidado_unificado = sum((c.saldo for c in ativos), Decimal('0'))
    caixa_padrao = next((c for c in caixas if c.tipo == Caixa.Tipo.GERAL), None)

    return render(
        request,
        'financeiro/caixa_lista.html',
        {
            'page_title': 'Caixas',
            'caixas': caixas,
            'saldo_consolidado': saldo_consolidado,
            'saldo_consolidado_unificado': saldo_consolidado_unificado,
            'caixa_padrao': caixa_padrao,
        },
    )


@login_required
def caixa_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = CaixaNovoForm(request.POST or None, empresa=empresa)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Subcaixa criada.')
        return redirect_empresa(request, 'financeiro:caixa_lista')

    return render(
        request,
        'financeiro/caixa_form.html',
        {
            'page_title': 'Nova subcaixa',
            'form': form,
            'modo': 'novo',
        },
    )


@login_required
def caixa_detalhe(request, pk):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    caixa = get_object_or_404(Caixa.objects.filter(empresa=empresa), pk=pk)

    periodo = _periodo_caixa_params(request)
    modo_extrato = (request.GET.get('modo') or 'detalhado').strip().lower()
    if modo_extrato not in {'detalhado', 'consolidado'}:
        modo_extrato = 'detalhado'
    fornecedor_id = _int_param(request, 'fornecedor')
    categoria_id = _int_param(request, 'categoria')
    query_base = {
        'modo': modo_extrato,
        'mes': periodo['mes'],
        'ano': periodo['ano'],
        'fornecedor': fornecedor_id,
        'categoria': categoria_id,
    }

    movimentacoes = _extrato_caixa(request, caixa, periodo['inicio'], periodo['fim'])
    movimentacoes_detalhadas = _extrato_caixa_detalhado(
        request,
        caixa,
        periodo['inicio'],
        periodo['fim'],
        incluir_recebimentos_abertos=True,
    )
    movimentacoes = _recalcular_saldo_linhas(
        _aplicar_filtros_extrato(
            movimentacoes,
            fornecedor_id=fornecedor_id,
            categoria_id=categoria_id,
        ),
        'valor',
    )
    movimentacoes_detalhadas = _recalcular_saldo_linhas(
        _aplicar_filtros_extrato(
            movimentacoes_detalhadas,
            fornecedor_id=fornecedor_id,
            categoria_id=categoria_id,
        ),
        'valor_liquido',
    )

    sort_key = (request.GET.get('sort') or 'data').strip()
    sort_dir = (request.GET.get('dir') or 'desc').strip()
    if sort_key not in {
        'numero',
        'data',
        'natureza',
        'caixa',
        'categoria',
        'nf',
        'pessoa',
        'descricao',
        'valor_bruto',
        'descontos',
        'valor_liquido',
        'saldo',
    }:
        sort_key = 'data'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'
    movimentacoes_detalhadas = _ordenar_extrato_detalhado(
        movimentacoes_detalhadas,
        sort_key,
        sort_dir,
    )

    linhas_ativas = movimentacoes if modo_extrato == 'consolidado' else movimentacoes_detalhadas
    paginator = Paginator(linhas_ativas, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    campo_total = 'valor' if modo_extrato == 'consolidado' else 'valor_liquido'
    entradas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas_ativas
            if linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    saidas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas_ativas
            if not linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    page_query = dict(query_base)
    if modo_extrato == 'detalhado':
        page_query.update({'sort': sort_key, 'dir': sort_dir})

    return render(
        request,
        'financeiro/caixa_detalhe.html',
        {
            'page_title': caixa.nome,
            'caixa': caixa,
            'movimentacoes': page_obj.object_list if modo_extrato == 'consolidado' else movimentacoes,
            'movimentacoes_detalhadas': (
                page_obj.object_list if modo_extrato == 'detalhado' else movimentacoes_detalhadas
            ),
            'page_obj': page_obj,
            'paginator': paginator,
            'total_movimentacoes_detalhadas': paginator.count,
            'modo_extrato': modo_extrato,
            'periodo': periodo,
            'detalhado_url': f'?{_query_string({**query_base, "modo": "detalhado", "sort": sort_key, "dir": sort_dir})}',
            'consolidado_url': f'?{_query_string({**query_base, "modo": "consolidado"})}',
            'pdf_url': reverse_empresa(
                request,
                'financeiro:caixa_extrato_pdf',
                kwargs={'pk': caixa.pk},
            )
            + f'?{_query_string(page_query)}',
            'fornecedores_filtro': Fornecedor.objects.filter(empresa=empresa).order_by('nome'),
            'categorias_filtro': CategoriaFinanceira.objects.filter(empresa=empresa).order_by(
                'movimentacao_tipo',
                'nome',
            ),
            'caixas_filtro': Caixa.objects.filter(empresa=empresa, ativo=True).order_by('tipo', 'nome'),
            'fornecedor_id': fornecedor_id,
            'categoria_id': categoria_id,
            'caixa_id': None,
            'unificado': False,
            'sort_key': sort_key,
            'sort_dir': sort_dir,
            'sort_links': _sort_links_extrato_detalhado(sort_key, sort_dir, query_base),
            'page_query': _query_string(page_query),
            'entradas_total': entradas_total,
            'saidas_total': saidas_total,
            'saldo_atual': entradas_total - saidas_total,
        },
    )


@login_required
def caixa_unificado(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    caixas_ativos = list(_caixas_com_saldo(empresa, somente_ativos=True))
    caixa_placeholder = Caixa(
        empresa=empresa,
        tipo=Caixa.Tipo.GERAL,
        nome='Caixa Unificado',
        ativo=True,
    )
    periodo = _periodo_caixa_params(request)
    modo_extrato = (request.GET.get('modo') or 'detalhado').strip().lower()
    if modo_extrato not in {'detalhado', 'consolidado'}:
        modo_extrato = 'detalhado'
    fornecedor_id = _int_param(request, 'fornecedor')
    categoria_id = _int_param(request, 'categoria')
    caixa_id = _int_param(request, 'caixa')
    query_base = {
        'modo': modo_extrato,
        'mes': periodo['mes'],
        'ano': periodo['ano'],
        'fornecedor': fornecedor_id,
        'categoria': categoria_id,
        'caixa': caixa_id,
    }

    movimentacoes = []
    movimentacoes_detalhadas = []
    for caixa in caixas_ativos:
        movimentacoes.extend(_extrato_caixa(request, caixa, periodo['inicio'], periodo['fim']))
        movimentacoes_detalhadas.extend(
            _extrato_caixa_detalhado(
                request,
                caixa,
                periodo['inicio'],
                periodo['fim'],
                incluir_recebimentos_abertos=True,
            )
        )

    movimentacoes = _recalcular_saldo_linhas(
        _aplicar_filtros_extrato(
            movimentacoes,
            fornecedor_id=fornecedor_id,
            categoria_id=categoria_id,
            caixa_id=caixa_id,
        ),
        'valor',
    )
    movimentacoes_detalhadas = _recalcular_saldo_linhas(
        _aplicar_filtros_extrato(
            movimentacoes_detalhadas,
            fornecedor_id=fornecedor_id,
            categoria_id=categoria_id,
            caixa_id=caixa_id,
        ),
        'valor_liquido',
    )

    sort_key = (request.GET.get('sort') or 'data').strip()
    sort_dir = (request.GET.get('dir') or 'desc').strip()
    if sort_key not in {
        'numero',
        'data',
        'natureza',
        'caixa',
        'categoria',
        'nf',
        'pessoa',
        'descricao',
        'valor_bruto',
        'descontos',
        'valor_liquido',
        'saldo',
    }:
        sort_key = 'data'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'
    movimentacoes_detalhadas = _ordenar_extrato_detalhado(
        movimentacoes_detalhadas,
        sort_key,
        sort_dir,
    )

    linhas_ativas = movimentacoes if modo_extrato == 'consolidado' else movimentacoes_detalhadas
    paginator = Paginator(linhas_ativas, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    campo_total = 'valor' if modo_extrato == 'consolidado' else 'valor_liquido'
    entradas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas_ativas
            if linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    saidas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas_ativas
            if not linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    page_query = dict(query_base)
    if modo_extrato == 'detalhado':
        page_query.update({'sort': sort_key, 'dir': sort_dir})

    return render(
        request,
        'financeiro/caixa_detalhe.html',
        {
            'page_title': 'Caixa Unificado',
            'caixa': caixa_placeholder,
            'movimentacoes': page_obj.object_list if modo_extrato == 'consolidado' else movimentacoes,
            'movimentacoes_detalhadas': (
                page_obj.object_list if modo_extrato == 'detalhado' else movimentacoes_detalhadas
            ),
            'page_obj': page_obj,
            'paginator': paginator,
            'total_movimentacoes_detalhadas': paginator.count,
            'modo_extrato': modo_extrato,
            'periodo': periodo,
            'detalhado_url': f'?{_query_string({**query_base, "modo": "detalhado", "sort": sort_key, "dir": sort_dir})}',
            'consolidado_url': f'?{_query_string({**query_base, "modo": "consolidado"})}',
            'pdf_url': reverse_empresa(request, 'financeiro:caixa_unificado_pdf')
            + f'?{_query_string(page_query)}',
            'fornecedores_filtro': Fornecedor.objects.filter(empresa=empresa).order_by('nome'),
            'categorias_filtro': CategoriaFinanceira.objects.filter(empresa=empresa).order_by(
                'movimentacao_tipo',
                'nome',
            ),
            'caixas_filtro': caixas_ativos,
            'fornecedor_id': fornecedor_id,
            'categoria_id': categoria_id,
            'caixa_id': caixa_id,
            'sort_key': sort_key,
            'sort_dir': sort_dir,
            'sort_links': _sort_links_extrato_detalhado(sort_key, sort_dir, query_base),
            'page_query': _query_string(page_query),
            'entradas_total': entradas_total,
            'saidas_total': saidas_total,
            'saldo_atual': entradas_total - saidas_total,
            'unificado': True,
        },
    )


def _safe_filename_part(text):
    s = re.sub(r'[^\w\s\-]', '_', str(text), flags=re.UNICODE)
    s = re.sub(r'\s+', '_', s.strip())
    return (s[:60] or 'extrato').strip('_')


def _format_data_pdf(valor) -> str:
    return valor.strftime('%d/%m/%Y') if valor else '-'


def _format_moeda_pdf(valor) -> str:
    return f'R$ {format_decimal_br_moeda(valor or Decimal("0"))}'


def _mes_nome_pt(mes: int) -> str:
    return dict(MESES_FILTRO_CAIXA).get(int(mes or 0), '')


def _nome_empresa_pdf(empresa) -> str:
    return (getattr(empresa, 'nome_fantasia', '') or getattr(empresa, 'razao_social', '') or 'Empresa').strip()


def _linha_extra_cabecalho_recibo(empresa):
    parts = []
    cnpj = (getattr(empresa, 'cnpj', None) or '').strip()
    if cnpj:
        parts.append(f'CNPJ: {cnpj}')
    end = (getattr(empresa, 'endereco', None) or '').strip()
    if end:
        parts.append(end)
    tel = (getattr(empresa, 'telefone', None) or '').strip()
    if tel:
        parts.append(f'Tel: {tel}')
    return ' · '.join(parts) if parts else ''


def _empresa_logo_flowable(empresa):
    logo = getattr(empresa, 'logo', None)
    path = getattr(logo, 'path', None)
    if not path:
        return None
    try:
        img = Image(path)
        max_w = 34 * mm
        max_h = 24 * mm
        scale = min(max_w / img.imageWidth, max_h / img.imageHeight, 1)
        img.drawWidth = img.imageWidth * scale
        img.drawHeight = img.imageHeight * scale
        return img
    except Exception:
        return None


def _header_text_html_caixa(empresa, periodo, caixa):
    nome = _nome_empresa_pdf(empresa)
    modo = periodo.get('modo_label') or 'Extrato'
    mes_nome = 'Todos os meses' if periodo.get('todos_os_meses') else _mes_nome_pt(periodo['mes'])
    bits = [f'<b>{xml_escape(nome)}</b>']
    razao = (getattr(empresa, 'razao_social', '') or '').strip()
    if razao and razao.upper() != (nome or '').upper():
        bits.append(f'<font size="8" color="#64748b">{xml_escape(razao)}</font>')
    bits.append(xml_escape(f'{modo} · {caixa.nome} · {mes_nome}/{periodo["ano"]}'))
    extra = _linha_extra_cabecalho_recibo(empresa)
    if extra:
        bits.append(xml_escape(extra))
    email = (getattr(empresa, 'email', None) or '').strip()
    if email:
        bits.append(xml_escape(f'E-mail: {email}'))
    return '<br/>'.join(bits)


def _flowables_header_compact(empresa, periodo, caixa, styles):
    hdr_style = ParagraphStyle(
        'cx_hdr_compact',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        alignment=0,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=0,
        spaceBefore=0,
    )
    logo_w = 34 * mm
    content_w = 273 * mm
    p = Paragraph(_header_text_html_caixa(empresa, periodo, caixa), hdr_style)
    logo = _empresa_logo_flowable(empresa) or ''
    tbl = Table([[logo, p]], colWidths=[logo_w, content_w - logo_w])
    tbl.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
                ('LEFTPADDING', (1, 0), (1, 0), 3 * mm),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
        )
    )
    return [tbl, Spacer(1, 2 * mm)]


def _flowables_titulo_pdf_centro(titulo_doc, styles):
    t = ParagraphStyle(
        'cx_titulo_doc_pdf',
        parent=styles['Normal'],
        fontSize=12,
        leading=14,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.HexColor('#0f172a'),
        fontName='Helvetica-Bold',
    )
    return [
        Spacer(1, 6 * mm),
        Paragraph(f'<b>{xml_escape(titulo_doc)}</b>', t),
        Spacer(1, 5 * mm),
    ]


def _flowables_rodape_impressao(request, styles, *, space_before_mm: float = 2.5):
    user = getattr(request, 'user', None)
    nome = '—'
    if getattr(user, 'is_authenticated', False):
        nome = (user.get_full_name() or getattr(user, 'username', '') or '—').strip()
    emitido = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')
    rodape_style = ParagraphStyle(
        'cx_rodape_impressao',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6,
        leading=7.5,
        textColor=colors.HexColor('#94a3b8'),
        alignment=2,
    )
    return [
        Spacer(1, space_before_mm * mm),
        Paragraph(xml_escape(f'{nome} · {emitido} · Gênesis ERP'), rodape_style),
    ]


def _pdf_cell(text, style):
    return Paragraph(xml_escape(str(text if text is not None else '-')), style)


def _pdf_table_style():
    return TableStyle(
        [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6.3),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2.2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2.2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2.8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2.8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]
    )


def _pdf_totais_table(entradas, saidas, saldo, quantidade):
    data = [[
        'Entradas',
        'Saídas',
        'Saldo do período',
        'Movimentações',
    ], [
        _format_moeda_pdf(entradas),
        _format_moeda_pdf(saidas),
        _format_moeda_pdf(saldo),
        str(quantidade),
    ]]
    table = Table(data, colWidths=[45 * mm, 45 * mm, 45 * mm, 35 * mm])
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#64748b')),
                ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#15803d')),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#b91c1c')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7.5),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('BOX', (0, 0), (-1, -1), 0.35, colors.HexColor('#cbd5e1')),
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
            ]
        )
    )
    return table


@login_required
def caixa_extrato_pdf(request, pk):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    caixa = get_object_or_404(Caixa.objects.filter(empresa=empresa), pk=pk)
    periodo = _periodo_caixa_params(request)
    modo_extrato = (request.GET.get('modo') or 'detalhado').strip().lower()
    if modo_extrato not in {'detalhado', 'consolidado'}:
        modo_extrato = 'detalhado'
    fornecedor_id = _int_param(request, 'fornecedor')
    categoria_id = _int_param(request, 'categoria')

    if modo_extrato == 'consolidado':
        linhas = _extrato_caixa(request, caixa, periodo['inicio'], periodo['fim'])
        linhas = _recalcular_saldo_linhas(
            _aplicar_filtros_extrato(
                linhas,
                fornecedor_id=fornecedor_id,
                categoria_id=categoria_id,
            ),
            'valor',
        )
        titulo = 'Extrato Consolidado'
        campo_total = 'valor'
    else:
        linhas = _extrato_caixa_detalhado(
            request,
            caixa,
            periodo['inicio'],
            periodo['fim'],
            incluir_recebimentos_abertos=True,
        )
        linhas = _recalcular_saldo_linhas(
            _aplicar_filtros_extrato(
                linhas,
                fornecedor_id=fornecedor_id,
                categoria_id=categoria_id,
            ),
            'valor_liquido',
        )
        linhas = _ordenar_extrato_detalhado(
            linhas,
            (request.GET.get('sort') or 'data').strip(),
            (request.GET.get('dir') or 'desc').strip(),
        )
        titulo = 'Extrato Detalhado'
        campo_total = 'valor_liquido'

    periodo['modo_label'] = titulo
    entradas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas
            if linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    saidas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas
            if not linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    cell = ParagraphStyle(
        'cx_pdf_cell',
        parent=styles['Normal'],
        fontSize=6.3,
        leading=7.4,
        spaceBefore=0,
        spaceAfter=0,
    )
    cell_right = ParagraphStyle('cx_pdf_cell_right', parent=cell, alignment=2)
    story = []
    story.extend(_flowables_header_compact(empresa, periodo, caixa, styles))
    story.extend(_flowables_titulo_pdf_centro(titulo, styles))
    story.append(_pdf_totais_table(entradas_total, saidas_total, entradas_total - saidas_total, len(linhas)))
    story.append(Spacer(1, 4 * mm))

    if modo_extrato == 'consolidado':
        data = [['#', 'Data', 'Natureza', 'Origem', 'Descrição', 'Valor', 'Saldo Após']]
        for linha in linhas:
            sinal = '+' if linha['entrada'] else '-'
            data.append([
                str(linha['numero']),
                _format_data_pdf(linha['data']),
                linha['natureza'],
                linha['origem'],
                _pdf_cell(linha['descricao'], cell),
                _pdf_cell(f'{sinal} {_format_moeda_pdf(linha["valor"])}', cell_right),
                _pdf_cell(_format_moeda_pdf(linha['saldo_apos_lancamento']), cell_right),
            ])
        widths = [8, 18, 18, 32, 115, 32, 32]
    else:
        data = [[
            '#',
            'Data',
            'Natureza',
            'Categoria',
            'NF',
            'Fornecedor/Cliente',
            'Descrição',
            'Bruto',
            'Desc.',
            'Líquido',
            'Saldo Após',
        ]]
        for linha in linhas:
            sinal = '+' if linha['entrada'] else '-'
            data.append([
                str(linha['numero']),
                _format_data_pdf(linha['data']),
                linha['natureza'],
                linha['categoria'],
                linha['nf'],
                _pdf_cell(linha['pessoa'], cell),
                _pdf_cell(linha['descricao'], cell),
                _pdf_cell(_format_moeda_pdf(linha['valor_bruto']), cell_right),
                _pdf_cell(_format_moeda_pdf(linha['descontos']), cell_right),
                _pdf_cell(f'{sinal} {_format_moeda_pdf(linha["valor_liquido"])}', cell_right),
                _pdf_cell(_format_moeda_pdf(linha['saldo_apos_lancamento']), cell_right),
            ])
        widths = [7, 16, 16, 25, 18, 40, 67, 22, 20, 24, 24]

    table = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    table.setStyle(_pdf_table_style())
    story.append(table)
    story.extend(_flowables_rodape_impressao(request, styles, space_before_mm=3))
    doc.build(story)

    ref = f'{periodo["ano"]}-{periodo["mes"]:02d}'
    filename = f'Caixa_{_safe_filename_part(caixa.nome)}_{modo_extrato}_{ref}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def caixa_unificado_pdf(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    caixas_ativos = list(_caixas_com_saldo(empresa, somente_ativos=True))
    caixa_placeholder = Caixa(
        empresa=empresa,
        tipo=Caixa.Tipo.GERAL,
        nome='Caixa Unificado',
        ativo=True,
    )
    periodo = _periodo_caixa_params(request)
    modo_extrato = (request.GET.get('modo') or 'detalhado').strip().lower()
    if modo_extrato not in {'detalhado', 'consolidado'}:
        modo_extrato = 'detalhado'
    fornecedor_id = _int_param(request, 'fornecedor')
    categoria_id = _int_param(request, 'categoria')
    caixa_id = _int_param(request, 'caixa')

    linhas = []
    for caixa in caixas_ativos:
        if modo_extrato == 'consolidado':
            linhas.extend(_extrato_caixa(request, caixa, periodo['inicio'], periodo['fim']))
        else:
            linhas.extend(
                _extrato_caixa_detalhado(
                    request,
                    caixa,
                    periodo['inicio'],
                    periodo['fim'],
                    incluir_recebimentos_abertos=True,
                )
            )

    campo_total = 'valor' if modo_extrato == 'consolidado' else 'valor_liquido'
    linhas = _recalcular_saldo_linhas(
        _aplicar_filtros_extrato(
            linhas,
            fornecedor_id=fornecedor_id,
            categoria_id=categoria_id,
            caixa_id=caixa_id,
        ),
        campo_total,
    )
    titulo = 'Extrato Consolidado' if modo_extrato == 'consolidado' else 'Extrato Detalhado'
    if modo_extrato == 'detalhado':
        linhas = _ordenar_extrato_detalhado(
            linhas,
            (request.GET.get('sort') or 'data').strip(),
            (request.GET.get('dir') or 'desc').strip(),
        )

    periodo['modo_label'] = f'{titulo} Unificado'
    entradas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas
            if linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    saidas_total = sum(
        (
            (linha.get(campo_total) or Decimal('0'))
            for linha in linhas
            if not linha['entrada'] and _linha_extrato_movimento_efetivo(linha)
        ),
        Decimal('0'),
    )
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    cell = ParagraphStyle(
        'cx_pdf_cell_unificado',
        parent=styles['Normal'],
        fontSize=6.1,
        leading=7.2,
        spaceBefore=0,
        spaceAfter=0,
    )
    cell_right = ParagraphStyle('cx_pdf_cell_unificado_right', parent=cell, alignment=2)
    story = []
    story.extend(_flowables_header_compact(empresa, periodo, caixa_placeholder, styles))
    story.extend(_flowables_titulo_pdf_centro(f'{titulo} - Caixa Unificado', styles))
    story.append(_pdf_totais_table(entradas_total, saidas_total, entradas_total - saidas_total, len(linhas)))
    story.append(Spacer(1, 4 * mm))

    if modo_extrato == 'consolidado':
        data = [['#', 'Data', 'Caixa', 'Natureza', 'Origem', 'Descrição', 'Valor', 'Saldo Após']]
        for linha in linhas:
            sinal = '+' if linha['entrada'] else '-'
            data.append([
                str(linha['numero']),
                _format_data_pdf(linha['data']),
                linha.get('caixa_nome') or '-',
                linha['natureza'],
                linha['origem'],
                _pdf_cell(linha['descricao'], cell),
                _pdf_cell(f'{sinal} {_format_moeda_pdf(linha["valor"])}', cell_right),
                _pdf_cell(_format_moeda_pdf(linha['saldo_apos_lancamento']), cell_right),
            ])
        widths = [7, 16, 30, 16, 27, 96, 28, 28]
    else:
        data = [['#', 'Data', 'Caixa', 'Natureza', 'Categoria', 'NF', 'Fornecedor/Cliente', 'Descrição', 'Bruto', 'Desc.', 'Líquido', 'Saldo Após']]
        for linha in linhas:
            sinal = '+' if linha['entrada'] else '-'
            data.append([
                str(linha['numero']),
                _format_data_pdf(linha['data']),
                linha.get('caixa_nome') or '-',
                linha['natureza'],
                linha['categoria'],
                linha['nf'],
                _pdf_cell(linha['pessoa'], cell),
                _pdf_cell(linha['descricao'], cell),
                _pdf_cell(_format_moeda_pdf(linha['valor_bruto']), cell_right),
                _pdf_cell(_format_moeda_pdf(linha['descontos']), cell_right),
                _pdf_cell(f'{sinal} {_format_moeda_pdf(linha["valor_liquido"])}', cell_right),
                _pdf_cell(_format_moeda_pdf(linha['saldo_apos_lancamento']), cell_right),
            ])
        widths = [7, 15, 24, 15, 22, 16, 34, 58, 20, 18, 22, 22]

    table = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    table.setStyle(_pdf_table_style())
    story.append(table)
    story.extend(_flowables_rodape_impressao(request, styles, space_before_mm=3))
    doc.build(story)

    ref = f'{periodo["ano"]}-{periodo["mes"]:02d}'
    filename = f'Caixa_Unificado_{modo_extrato}_{ref}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def caixa_editar(request, pk):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    caixa = get_object_or_404(Caixa.objects.filter(empresa=empresa), pk=pk)
    form = CaixaEditForm(
        request.POST or None,
        instance=caixa,
        empresa=empresa,
    )
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Caixa atualizado.')
        return redirect_empresa(request, 'financeiro:caixa_detalhe', kwargs={'pk': caixa.pk})

    return render(
        request,
        'financeiro/caixa_form.html',
        {
            'page_title': f'Editar — {caixa.nome}',
            'form': form,
            'caixa': caixa,
            'modo': 'editar',
        },
    )


@login_required
def caixa_inativar(request, pk):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    if request.method != 'POST':
        return redirect_empresa(request, 'financeiro:caixa_detalhe', kwargs={'pk': pk})

    caixa = get_object_or_404(Caixa.objects.filter(empresa=empresa), pk=pk)
    caixa.ativo = False
    caixa.save()
    messages.success(request, 'Caixa inativado. O histórico de movimentos foi preservado.')
    return redirect_empresa(request, 'financeiro:caixa_lista')


@login_required
def caixa_reativar(request, pk):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    if request.method != 'POST':
        return redirect_empresa(request, 'financeiro:caixa_detalhe', kwargs={'pk': pk})

    caixa = get_object_or_404(Caixa.objects.filter(empresa=empresa), pk=pk)
    caixa.ativo = True
    caixa.save()
    messages.success(request, 'Caixa reativado.')
    return redirect_empresa(request, 'financeiro:caixa_detalhe', kwargs={'pk': caixa.pk})


@login_required
def movimentar_caixa(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    filtros_abertos = _recebimentos_filtros(
        request,
        empresa,
        prefixo='abertos',
        usar_mes_corrente_padrao=False,
    )
    filtros_pagos = _recebimentos_filtros(
        request,
        empresa,
        prefixo='pagos',
        usar_mes_corrente_padrao=True,
    )
    recebimentos_abertos, recebimentos_pagos = _recebimentos_movimentar_listas(
        empresa,
        filtros_abertos=filtros_abertos,
        filtros_pagos=filtros_pagos,
    )
    return render(
        request,
        'financeiro/movimentar_caixa.html',
        {
            'page_title': 'Recebimentos',
            'filtros_abertos': filtros_abertos,
            'filtros_pagos': filtros_pagos,
            'recebimentos_abertos': recebimentos_abertos,
            'recebimentos_pagos': recebimentos_pagos,
            'totais_abertos': _totais_recebimentos(recebimentos_abertos),
            'totais_pagos': _totais_recebimentos(recebimentos_pagos),
            'abertos_pdf_query': _recebimentos_query_string(filtros_abertos),
            'pagos_pdf_query': _recebimentos_query_string(filtros_pagos),
            'abertos_limpar_query': _recebimentos_query_string(filtros_pagos),
            'pagos_limpar_query': _recebimentos_query_string(filtros_abertos),
        },
    )


def _recebimentos_filtros(request, empresa, *, prefixo: str, usar_mes_corrente_padrao: bool):
    campo_data_inicio = f'{prefixo}_data_inicio'
    campo_data_fim = f'{prefixo}_data_fim'
    campo_obra = f'{prefixo}_obra'
    obra_id = _int_param(request, campo_obra)
    obra_selecionada = None
    obras = list(Obra.objects.filter(empresa=empresa).order_by('nome'))
    if obra_id:
        obra_selecionada = next((obra for obra in obras if obra.pk == obra_id), None)

    hoje = timezone.localdate()
    primeiro_dia_mes = hoje.replace(day=1)
    proximo_mes = (
        primeiro_dia_mes.replace(year=primeiro_dia_mes.year + 1, month=1)
        if primeiro_dia_mes.month == 12
        else primeiro_dia_mes.replace(month=primeiro_dia_mes.month + 1)
    )
    ultimo_dia_mes = proximo_mes - timedelta(days=1)
    data_inicio_raw = (request.GET.get(campo_data_inicio) or '').strip()
    data_fim_raw = (request.GET.get(campo_data_fim) or '').strip()
    data_inicio = parse_date(data_inicio_raw)
    data_fim = parse_date(data_fim_raw)
    if usar_mes_corrente_padrao:
        data_inicio = data_inicio or primeiro_dia_mes
        data_fim = data_fim or ultimo_dia_mes

    return {
        'prefixo': prefixo,
        'campo_data_inicio': campo_data_inicio,
        'campo_data_fim': campo_data_fim,
        'campo_obra': campo_obra,
        'data_inicio_raw': data_inicio.isoformat() if data_inicio else data_inicio_raw,
        'data_fim_raw': data_fim.isoformat() if data_fim else data_fim_raw,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'obra_id': obra_selecionada.pk if obra_selecionada else None,
        'obra': obra_selecionada,
        'obras': obras,
    }


def _recebimentos_query_string(filtros):
    params = {}
    if filtros.get('data_inicio_raw'):
        params[filtros['campo_data_inicio']] = filtros['data_inicio_raw']
    if filtros.get('data_fim_raw'):
        params[filtros['campo_data_fim']] = filtros['data_fim_raw']
    if filtros.get('obra_id'):
        params[filtros['campo_obra']] = filtros['obra_id']
    return urlencode(params)


def _data_referencia_recebimento(recebimento):
    if recebimento['status'] in (RecebimentoAvulso.Status.PAGO, RecebimentoMedicao.Status.PAGO):
        return recebimento.get('data_pagamento') or recebimento.get('data')
    return recebimento.get('data')


def _aplicar_filtros_recebimentos(recebimentos, filtros):
    data_inicio = filtros.get('data_inicio')
    data_fim = filtros.get('data_fim')
    obra_id = filtros.get('obra_id')

    filtrados = []
    for recebimento in recebimentos:
        data_ref = _data_referencia_recebimento(recebimento)
        if data_inicio and (not data_ref or data_ref < data_inicio):
            continue
        if data_fim and (not data_ref or data_ref > data_fim):
            continue
        if obra_id:
            obra = recebimento.get('obra')
            if not obra or obra.pk != obra_id:
                continue
        filtrados.append(recebimento)
    return filtrados


def _recebimentos_movimentar_listas(empresa, *, filtros_abertos=None, filtros_pagos=None):
    recebimentos_avulsos = (
        RecebimentoAvulso.objects.filter(empresa=empresa)
        .select_related('caixa', 'cliente', 'conta_bancaria', 'movimento')
        .order_by('-data', '-pk')
    )
    recebimentos_medicao = (
        RecebimentoMedicao.objects.filter(empresa=empresa)
        .select_related('caixa', 'cliente', 'obra', 'conta_bancaria', 'movimento')
        .order_by('-data', '-pk')
    )
    recebimentos = [
        *(_recebimento_para_linha(r, 'avulso') for r in recebimentos_avulsos),
        *(_recebimento_para_linha(r, 'medicao') for r in recebimentos_medicao),
    ]
    recebimentos.sort(key=lambda r: (r['data'] or timezone.localdate(), r['pk']), reverse=True)
    recebimentos_abertos = [
        r for r in recebimentos
        if r['status'] in (RecebimentoAvulso.Status.ABERTO, RecebimentoMedicao.Status.ABERTO)
    ]
    recebimentos_pagos = [
        r for r in recebimentos
        if r['status'] in (RecebimentoAvulso.Status.PAGO, RecebimentoMedicao.Status.PAGO)
    ]
    if filtros_abertos:
        recebimentos_abertos = _aplicar_filtros_recebimentos(recebimentos_abertos, filtros_abertos)
    if filtros_pagos:
        recebimentos_pagos = _aplicar_filtros_recebimentos(recebimentos_pagos, filtros_pagos)
    recebimentos_pagos = recebimentos_pagos[:ULTIMOS_RECEBIMENTOS_LISTA_LIMITE]
    return recebimentos_abertos, recebimentos_pagos


def _flowables_header_recebimentos(empresa, subtitulo, styles):
    hdr_style = ParagraphStyle(
        'rec_hdr_compact',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        alignment=0,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=0,
        spaceBefore=0,
    )
    nome = _nome_empresa_pdf(empresa)
    bits = [f'<b>{xml_escape(nome)}</b>']
    razao = (getattr(empresa, 'razao_social', '') or '').strip()
    if razao and razao.upper() != (nome or '').upper():
        bits.append(f'<font size="8" color="#64748b">{xml_escape(razao)}</font>')
    bits.append(xml_escape(subtitulo))
    extra = _linha_extra_cabecalho_recibo(empresa)
    if extra:
        bits.append(xml_escape(extra))
    email = (getattr(empresa, 'email', None) or '').strip()
    if email:
        bits.append(xml_escape(f'E-mail: {email}'))

    logo_w = 34 * mm
    content_w = 273 * mm
    logo = _empresa_logo_flowable(empresa) or ''
    texto = Paragraph('<br/>'.join(bits), hdr_style)
    table = Table([[logo, texto]], colWidths=[logo_w, content_w - logo_w])
    table.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
                ('LEFTPADDING', (1, 0), (1, 0), 3 * mm),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
        )
    )
    return [table, Spacer(1, 2 * mm)]


@login_required
def recebimentos_pdf(request, status: str):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    if status == 'abertos':
        filtros = _recebimentos_filtros(
            request,
            empresa,
            prefixo='abertos',
            usar_mes_corrente_padrao=False,
        )
        recebimentos_abertos, _recebimentos_pagos = _recebimentos_movimentar_listas(
            empresa,
            filtros_abertos=filtros,
        )
        recebimentos = recebimentos_abertos
        titulo = 'Recebimentos em Aberto'
        subtitulo = 'Recebimentos em aberto · Ainda não lançados no caixa'
        filename_status = 'Abertos'
    elif status == 'pagos':
        filtros = _recebimentos_filtros(
            request,
            empresa,
            prefixo='pagos',
            usar_mes_corrente_padrao=True,
        )
        _recebimentos_abertos, recebimentos_pagos = _recebimentos_movimentar_listas(
            empresa,
            filtros_pagos=filtros,
        )
        recebimentos = recebimentos_pagos
        titulo = 'Recebimentos Pagos/Liquidados'
        subtitulo = f'Pagos/Liquidados · Últimos {ULTIMOS_RECEBIMENTOS_LISTA_LIMITE} lançados no caixa'
        filename_status = 'Liquidados'
    else:
        messages.error(request, 'Tipo de relatório de recebimentos inválido.')
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    filtro_partes = []
    if filtros.get('data_inicio'):
        filtro_partes.append(f'A partir de {_format_data_pdf(filtros["data_inicio"])}')
    if filtros.get('data_fim'):
        filtro_partes.append(f'Até {_format_data_pdf(filtros["data_fim"])}')
    if filtros.get('obra'):
        filtro_partes.append(f'Obra: {filtros["obra"].nome}')
    if filtro_partes:
        subtitulo = f'{subtitulo} · {" · ".join(filtro_partes)}'

    totais = _totais_recebimentos(recebimentos)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle(
        'rec_pdf_cell',
        parent=styles['Normal'],
        fontSize=5.8,
        leading=7,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.HexColor('#0f172a'),
    )
    cell_right = ParagraphStyle('rec_pdf_cell_right', parent=cell, alignment=2)

    if status == 'pagos':
        data = [[
            '#',
            'Data',
            'Data Pgto.',
            'Tipo',
            'Medição',
            'NF',
            'Cliente',
            'Obra',
            'Bruto',
            'Impostos',
            'Líquido',
            'Recebido em',
        ]]
        col_widths = [7, 15, 16, 14, 18, 16, 40, 32, 22, 20, 22, 28]
    else:
        data = [[
            '#',
            'Data',
            'Tipo',
            'Medição',
            'NF',
            'Cliente',
            'Obra',
            'Bruto',
            'Impostos',
            'Líquido',
        ]]
        col_widths = [8, 18, 18, 22, 18, 56, 42, 28, 25, 28]

    for idx, recebimento in enumerate(recebimentos, start=1):
        obra = getattr(recebimento.get('obra'), 'nome', '') or '-'
        row = [
            str(idx),
            _format_data_pdf(recebimento.get('data')),
        ]
        if status == 'pagos':
            row.append(_format_data_pdf(recebimento.get('data_pagamento')))
        row.extend(
            [
                recebimento.get('tipo_label') or '-',
                recebimento.get('medicao') or '-',
                recebimento.get('nf') or '-',
                _pdf_cell(getattr(recebimento.get('cliente'), 'nome', '') or '-', cell),
                _pdf_cell(obra, cell),
                _pdf_cell(_format_moeda_pdf(recebimento.get('valor')), cell_right),
                _pdf_cell(_format_moeda_pdf(recebimento.get('impostos')), cell_right),
                _pdf_cell(_format_moeda_pdf(recebimento.get('valor_liquido')), cell_right),
            ]
        )
        if status == 'pagos':
            recebido_em = getattr(recebimento.get('conta_bancaria'), 'nome', '') or 'Dinheiro'
            row.append(_pdf_cell(recebido_em, cell))
        data.append(row)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=titulo,
    )
    story = []
    story.extend(_flowables_header_recebimentos(empresa, subtitulo, styles))
    story.extend(_flowables_titulo_pdf_centro(titulo, styles))

    table = Table(
        data,
        colWidths=[w * mm for w in col_widths],
        repeatRows=1,
    )
    table.setStyle(_pdf_table_style())
    story.append(table)
    story.append(Spacer(1, 3 * mm))

    totais_table = Table(
        [[
            'Quantidade',
            'Valor Bruto',
            'Impostos',
            'Valor Líquido',
        ], [
            str(len(recebimentos)),
            _format_moeda_pdf(totais['valor']),
            _format_moeda_pdf(totais['impostos']),
            _format_moeda_pdf(totais['valor_liquido']),
        ]],
        colWidths=[34 * mm, 48 * mm, 48 * mm, 48 * mm],
    )
    totais_table.setStyle(_pdf_table_style())
    story.append(totais_table)
    story.extend(_flowables_rodape_impressao(request, styles, space_before_mm=3))
    doc.build(story)

    ref = timezone.localdate().strftime('%Y%m%d')
    filename = f'Recebimentos_{filename_status}_{_safe_filename_part(_nome_empresa_pdf(empresa))}_{ref}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def recebimento_liquidar(request, tipo: str, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    if tipo == 'avulso':
        recebimento = get_object_or_404(
            RecebimentoAvulso.objects.filter(empresa=empresa).select_related('caixa'),
            pk=pk,
        )
        categoria_origem = MovimentoCaixa.CategoriaOrigem.RECEBIMENTO_AVULSO
    elif tipo == 'medicao':
        recebimento = get_object_or_404(
            RecebimentoMedicao.objects.filter(empresa=empresa).select_related('caixa'),
            pk=pk,
        )
        categoria_origem = MovimentoCaixa.CategoriaOrigem.RECEBIMENTO_MEDICAO
    else:
        messages.error(request, 'Tipo de recebimento inválido.')
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    if recebimento.status == recebimento.Status.PAGO:
        messages.info(request, 'Este recebimento já está liquidado.')
        return redirect_empresa(request, 'financeiro:movimentar_caixa')
    if not recebimento.caixa_id:
        messages.error(request, 'Informe um caixa antes de liquidar este recebimento.')
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    form = RecebimentoLiquidacaoForm(
        request.POST or None,
        recebimento=recebimento,
        empresa=empresa,
    )
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        recebimento.valor = cd['valor']
        recebimento.impostos = cd['impostos']
        recebimento.valor_liquido = cd['valor_liquido']
        recebimento.conta_bancaria = cd.get('conta_bancaria')
        with transaction.atomic():
            _criar_movimento_recebimento(
                recebimento,
                categoria_origem,
                data_liquidacao=cd['data_pagamento'],
            )

        messages.success(request, 'Recebimento liquidado e lançado no caixa.')
        if _is_htmx(request):
            response = HttpResponse(status=200)
            response['HX-Redirect'] = reverse_empresa(request, 'financeiro:movimentar_caixa')
            return response
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    if _is_htmx(request):
        return render(
            request,
            'financeiro/partials/recebimento_liquidar_modal.html',
            {
                'page_title': 'Liquidar recebimento',
                'form': form,
                'recebimento': recebimento,
                'tipo': tipo,
                'post_url': reverse_empresa(
                    request,
                    'financeiro:recebimento_liquidar',
                    kwargs={'tipo': tipo, 'pk': recebimento.pk},
                ),
            },
        )
    return redirect_empresa(request, 'financeiro:movimentar_caixa')


@login_required
def recebimento_editar(request, tipo: str, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    if tipo == 'avulso':
        recebimento = get_object_or_404(
            RecebimentoAvulso.objects.filter(empresa=empresa).select_related('caixa', 'movimento'),
            pk=pk,
        )
        form_class = RecebimentoAvulsoEditForm
    elif tipo == 'medicao':
        recebimento = get_object_or_404(
            RecebimentoMedicao.objects.filter(empresa=empresa).select_related(
                'caixa',
                'movimento',
                'obra',
            ),
            pk=pk,
        )
        form_class = RecebimentoMedicaoEditForm
    else:
        messages.error(request, 'Tipo de recebimento inválido.')
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    if request.method == 'POST' and request.POST.get('action') == 'excluir':
        with transaction.atomic():
            movimento = getattr(recebimento, 'movimento', None)
            recebimento.delete()
            if movimento:
                movimento.delete()
        messages.success(request, 'Recebimento excluído.')
        if _is_htmx(request):
            response = HttpResponse(status=200)
            response['HX-Redirect'] = reverse_empresa(request, 'financeiro:movimentar_caixa')
            return response
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    form = form_class(
        request.POST or None,
        request.FILES or None,
        empresa=empresa,
        instance=recebimento,
    )
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        with transaction.atomic():
            recebimento.caixa = cd['caixa']
            recebimento.cliente = cd['cliente']
            recebimento.categoria = cd.get('categoria')
            recebimento.data = cd['data']
            recebimento.data_pagamento = cd.get('data_pagamento')
            recebimento.valor = cd['valor']
            recebimento.impostos = cd['impostos']
            recebimento.valor_liquido = cd['valor_liquido']
            recebimento.descricao = cd['descricao']
            recebimento.observacao = cd.get('observacao') or ''
            if cd.get('comprovante'):
                recebimento.comprovante = cd['comprovante']
            if tipo == 'medicao':
                recebimento.obra = cd['obra']
                recebimento.medicao_numero = cd['medicao_numero']
                recebimento.nota_fiscal_numero = (cd.get('nota_fiscal_numero') or '').strip()
            recebimento.full_clean()
            recebimento.save()

            movimento = getattr(recebimento, 'movimento', None)
            if movimento:
                movimento.caixa = recebimento.caixa
                movimento.valor = recebimento.valor_liquido
                movimento.data = recebimento.data_pagamento or recebimento.data
                movimento.descricao = recebimento.descricao
                movimento.observacao = recebimento.observacao
                movimento.full_clean()
                movimento.save()

        messages.success(request, 'Recebimento atualizado.')
        if _is_htmx(request):
            response = HttpResponse(status=200)
            response['HX-Redirect'] = reverse_empresa(request, 'financeiro:movimentar_caixa')
            return response
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    if _is_htmx(request):
        return render(
            request,
            'financeiro/partials/recebimento_editar_modal.html',
            {
                'page_title': 'Editar recebimento',
                'form': form,
                'recebimento': recebimento,
                'tipo': tipo,
                'post_url': reverse_empresa(
                    request,
                    'financeiro:recebimento_editar',
                    kwargs={'tipo': tipo, 'pk': recebimento.pk},
                ),
            },
        )
    return redirect_empresa(request, 'financeiro:movimentar_caixa')


@login_required
def movimentar_pagamento(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    return render(
        request,
        'financeiro/movimentar_pagamento.html',
        {'page_title': 'Pagamentos'},
    )


def _periodo_pagamento_bancario_params(request):
    inicio_mes, fim_mes = _periodo_mensal_atual()
    inicio_raw = (request.GET.get('data_inicio') or '').strip()
    fim_raw = (request.GET.get('data_fim') or '').strip()
    inicio = parse_date(inicio_raw) or inicio_mes
    fim_inclusive = parse_date(fim_raw) or (fim_mes - timedelta(days=1))
    if fim_inclusive < inicio:
        fim_inclusive = inicio
    return {
        'data_inicio_raw': inicio.isoformat(),
        'data_fim_raw': fim_inclusive.isoformat(),
        'inicio': inicio,
        'fim_inclusive': fim_inclusive,
        'fim_exclusive': fim_inclusive + timedelta(days=1),
    }


def _periodo_pagamentos_avulsos_mensais(request):
    """Mês/ano para o card de pagamentos avulsos (padrão: mês corrente)."""
    hoje = timezone.localdate()
    mes_raw = (request.GET.get('av_mes') or '').strip()
    ano_raw = (request.GET.get('av_ano') or '').strip()
    try:
        mes = int(mes_raw) if mes_raw else hoje.month
    except (TypeError, ValueError):
        mes = hoje.month
    if mes < 1 or mes > 12:
        mes = hoje.month
    try:
        ano = int(ano_raw) if ano_raw else hoje.year
    except (TypeError, ValueError):
        ano = hoje.year
    if ano < 2000 or ano > hoje.year + 10:
        ano = hoje.year
    dia_ultimo = calendar.monthrange(ano, mes)[1]
    inicio = date(ano, mes, 1)
    fim_inclusive = date(ano, mes, dia_ultimo)
    return {
        'mes': mes,
        'ano': ano,
        'inicio': inicio,
        'fim_inclusive': fim_inclusive,
    }


def _contar_parcelas_previstas_sem_qtd(recorrencia: PagamentoBancarioRecorrente) -> int:
    """Quantidade prevista quando só há data_fim (sem qtd_parcelas)."""
    dia = recorrencia.dia_pagamento
    ini = recorrencia.data_inicio
    limite = recorrencia.data_fim
    if not limite:
        return 0
    resultado = 0
    for indice in range(2400):
        vencimento = _data_pagamento_bancario(ini, dia, indice)
        if vencimento < ini:
            continue
        if vencimento > limite:
            break
        resultado += 1
    return resultado


def _parcelas_previstas_count(recorrencia: PagamentoBancarioRecorrente) -> int | None:
    """Número de parcelas quando o contrato tem término definido (qtd ou data fim)."""
    if recorrencia.qtd_parcelas:
        return int(recorrencia.qtd_parcelas)
    if recorrencia.data_fim:
        return _contar_parcelas_previstas_sem_qtd(recorrencia)
    return None


def _valor_total_exibicao_recorrencia(
    recorrencia: PagamentoBancarioRecorrente, soma_todas_parcelas_db: Decimal
) -> Decimal:
    parcelas_prev = _parcelas_previstas_count(recorrencia)
    if parcelas_prev:
        vp = recorrencia.valor_parcela or Decimal('0')
        return vp * Decimal(parcelas_prev)
    base = soma_todas_parcelas_db or Decimal('0')
    return base


def _garantir_parcelas_bancarias(empresa, inicio: date, fim_exclusive: date):
    recorrencias = (
        PagamentoBancarioRecorrente.objects.filter(
            empresa=empresa,
            ativo=True,
            data_inicio__lt=fim_exclusive,
        )
        .select_related('caixa', 'conta_bancaria', 'categoria')
        .order_by('data_inicio', 'pk')
    )
    for recorrencia in recorrencias:
        if recorrencia.data_fim and recorrencia.data_fim < inicio:
            continue
        inicio_meses = max(
            0,
            _meses_entre(recorrencia.data_inicio.replace(day=1), inicio.replace(day=1)),
        )
        fim_meses = max(
            inicio_meses,
            _meses_entre(recorrencia.data_inicio.replace(day=1), (fim_exclusive - timedelta(days=1)).replace(day=1)),
        )
        if recorrencia.qtd_parcelas:
            fim_meses = min(fim_meses, recorrencia.qtd_parcelas - 1)
        for indice in range(inicio_meses, fim_meses + 1):
            vencimento = _data_pagamento_bancario(
                recorrencia.data_inicio,
                recorrencia.dia_pagamento,
                indice,
            )
            if vencimento < recorrencia.data_inicio:
                continue
            if recorrencia.data_fim and vencimento > recorrencia.data_fim:
                continue
            if not (inicio <= vencimento < fim_exclusive):
                continue
            PagamentoBancarioParcela.objects.get_or_create(
                recorrencia=recorrencia,
                numero_parcela=indice + 1,
                defaults={
                    'data_vencimento': vencimento,
                    'valor': recorrencia.valor_parcela,
                    'conta_bancaria': recorrencia.conta_bancaria,
                },
            )


def _gerar_parcelas_iniciais_bancarias(recorrencia):
    if recorrencia.qtd_parcelas:
        fim = _data_pagamento_bancario(
            recorrencia.data_inicio,
            recorrencia.dia_pagamento,
            recorrencia.qtd_parcelas - 1,
        ) + timedelta(days=1)
    elif recorrencia.data_fim:
        fim = recorrencia.data_fim + timedelta(days=1)
    else:
        fim = _somar_meses(timezone.localdate(), 12)
    _garantir_parcelas_bancarias(recorrencia.empresa, recorrencia.data_inicio, fim)


@login_required
def pagamento_bancario_lista(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    periodo = _periodo_pagamento_bancario_params(request)
    av_periodo = _periodo_pagamentos_avulsos_mensais(request)
    filtro_conta_id = _int_param(request, 'conta_bancaria')
    filtro_categoria_id = _int_param(request, 'categoria')
    av_conta_id = _int_param(request, 'av_conta_bancaria')
    av_categoria_id = _int_param(request, 'av_categoria')
    _garantir_parcelas_bancarias(empresa, periodo['inicio'], periodo['fim_exclusive'])
    parcelas_abertas = (
        PagamentoBancarioParcela.objects.filter(
            recorrencia__empresa=empresa,
            status=PagamentoBancarioParcela.Status.ABERTO,
            data_vencimento__gte=periodo['inicio'],
            data_vencimento__lt=periodo['fim_exclusive'],
        )
        .select_related('recorrencia', 'recorrencia__caixa', 'recorrencia__conta_bancaria', 'recorrencia__categoria')
        .order_by('data_vencimento', 'pk')
    )
    recorrencias_qs = (
        PagamentoBancarioRecorrente.objects.filter(empresa=empresa)
        .select_related('caixa', 'conta_bancaria', 'categoria')
    )
    if filtro_conta_id:
        recorrencias_qs = recorrencias_qs.filter(conta_bancaria_id=filtro_conta_id)
    if filtro_categoria_id:
        recorrencias_qs = recorrencias_qs.filter(categoria_id=filtro_categoria_id)
    decimal_out = DecimalField(max_digits=16, decimal_places=2)
    recorrencias_qs = recorrencias_qs.annotate(
        valor_pago_parcelas=Coalesce(
            Sum(
                'parcelas__valor',
                filter=Q(parcelas__status=PagamentoBancarioParcela.Status.PAGO),
            ),
            Value(Decimal('0')),
            output_field=decimal_out,
        ),
        soma_valor_parcelas=Coalesce(
            Sum('parcelas__valor'),
            Value(Decimal('0')),
            output_field=decimal_out,
        ),
    ).order_by('-ativo', 'descricao')

    recorrencias_lista = list(recorrencias_qs)
    ids_rec = [r.pk for r in recorrencias_lista]
    qtd_pagas_por_rec: dict[int, int] = {rid: 0 for rid in ids_rec}
    if ids_rec:
        for rid, n in (
            PagamentoBancarioParcela.objects.filter(
                recorrencia_id__in=ids_rec,
                status=PagamentoBancarioParcela.Status.PAGO,
            )
            .values('recorrencia_id')
            .annotate(n=Count('pk'))
            .values_list('recorrencia_id', 'n')
        ):
            qtd_pagas_por_rec[int(rid)] = int(n)

    for indice_linha, rec in enumerate(recorrencias_lista, start=1):
        setattr(rec, 'linha_indice_recorrencia', indice_linha)
        setattr(rec, 'qtd_parcelas_pagas', qtd_pagas_por_rec.get(rec.pk, 0))
        soma_parcelas_db = getattr(rec, 'soma_valor_parcelas', Decimal('0'))
        setattr(rec, 'valor_total_calculado', _valor_total_exibicao_recorrencia(rec, soma_parcelas_db))

    contas_banco_filtro = ContaBancaria.objects.filter(empresa=empresa, ativo=True).order_by('banco', 'nome')
    categorias_filtro = CategoriaFinanceira.objects.filter(
        empresa=empresa,
        movimentacao_tipo=CategoriaFinanceira.MovimentacaoTipo.PAGAMENTO_BANCARIO,
    ).order_by('nome')
    pagamentos_avulsos_qs = PagamentoBancarioAvulso.objects.filter(
        empresa=empresa,
        data_pagamento__gte=av_periodo['inicio'],
        data_pagamento__lte=av_periodo['fim_inclusive'],
    )
    if av_conta_id:
        pagamentos_avulsos_qs = pagamentos_avulsos_qs.filter(conta_bancaria_id=av_conta_id)
    if av_categoria_id:
        pagamentos_avulsos_qs = pagamentos_avulsos_qs.filter(categoria_id=av_categoria_id)
    pagamentos_avulsos_periodo = (
        pagamentos_avulsos_qs.select_related('caixa', 'conta_bancaria', 'categoria')
        .order_by('-data_pagamento', '-pk')
    )
    return render(
        request,
        'financeiro/pagamento_bancario_lista.html',
        {
            'page_title': 'Pagamentos Bancários',
            'periodo': periodo,
            'parcelas_abertas': parcelas_abertas,
            'pagamentos_avulsos_periodo': pagamentos_avulsos_periodo,
            'av_periodo': av_periodo,
            'filtro_av_conta_pk': av_conta_id,
            'filtro_av_categoria_pk': av_categoria_id,
            'meses_filtro_caixa': MESES_FILTRO_CAIXA,
            'recorrencias': recorrencias_lista,
            'contas_banco_filtro': contas_banco_filtro,
            'categorias_filtro': categorias_filtro,
            'filtro_rec_conta_pk': filtro_conta_id,
            'filtro_rec_categoria_pk': filtro_categoria_id,
        },
    )


@login_required
def pagamento_bancario_avulso_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    form = PagamentoBancarioAvulsoForm(request.POST or None, empresa=empresa)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            form.save()
        messages.success(request, 'Pagamento bancário avulso registrado.')
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        return response
    return render(
        request,
        'financeiro/partials/pagamento_bancario_avulso_modal.html',
        {
            'form': form,
            'post_url': reverse_empresa(request, 'financeiro:pagamento_bancario_avulso_novo'),
            'modal_title': 'Novo pagamento bancário avulso',
            'modal_subtitle': 'Registre um pagamento já liquidado fora da recorrência.',
        },
    )


@login_required
def pagamento_bancario_avulso_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    pagamento = get_object_or_404(
        PagamentoBancarioAvulso.objects.filter(empresa=empresa),
        pk=pk,
    )
    if request.method == 'POST' and request.POST.get('action') == 'excluir':
        with transaction.atomic():
            pagamento.delete()
        messages.success(request, 'Pagamento bancário avulso excluído.')
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        return response
    form = PagamentoBancarioAvulsoForm(request.POST or None, empresa=empresa, instance=pagamento)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            form.save()
        messages.success(request, 'Pagamento bancário avulso atualizado.')
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        return response
    return render(
        request,
        'financeiro/partials/pagamento_bancario_avulso_modal.html',
        {
            'form': form,
            'post_url': reverse_empresa(
                request,
                'financeiro:pagamento_bancario_avulso_editar',
                kwargs={'pk': pagamento.pk},
            ),
            'modal_title': 'Editar pagamento bancário avulso',
            'modal_subtitle': 'Altere os dados ou exclua o registro.',
            'pagamento_avulso': pagamento,
        },
    )


@login_required
def pagamento_bancario_detalhe(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    recorrencia = get_object_or_404(
        PagamentoBancarioRecorrente.objects.filter(empresa=empresa)
        .select_related('caixa', 'conta_bancaria', 'categoria'),
        pk=pk,
    )
    if recorrencia.qtd_parcelas:
        fim = _data_pagamento_bancario(
            recorrencia.data_inicio,
            recorrencia.dia_pagamento,
            recorrencia.qtd_parcelas - 1,
        ) + timedelta(days=1)
    elif recorrencia.data_fim:
        fim = recorrencia.data_fim + timedelta(days=1)
    else:
        fim = _somar_meses(timezone.localdate(), 12)
    _garantir_parcelas_bancarias(empresa, recorrencia.data_inicio, fim)
    parcelas = (
        recorrencia.parcelas.select_related('conta_bancaria')
        .order_by('data_vencimento', 'numero_parcela', 'pk')
    )
    parcelas_lista = list(parcelas)
    total_geral = sum((parcela.valor for parcela in parcelas_lista), Decimal('0'))
    total_pago = sum(
        (
            parcela.valor
            for parcela in parcelas_lista
            if parcela.status == PagamentoBancarioParcela.Status.PAGO
        ),
        Decimal('0'),
    )
    total_aberto = sum(
        (
            parcela.valor
            for parcela in parcelas_lista
            if parcela.status == PagamentoBancarioParcela.Status.ABERTO
        ),
        Decimal('0'),
    )
    return render(
        request,
        'financeiro/pagamento_bancario_detalhe.html',
        {
            'page_title': recorrencia.descricao,
            'recorrencia': recorrencia,
            'parcelas': parcelas_lista,
            'total_geral': total_geral,
            'total_pago': total_pago,
            'total_aberto': total_aberto,
        },
    )


@login_required
def pagamento_bancario_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    form = PagamentoBancarioRecorrenteForm(request.POST or None, empresa=empresa)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            recorrencia = form.save()
            _gerar_parcelas_iniciais_bancarias(recorrencia)
        messages.success(request, 'Pagamento bancário recorrente cadastrado.')
        return redirect_empresa(request, 'financeiro:pagamento_bancario_lista')
    return render(
        request,
        'financeiro/pagamento_bancario_form.html',
        {
            'page_title': 'Novo pagamento bancário',
            'form': form,
            'modo': 'novo',
        },
    )


@login_required
def pagamento_bancario_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    recorrencia = get_object_or_404(
        PagamentoBancarioRecorrente.objects.filter(empresa=empresa),
        pk=pk,
    )
    form = PagamentoBancarioRecorrenteForm(
        request.POST or None,
        empresa=empresa,
        instance=recorrencia,
    )
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            recorrencia = form.save()
            _gerar_parcelas_iniciais_bancarias(recorrencia)
        messages.success(request, 'Pagamento bancário atualizado.')
        return redirect_empresa(request, 'financeiro:pagamento_bancario_lista')
    return render(
        request,
        'financeiro/pagamento_bancario_form.html',
        {
            'page_title': 'Editar pagamento bancário',
            'form': form,
            'recorrencia': recorrencia,
            'modo': 'editar',
        },
    )


@login_required
def pagamento_bancario_excluir(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    recorrencia = get_object_or_404(
        PagamentoBancarioRecorrente.objects.filter(empresa=empresa),
        pk=pk,
    )
    if request.method == 'POST':
        if recorrencia.parcelas.filter(status=PagamentoBancarioParcela.Status.PAGO).exists():
            recorrencia.ativo = False
            recorrencia.save(update_fields=['ativo', 'atualizado_em'])
            messages.success(request, 'Pagamento bancário inativado; parcelas pagas foram preservadas.')
        else:
            recorrencia.delete()
            messages.success(request, 'Pagamento bancário excluído.')
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        return response
    return render(
        request,
        'financeiro/partials/pagamento_bancario_excluir_modal.html',
        {
            'recorrencia': recorrencia,
            'post_url': reverse_empresa(
                request,
                'financeiro:pagamento_bancario_excluir',
                kwargs={'pk': recorrencia.pk},
            ),
        },
    )


@login_required
def pagamento_bancario_pagar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    parcela = get_object_or_404(
        PagamentoBancarioParcela.objects.filter(recorrencia__empresa=empresa)
        .select_related('recorrencia', 'recorrencia__conta_bancaria'),
        pk=pk,
    )
    form = PagamentoBancarioPagarForm(
        request.POST or None,
        empresa=empresa,
        parcela=parcela,
    )
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        ja_pago = parcela.status == PagamentoBancarioParcela.Status.PAGO
        parcela.status = PagamentoBancarioParcela.Status.PAGO
        parcela.data_pagamento = cd['data_pagamento']
        parcela.valor = cd['valor']
        parcela.conta_bancaria = cd['conta_bancaria']
        parcela.observacao = cd.get('observacao') or ''
        parcela.full_clean()
        parcela.save()
        messages.success(
            request,
            'Pagamento da parcela atualizado.' if ja_pago else 'Pagamento bancário realizado.',
        )
        response = HttpResponse(status=200)
        next_view = (request.POST.get('next') or request.GET.get('next') or '').strip()
        if next_view == 'detalhe':
            redirect_url = reverse_empresa(
                request,
                'financeiro:pagamento_bancario_detalhe',
                kwargs={'pk': parcela.recorrencia_id},
            )
        else:
            redirect_url = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        response['HX-Redirect'] = redirect_url
        return response
    editando_parcela = parcela.status == PagamentoBancarioParcela.Status.PAGO
    return render(
        request,
        'financeiro/partials/pagamento_bancario_pagar_modal.html',
        {
            'parcela': parcela,
            'form': form,
            'editando_parcela': editando_parcela,
            'next_view': (request.GET.get('next') or '').strip(),
            'post_url': reverse_empresa(
                request,
                'financeiro:pagamento_bancario_pagar',
                kwargs={'pk': parcela.pk},
            ),
        },
    )


@login_required
def pagamento_bancario_parcela_estornar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    parcela = get_object_or_404(
        PagamentoBancarioParcela.objects.filter(
            recorrencia__empresa=empresa,
            status=PagamentoBancarioParcela.Status.PAGO,
        ).select_related('recorrencia'),
        pk=pk,
    )
    if request.method == 'POST':
        with transaction.atomic():
            parcela.status = PagamentoBancarioParcela.Status.ABERTO
            parcela.data_pagamento = None
            parcela.conta_bancaria = None
            parcela.observacao = ''
            parcela.valor = parcela.recorrencia.valor_parcela
            parcela.full_clean()
            parcela.save()
        messages.success(request, 'Pagamento da parcela excluído; a parcela voltou para em aberto.')
        response = HttpResponse(status=200)
        next_view = (request.POST.get('next') or '').strip()
        if next_view == 'detalhe':
            redirect_url = reverse_empresa(
                request,
                'financeiro:pagamento_bancario_detalhe',
                kwargs={'pk': parcela.recorrencia_id},
            )
        else:
            redirect_url = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        response['HX-Redirect'] = redirect_url
        return response
    return render(
        request,
        'financeiro/partials/pagamento_bancario_parcela_estornar_modal.html',
        {
            'parcela': parcela,
            'post_url': reverse_empresa(
                request,
                'financeiro:pagamento_bancario_parcela_estornar',
                kwargs={'pk': parcela.pk},
            ),
            'next_view': (request.GET.get('next') or '').strip(),
        },
    )


@login_required
def pagamento_bancario_parcela_valor(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    parcela = get_object_or_404(
        PagamentoBancarioParcela.objects.filter(
            recorrencia__empresa=empresa,
            status=PagamentoBancarioParcela.Status.ABERTO,
        ).select_related('recorrencia'),
        pk=pk,
    )
    form = PagamentoBancarioParcelaValorForm(request.POST or None, parcela=parcela)
    if request.method == 'POST' and form.is_valid():
        parcela.valor = form.cleaned_data['valor']
        parcela.full_clean()
        parcela.save()
        messages.success(request, 'Valor da parcela atualizado.')
        response = HttpResponse(status=200)
        next_view = (request.POST.get('next') or '').strip()
        if next_view == 'detalhe':
            redirect_url = reverse_empresa(
                request,
                'financeiro:pagamento_bancario_detalhe',
                kwargs={'pk': parcela.recorrencia_id},
            )
        else:
            redirect_url = reverse_empresa(request, 'financeiro:pagamento_bancario_lista')
        response['HX-Redirect'] = redirect_url
        return response
    return render(
        request,
        'financeiro/partials/pagamento_bancario_parcela_valor_modal.html',
        {
            'parcela': parcela,
            'form': form,
            'next_view': (request.GET.get('next') or '').strip(),
            'post_url': reverse_empresa(
                request,
                'financeiro:pagamento_bancario_parcela_valor',
                kwargs={'pk': parcela.pk},
            ),
        },
    )


@login_required
def recebimento_avulso_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = RecebimentoAvulsoForm(
        request.POST or None,
        request.FILES or None,
        empresa=empresa,
    )
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        with transaction.atomic():
            extra = {
                'empresa': empresa,
                'caixa': cd['caixa'],
                'cliente': cd['cliente'],
                'categoria': cd.get('categoria'),
                'status': RecebimentoAvulso.Status.ABERTO,
                'data': cd['data'],
                'valor': cd['valor'],
                'impostos': cd['impostos'],
                'valor_liquido': cd['valor_liquido'],
                'descricao': cd['descricao'],
                'observacao': cd.get('observacao') or '',
            }
            if cd.get('comprovante'):
                extra['comprovante'] = cd['comprovante']
            RecebimentoAvulso.objects.create(**extra)
        messages.success(request, 'Recebimento avulso registrado em aberto.')
        if _is_htmx(request):
            response = HttpResponse(status=200)
            response['HX-Redirect'] = reverse_empresa(request, 'financeiro:movimentar_caixa')
            return response
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    if _is_htmx(request):
        return render(
            request,
            'financeiro/partials/recebimento_form_modal.html',
            {
                'page_title': 'Recebimento avulso',
                'modal_title': 'Lançar recebimento avulso',
                'modal_subtitle': 'O recebimento será registrado em aberto até a liquidação.',
                'form': form,
                'post_url': reverse_empresa(request, 'financeiro:recebimento_avulso_novo'),
            },
        )

    return render(
        request,
        'financeiro/recebimento_avulso_form.html',
        {
            'page_title': 'Recebimento avulso',
            'form': form,
            'post_url': reverse_empresa(request, 'financeiro:recebimento_avulso_novo'),
        },
    )


@login_required
def recebimento_medicao_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = RecebimentoMedicaoForm(
        request.POST or None,
        request.FILES or None,
        empresa=empresa,
    )
    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        desc_extra = f' — Obra: {cd["obra"].nome} — Medição nº {cd["medicao_numero"]}'
        base = cd['descricao']
        max_base = 500 - len(desc_extra)
        if max_base >= 1:
            descricao = (base[:max_base].rstrip() + desc_extra)[:500]
        else:
            descricao = desc_extra[:500]
        with transaction.atomic():
            extra_m = {
                'empresa': empresa,
                'caixa': cd['caixa'],
                'cliente': cd['cliente'],
                'categoria': cd.get('categoria'),
                'obra': cd['obra'],
                'status': RecebimentoMedicao.Status.ABERTO,
                'data': cd['data'],
                'valor': cd['valor'],
                'impostos': cd['impostos'],
                'valor_liquido': cd['valor_liquido'],
                'descricao': descricao[:500],
                'observacao': cd.get('observacao') or '',
                'medicao_numero': cd['medicao_numero'],
                'nota_fiscal_numero': (cd.get('nota_fiscal_numero') or '').strip(),
            }
            if cd.get('comprovante'):
                extra_m['comprovante'] = cd['comprovante']
            RecebimentoMedicao.objects.create(**extra_m)
        messages.success(request, 'Recebimento por medição registrado em aberto.')
        if _is_htmx(request):
            response = HttpResponse(status=200)
            response['HX-Redirect'] = reverse_empresa(request, 'financeiro:movimentar_caixa')
            return response
        return redirect_empresa(request, 'financeiro:movimentar_caixa')

    if _is_htmx(request):
        return render(
            request,
            'financeiro/partials/recebimento_form_modal.html',
            {
                'page_title': 'Recebimento por medição',
                'modal_title': 'Lançar recebimento de medição',
                'modal_subtitle': 'O recebimento será registrado em aberto até a liquidação.',
                'form': form,
                'post_url': reverse_empresa(request, 'financeiro:recebimento_medicao_novo'),
            },
        )

    return render(
        request,
        'financeiro/recebimento_medicao_form.html',
        {
            'page_title': 'Recebimento por medição',
            'form': form,
            'post_url': reverse_empresa(request, 'financeiro:recebimento_medicao_novo'),
        },
    )


@login_required
def categoria_lista(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    categorias = list(
        CategoriaFinanceira.objects.filter(empresa=empresa).order_by('movimentacao_tipo', 'nome')
    )
    return render(
        request,
        'financeiro/categoria_lista.html',
        {
            'page_title': 'Categorias',
            'categorias': categorias,
        },
    )


@login_required
def conta_bancaria_lista(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    contas = list(ContaBancaria.objects.filter(empresa=empresa).order_by('-ativo', 'banco', 'nome'))
    return render(
        request,
        'financeiro/conta_bancaria_lista.html',
        {
            'page_title': 'Contas Bancárias',
            'contas': contas,
            'total_contas': len(contas),
            'total_ativas': sum(1 for conta in contas if conta.ativo),
            'total_inativas': sum(1 for conta in contas if not conta.ativo),
        },
    )


def _conta_bancaria_modal_response(request, template, context):
    return render(request, template, context)


def _conta_bancaria_redirect_response(request):
    url = reverse_empresa(request, 'financeiro:conta_bancaria_lista')
    if _is_htmx(request):
        response = HttpResponse(status=204)
        response['HX-Redirect'] = url
        return response
    return redirect(url)


@login_required
def conta_bancaria_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = ContaBancariaForm(request.POST or None, empresa=empresa)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Conta bancária criada.')
        return _conta_bancaria_redirect_response(request)

    return _conta_bancaria_modal_response(
        request,
        'financeiro/partials/conta_bancaria_form_modal.html',
        {
            'form': form,
            'modal_title': 'Nova conta bancária',
            'modal_subtitle': 'Cadastre uma conta bancária da empresa.',
            'post_url': reverse_empresa(request, 'financeiro:conta_bancaria_novo'),
            'submit_label': 'Salvar conta',
        },
    )


@login_required
def conta_bancaria_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    conta = get_object_or_404(ContaBancaria.objects.filter(empresa=empresa), pk=pk)
    form = ContaBancariaForm(request.POST or None, instance=conta, empresa=empresa)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Conta bancária atualizada.')
        return _conta_bancaria_redirect_response(request)

    return _conta_bancaria_modal_response(
        request,
        'financeiro/partials/conta_bancaria_form_modal.html',
        {
            'form': form,
            'conta': conta,
            'modal_title': f'Editar — {conta.nome}',
            'modal_subtitle': 'Atualize os dados bancários cadastrados.',
            'post_url': reverse_empresa(
                request,
                'financeiro:conta_bancaria_editar',
                kwargs={'pk': conta.pk},
            ),
            'submit_label': 'Salvar alterações',
        },
    )


@login_required
def conta_bancaria_excluir(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    conta = get_object_or_404(ContaBancaria.objects.filter(empresa=empresa), pk=pk)
    if request.method == 'POST':
        conta.delete()
        messages.success(request, 'Conta bancária excluída.')
        return _conta_bancaria_redirect_response(request)

    return _conta_bancaria_modal_response(
        request,
        'financeiro/partials/conta_bancaria_excluir_modal.html',
        {'conta': conta},
    )


@login_required
def categoria_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = CategoriaFinanceiraForm(
        request.POST or None,
        empresa=empresa,
    )
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Categoria criada.')
        return redirect_empresa(request, 'financeiro:categoria_lista')

    return render(
        request,
        'financeiro/categoria_form.html',
        {
            'page_title': 'Nova categoria',
            'form': form,
            'modo': 'novo',
        },
    )


@login_required
def categoria_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    categoria = get_object_or_404(CategoriaFinanceira.objects.filter(empresa=empresa), pk=pk)
    form = CategoriaFinanceiraForm(
        request.POST or None,
        instance=categoria,
        empresa=empresa,
    )
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Categoria atualizada.')
        return redirect_empresa(request, 'financeiro:categoria_lista')

    return render(
        request,
        'financeiro/categoria_form.html',
        {
            'page_title': f'Editar — {categoria.nome}',
            'form': form,
            'modo': 'editar',
            'categoria': categoria,
        },
    )


@login_required
def categoria_inativar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    if request.method != 'POST':
        return redirect_empresa(request, 'financeiro:dashboard')

    categoria = get_object_or_404(CategoriaFinanceira.objects.filter(empresa=empresa), pk=pk)
    categoria.ativo = False
    categoria.save()
    messages.success(request, 'Categoria inativada.')
    return redirect_empresa(request, 'financeiro:categoria_lista')


@login_required
def categoria_reativar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    if request.method != 'POST':
        return redirect_empresa(request, 'financeiro:dashboard')

    categoria = get_object_or_404(CategoriaFinanceira.objects.filter(empresa=empresa), pk=pk)
    categoria.ativo = True
    categoria.save()
    messages.success(request, 'Categoria reativada.')
    return redirect_empresa(request, 'financeiro:categoria_lista')


@login_required
def pagamento_nf_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = PagamentoNotaFiscalForm(request.POST or None, empresa=empresa)
    itens_fs = PagamentoNotaFiscalItemFormSet(
        request.POST or None,
        prefix='itens',
        form_kwargs={
            'empresa': empresa,
            'default_caixa_id': form.initial.get('caixa'),
        },
    )
    pagamentos_fs = PagamentoNotaFiscalPagamentoFormSet(
        request.POST or None,
        prefix='pagamentos',
        form_kwargs={'default_data': timezone.localdate(), 'empresa': empresa},
    )
    boletos_fs = BoletoRascunhoFormSet(request.POST or None, prefix='boletos')
    dup_nf = None
    force_save = str(request.POST.get('force_save', '')).strip() == '1'

    if request.method == 'POST':
        form_ok = form.is_valid()
        itens_ok = itens_fs.is_valid()
        pagamentos_ok = pagamentos_fs.is_valid()
        boletos_ok = boletos_fs.is_valid()
        ok = form_ok and itens_ok and pagamentos_ok and boletos_ok

        if ok:
            cd = form.cleaned_data
            dup_nf = None
            if nf_numero_exige_unicidade_para_fornecedor(cd.get('numero_nf') or ''):
                dup_nf = (
                    PagamentoNotaFiscal.objects.filter(
                        empresa=empresa,
                        fornecedor=cd['fornecedor'],
                        numero_nf=cd['numero_nf'],
                    )
                    .order_by('-pk')
                    .first()
                )
            if dup_nf and not force_save:
                ok = False

        if ok:
            with transaction.atomic():
                nf = form.save()

                itens_objs = []
                for f in itens_fs:
                    if not getattr(f, 'cleaned_data', None):
                        continue
                    if f.cleaned_data.get('DELETE'):
                        continue
                    if f.cleaned_data.get('_skip_form'):
                        continue
                    item = f.save(commit=False)
                    item.pagamento_nf = nf
                    # Caixa já veio no form; se vazio por algum motivo, default da NF.
                    if not item.caixa_id:
                        item.caixa = nf.caixa
                    item.full_clean()
                    itens_objs.append(item)
                if not itens_objs:
                    raise ValidationError('Informe pelo menos 1 item da NF.')
                PagamentoNotaFiscalItem.objects.bulk_create(itens_objs)

                _salvar_pagamentos_nf(nf, pagamentos_fs, boletos_fs, nf.total_itens())

            messages.success(request, 'Pagamento por NF registrado.')
            _audit_nf(
                request,
                nf,
                acao='create',
                resumo=f'Criou Nota Fiscal {nf.numero_nf}',
                alteracoes=[
                    f'Fornecedor: {nf.fornecedor.nome}',
                    f'Valor total: {_decimal_audit(nf.total_itens())}',
                ],
            )
            return redirect_empresa(
                request,
                'financeiro:pagamento_nf_detalhe',
                kwargs={'pk': nf.pk},
            )

    return render(
        request,
        'financeiro/pagamento_nf_form.html',
        {
            'page_title': 'Nota Fiscal',
            'form': form,
            'itens_fs': itens_fs,
            'pagamentos_fs': pagamentos_fs,
            'boletos_fs': boletos_fs,
            'modo': 'novo',
            'dup_nf': dup_nf,
            'show_dup_modal': bool(dup_nf) and request.method == 'POST' and not force_save,
            'has_validation_errors': _pagamento_nf_tem_erros_validacao(
                form,
                itens_fs,
                pagamentos_fs,
                boletos_fs,
            ),
            'active_tab': _active_tab_pagamento_nf(
                request,
                form,
                itens_fs,
                pagamentos_fs,
                boletos_fs,
            ),
            'pagamento_tipo_boletos': PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS,
            'contas_bancarias_pagamento': ContaBancaria.objects.filter(
                empresa=empresa,
                ativo=True,
            ).order_by('banco', 'nome'),
        },
    )


def _caixa_padrao_funcionario(empresa, funcionario) -> Caixa | None:
    if funcionario and funcionario.lotacao_id:
        caixa = Caixa.objects.filter(
            empresa=empresa,
            ativo=True,
            nome__iexact=funcionario.lotacao.nome,
        ).first()
        if caixa:
            return caixa
    return Caixa.objects.filter(empresa=empresa, ativo=True, tipo=Caixa.Tipo.GERAL).first()


@login_required
def pagamento_pessoal_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = PagamentoPessoalForm(request.POST or None, empresa=empresa)
    itens_fs = PagamentoPessoalItemFormSet(
        request.POST or None,
        prefix='itens',
        form_kwargs={'empresa': empresa},
    )
    funcionarios_caixa = {
        str(funcionario.pk): caixa.pk
        for funcionario in Funcionario.objects.filter(empresa=empresa)
        .select_related('lotacao')
        .order_by('nome')
        for caixa in [_caixa_padrao_funcionario(empresa, funcionario)]
        if caixa
    }
    funcionarios_lotacao = {
        str(funcionario.pk): funcionario.lotacao.nome
        for funcionario in Funcionario.objects.filter(empresa=empresa)
        .select_related('lotacao')
        .order_by('nome')
        if funcionario.lotacao_id
    }

    if request.method == 'POST':
        form = PagamentoPessoalForm(request.POST, empresa=empresa)

        if form.is_valid() and itens_fs.is_valid():
            with transaction.atomic():
                pagamento = form.save(commit=False)
                itens_forms = []
                data_pagamento = None
                for f in itens_fs:
                    if itens_forms:
                        break
                    if not getattr(f, 'cleaned_data', None):
                        continue
                    if f.cleaned_data.get('DELETE') or f.cleaned_data.get('_skip_form'):
                        continue
                    if data_pagamento is None:
                        data_pagamento = f.cleaned_data.get('data_pagamento')
                    itens_forms.append(f)
                if not itens_forms:
                    raise ValidationError('Informe a descrição do pagamento.')
                pagamento.data_pagamento = data_pagamento or timezone.localdate()
                pagamento.full_clean()
                pagamento.save()

                itens_objs = []
                for f in itens_forms:
                    item = f.save(commit=False)
                    item.pagamento = pagamento
                    item.full_clean()
                    itens_objs.append(item)
                PagamentoPessoalItem.objects.bulk_create(itens_objs)
            messages.success(request, 'Pagamento pessoal registrado.')
            return redirect_empresa(
                request,
                'financeiro:pagamento_pessoal_detalhe',
                kwargs={'pk': pagamento.pk},
            )

    return render(
        request,
        'financeiro/pagamento_pessoal_form.html',
        {
            'page_title': 'Pagamento Pessoal',
            'form': form,
            'itens_fs': itens_fs,
            'funcionarios_caixa': funcionarios_caixa,
            'funcionarios_lotacao': funcionarios_lotacao,
        },
    )


@login_required
def pagamento_pessoal_detalhe(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    pagamento = get_object_or_404(
        PagamentoPessoal.objects.filter(empresa=empresa).select_related(
            'funcionario',
            'caixa',
        ),
        pk=pk,
    )
    itens = list(pagamento.itens.select_related('categoria').order_by('pk'))
    total_itens = pagamento.total_itens()
    return render(
        request,
        'financeiro/pagamento_pessoal_detalhe.html',
        {
            'page_title': 'Pagamento Pessoal',
            'pagamento': pagamento,
            'itens': itens,
            'total_itens': total_itens,
        },
    )


@login_required
def pagamento_pessoal_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    pagamento = get_object_or_404(
        PagamentoPessoal.objects.filter(empresa=empresa).select_related('funcionario', 'caixa'),
        pk=pk,
    )
    item = pagamento.itens.order_by('pk').first()
    initial_item = []
    if item:
        initial_item.append(
            {
                'descricao': item.descricao,
                'categoria': item.categoria_id,
                'data_pagamento': pagamento.data_pagamento.isoformat()
                if pagamento.data_pagamento
                else '',
                'valor_total': format_decimal_br_moeda(item.valor_total),
            }
        )

    form = PagamentoPessoalForm(request.POST or None, instance=pagamento, empresa=empresa)
    itens_fs = PagamentoPessoalItemFormSet(
        request.POST or None,
        prefix='itens',
        form_kwargs={'empresa': empresa},
        initial=initial_item,
    )
    funcionarios_caixa = {
        str(funcionario.pk): caixa.pk
        for funcionario in Funcionario.objects.filter(empresa=empresa)
        .select_related('lotacao')
        .order_by('nome')
        for caixa in [_caixa_padrao_funcionario(empresa, funcionario)]
        if caixa
    }
    funcionarios_lotacao = {
        str(funcionario.pk): funcionario.lotacao.nome
        for funcionario in Funcionario.objects.filter(empresa=empresa)
        .select_related('lotacao')
        .order_by('nome')
        if funcionario.lotacao_id
    }

    if request.method == 'POST':
        form = PagamentoPessoalForm(request.POST, instance=pagamento, empresa=empresa)

        if form.is_valid() and itens_fs.is_valid():
            with transaction.atomic():
                pagamento = form.save(commit=False)
                item_form = None
                for f in itens_fs:
                    if not getattr(f, 'cleaned_data', None):
                        continue
                    if f.cleaned_data.get('DELETE') or f.cleaned_data.get('_skip_form'):
                        continue
                    item_form = f
                    break
                if item_form is None:
                    raise ValidationError('Informe a descrição do pagamento.')
                pagamento.data_pagamento = (
                    item_form.cleaned_data.get('data_pagamento') or timezone.localdate()
                )
                pagamento.full_clean()
                pagamento.save()
                pagamento.itens.all().delete()
                novo_item = item_form.save(commit=False)
                novo_item.pagamento = pagamento
                novo_item.full_clean()
                novo_item.save()
            messages.success(request, 'Pagamento pessoal atualizado.')
            return redirect_empresa(
                request,
                'financeiro:pagamento_pessoal_detalhe',
                kwargs={'pk': pagamento.pk},
            )

    return render(
        request,
        'financeiro/pagamento_pessoal_form.html',
        {
            'page_title': 'Editar Pagamento Pessoal',
            'form': form,
            'itens_fs': itens_fs,
            'funcionarios_caixa': funcionarios_caixa,
            'funcionarios_lotacao': funcionarios_lotacao,
            'pagamento': pagamento,
            'modo': 'editar',
        },
    )


@login_required
def pagamento_pessoal_excluir(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    pagamento = get_object_or_404(
        PagamentoPessoal.objects.filter(empresa=empresa).select_related('funcionario'),
        pk=pk,
    )
    if request.method != 'POST':
        if not _is_htmx(request):
            return redirect_empresa(
                request,
                'financeiro:pagamento_pessoal_detalhe',
                kwargs={'pk': pagamento.pk},
            )
        itens = list(pagamento.itens.select_related('categoria').order_by('pk'))
        return render(
            request,
            'financeiro/partials/pagamento_pessoal_excluir_modal.html',
            {
                'pagamento': pagamento,
                'itens': itens,
                'total_itens': pagamento.total_itens(),
            },
        )

    funcionario_nome = pagamento.funcionario.nome if pagamento.funcionario_id else 'Geral'
    with transaction.atomic():
        pagamento.itens.all().delete()
        pagamento.delete()
    messages.success(request, f'Pagamento pessoal de {funcionario_nome} excluído.')
    if _is_htmx(request):
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:dashboard')
        return response
    return redirect_empresa(request, 'financeiro:dashboard')


@login_required
def pagamento_imposto_novo(request):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    form = PagamentoImpostoForm(request.POST or None, empresa=empresa)
    itens_fs = PagamentoImpostoItemFormSet(
        request.POST or None,
        prefix='itens',
        form_kwargs={'empresa': empresa},
    )

    if request.method == 'POST' and form.is_valid() and itens_fs.is_valid():
        with transaction.atomic():
            pagamento = form.save(commit=False)
            item_form = None
            for f in itens_fs:
                if not getattr(f, 'cleaned_data', None):
                    continue
                if f.cleaned_data.get('DELETE') or f.cleaned_data.get('_skip_form'):
                    continue
                item_form = f
                break
            if item_form is None:
                raise ValidationError('Informe a descrição do pagamento.')
            pagamento.data_pagamento = item_form.cleaned_data.get('data_pagamento') or timezone.localdate()
            pagamento.full_clean()
            pagamento.save()
            item = item_form.save(commit=False)
            item.pagamento = pagamento
            item.full_clean()
            item.save()
        messages.success(request, 'Pagamento de imposto registrado.')
        return redirect_empresa(
            request,
            'financeiro:pagamento_imposto_detalhe',
            kwargs={'pk': pagamento.pk},
        )

    return render(
        request,
        'financeiro/pagamento_imposto_form.html',
        {
            'page_title': 'Pagamento Impostos',
            'form': form,
            'itens_fs': itens_fs,
        },
    )


@login_required
def pagamento_imposto_detalhe(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    pagamento = get_object_or_404(
        PagamentoImposto.objects.filter(empresa=empresa).select_related('autoridade', 'caixa'),
        pk=pk,
    )
    itens = list(pagamento.itens.select_related('categoria').order_by('pk'))
    return render(
        request,
        'financeiro/pagamento_imposto_detalhe.html',
        {
            'page_title': 'Pagamento Impostos',
            'pagamento': pagamento,
            'itens': itens,
            'total_itens': pagamento.total_itens(),
        },
    )


@login_required
def pagamento_imposto_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    pagamento = get_object_or_404(
        PagamentoImposto.objects.filter(empresa=empresa).select_related('autoridade', 'caixa'),
        pk=pk,
    )
    item = pagamento.itens.order_by('pk').first()
    initial_item = []
    if item:
        initial_item.append(
            {
                'descricao': item.descricao,
                'categoria': item.categoria_id,
                'data_pagamento': pagamento.data_pagamento.isoformat() if pagamento.data_pagamento else '',
                'valor_total': format_decimal_br_moeda(item.valor_total),
            }
        )
    form = PagamentoImpostoForm(request.POST or None, instance=pagamento, empresa=empresa)
    itens_fs = PagamentoImpostoItemFormSet(
        request.POST or None,
        prefix='itens',
        form_kwargs={'empresa': empresa},
        initial=initial_item,
    )
    if request.method == 'POST' and form.is_valid() and itens_fs.is_valid():
        with transaction.atomic():
            pagamento = form.save(commit=False)
            item_form = None
            for f in itens_fs:
                if not getattr(f, 'cleaned_data', None):
                    continue
                if f.cleaned_data.get('DELETE') or f.cleaned_data.get('_skip_form'):
                    continue
                item_form = f
                break
            if item_form is None:
                raise ValidationError('Informe a descrição do pagamento.')
            pagamento.data_pagamento = item_form.cleaned_data.get('data_pagamento') or timezone.localdate()
            pagamento.full_clean()
            pagamento.save()
            pagamento.itens.all().delete()
            novo_item = item_form.save(commit=False)
            novo_item.pagamento = pagamento
            novo_item.full_clean()
            novo_item.save()
        messages.success(request, 'Pagamento de imposto atualizado.')
        return redirect_empresa(
            request,
            'financeiro:pagamento_imposto_detalhe',
            kwargs={'pk': pagamento.pk},
        )
    return render(
        request,
        'financeiro/pagamento_imposto_form.html',
        {
            'page_title': 'Editar Pagamento Impostos',
            'form': form,
            'itens_fs': itens_fs,
            'pagamento': pagamento,
            'modo': 'editar',
        },
    )


@login_required
def pagamento_imposto_excluir(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')
    pagamento = get_object_or_404(
        PagamentoImposto.objects.filter(empresa=empresa).select_related('autoridade'),
        pk=pk,
    )
    if request.method != 'POST':
        if not _is_htmx(request):
            return redirect_empresa(
                request,
                'financeiro:pagamento_imposto_detalhe',
                kwargs={'pk': pagamento.pk},
            )
        return render(
            request,
            'financeiro/partials/pagamento_imposto_excluir_modal.html',
            {
                'pagamento': pagamento,
                'total_itens': pagamento.total_itens(),
            },
        )
    autoridade_nome = pagamento.autoridade.nome
    with transaction.atomic():
        pagamento.itens.all().delete()
        pagamento.delete()
    messages.success(request, f'Pagamento de imposto de {autoridade_nome} excluído.')
    if _is_htmx(request):
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:dashboard')
        return response
    return redirect_empresa(request, 'financeiro:dashboard')


@login_required
def pagamento_nf_editar(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    nf = get_object_or_404(PagamentoNotaFiscal.objects.filter(empresa=empresa), pk=pk)
    form = PagamentoNotaFiscalForm(request.POST or None, instance=nf, empresa=empresa)
    # Para edição: por ora, regrava itens apagando e recriando (MVP).
    itens_fs = PagamentoNotaFiscalItemEditFormSet(
        request.POST or None,
        prefix='itens',
        form_kwargs={
            'empresa': empresa,
            'default_caixa_id': nf.caixa_id,
        },
        initial=[
            {
                'tipo': i.tipo,
                'descricao': i.descricao,
                'categoria': i.categoria_id,
                'quantidade': i.quantidade,
                'unidade': i.unidade,
                'valor_unitario': format_decimal_br_moeda(i.valor_unitario),
                'valor_total': format_decimal_br_moeda(i.valor_total),
                'caixa': i.caixa_id,
            }
            for i in nf.itens.all().order_by('pk')
        ],
    )
    pagamentos_fs = PagamentoNotaFiscalPagamentoEditFormSet(
        request.POST or None,
        prefix='pagamentos',
        initial=_initial_pagamentos_nf(nf),
        form_kwargs={'default_data': nf.data_emissao, 'empresa': empresa},
    )
    boletos_fs = BoletoRascunhoFormSet(
        request.POST or None,
        prefix='boletos',
        initial=[
            {
                'numero_doc': b.numero_doc,
                'parcela': b.parcela,
                'vencimento': b.vencimento.isoformat() if b.vencimento else '',
                'valor': format_decimal_br_moeda(b.valor),
            }
            for b in nf.boletos.all().order_by('vencimento', 'parcela', 'pk')
        ],
    )
    dup_nf = None
    force_save = str(request.POST.get('force_save', '')).strip() == '1'

    if request.method == 'POST':
        snapshot_antes = _snapshot_pagamento_nf(nf)
        form_ok = form.is_valid()
        itens_ok = itens_fs.is_valid()
        pagamentos_ok = pagamentos_fs.is_valid()
        boletos_ok = boletos_fs.is_valid()
        ok = form_ok and itens_ok and pagamentos_ok and boletos_ok
        if ok:
            cd = form.cleaned_data
            dup_nf = None
            if nf_numero_exige_unicidade_para_fornecedor(cd.get('numero_nf') or ''):
                dup_nf = (
                    PagamentoNotaFiscal.objects.filter(
                        empresa=empresa,
                        fornecedor=cd['fornecedor'],
                        numero_nf=cd['numero_nf'],
                    )
                    .exclude(pk=nf.pk)
                    .order_by('-pk')
                    .first()
                )
            if dup_nf and not force_save:
                ok = False

        if ok:
            with transaction.atomic():
                nf = form.save()

                nf.itens.all().delete()
                itens_objs = []
                for f in itens_fs:
                    if not getattr(f, 'cleaned_data', None):
                        continue
                    if f.cleaned_data.get('DELETE'):
                        continue
                    if f.cleaned_data.get('_skip_form'):
                        continue
                    item = f.save(commit=False)
                    item.pagamento_nf = nf
                    if not item.caixa_id:
                        item.caixa = nf.caixa
                    item.full_clean()
                    itens_objs.append(item)
                if not itens_objs:
                    raise ValidationError('Informe pelo menos 1 item da NF.')
                PagamentoNotaFiscalItem.objects.bulk_create(itens_objs)

                _salvar_pagamentos_nf(nf, pagamentos_fs, boletos_fs, nf.total_itens())

            messages.success(request, 'Pagamento por NF atualizado.')
            snapshot_depois = _snapshot_pagamento_nf(nf)
            alteracoes = _diff_snapshot_nf(snapshot_antes, snapshot_depois)
            _audit_nf(
                request,
                nf,
                acao='update',
                resumo=f'Editou Nota Fiscal {nf.numero_nf}',
                alteracoes=alteracoes or ['Registro salvo sem alteração detectada'],
            )
            return redirect_empresa(
                request,
                'financeiro:pagamento_nf_detalhe',
                kwargs={'pk': nf.pk},
            )

    return render(
        request,
        'financeiro/pagamento_nf_form.html',
        {
            'page_title': 'Nota Fiscal',
            'form': form,
            'itens_fs': itens_fs,
            'pagamentos_fs': pagamentos_fs,
            'boletos_fs': boletos_fs,
            'modo': 'editar',
            'nf': nf,
            'dup_nf': dup_nf,
            'show_dup_modal': bool(dup_nf) and request.method == 'POST' and not force_save,
            'has_validation_errors': _pagamento_nf_tem_erros_validacao(
                form,
                itens_fs,
                pagamentos_fs,
                boletos_fs,
            ),
            'active_tab': _active_tab_pagamento_nf(
                request,
                form,
                itens_fs,
                pagamentos_fs,
                boletos_fs,
            ),
            'pagamento_tipo_boletos': PagamentoNotaFiscalPagamento.TipoPagamento.BOLETOS,
            'contas_bancarias_pagamento': ContaBancaria.objects.filter(
                empresa=empresa,
                ativo=True,
            ).order_by('banco', 'nome'),
        },
    )


@login_required
def pagamento_nf_detalhe(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    nf = get_object_or_404(
        PagamentoNotaFiscal.objects.filter(empresa=empresa).select_related(
            'fornecedor',
            'caixa',
        ),
        pk=pk,
    )
    itens = list(
        nf.itens.select_related('categoria', 'caixa').order_by('pk')
    )
    pagamentos = list(nf.pagamentos.order_by('data', 'pk'))
    hoje = timezone.localdate()
    boletos = [
        _aplicar_situacao_boleto(boleto, hoje)
        for boleto in nf.boletos.order_by('vencimento', 'parcela', 'pk')
    ]
    pode_pagar_boletos = any(
        boleto.status not in (
            BoletoPagamento.Status.PAGO,
            BoletoPagamento.Status.CANCELADO,
        )
        for boleto in boletos
    )

    total_itens = nf.total_itens()
    total_boletos = sum((b.valor for b in boletos), Decimal('0'))
    resumo_pagamento = _resumo_pagamento_nf(pagamentos, boletos, total_itens, hoje)
    registros = []
    registros_qs = (
        RegistroAuditoria.objects.filter(
            empresa=empresa,
            modulo='financeiro',
            detalhes__nf_pk=nf.pk,
        )
        .select_related('usuario')
        .order_by('-criado_em', '-pk')[:30]
    )
    for registro in registros_qs:
        detalhes = registro.detalhes or {}
        registros.append(
            {
                'criado_em': registro.criado_em,
                'usuario': registro.usuario,
                'acao': registro.acao,
                'acao_label': registro.get_acao_display(),
                'resumo': registro.resumo,
                'alteracoes': detalhes.get('alteracoes') or [],
            }
        )

    return render(
        request,
        'financeiro/pagamento_nf_detalhe.html',
        {
            'page_title': f'Nota Fiscal nº {nf.numero_nf}',
            'nf': nf,
            'itens': itens,
            'pagamentos': pagamentos,
            'boletos': boletos,
            'pode_pagar_boletos': pode_pagar_boletos,
            'total_itens': total_itens,
            'total_boletos': total_boletos,
            'resumo_pagamento': resumo_pagamento,
            'registros': registros,
        },
    )


def _format_quantidade_nf_pdf(valor) -> str:
    if valor is None:
        return '-'
    return format_decimal_br_moeda(valor, decimal_places=4)


def _nf_pdf_header(empresa, nf, styles):
    nome = _nome_empresa_pdf(empresa)
    linhas = [f'<b>{xml_escape(nome)}</b>']
    cnpj = (getattr(empresa, 'cnpj', None) or '').strip()
    if cnpj:
        linhas.append(xml_escape(f'CNPJ: {cnpj}'))
    endereco = (getattr(empresa, 'endereco', None) or '').strip()
    if endereco:
        linhas.append(xml_escape(endereco))
    telefone = (getattr(empresa, 'telefone', None) or '').strip()
    if telefone:
        linhas.append(xml_escape(f'Tel: {telefone}'))
    email = (getattr(empresa, 'email', None) or '').strip()
    if email:
        linhas.append(xml_escape(f'E-mail: {email}'))

    hdr_style = ParagraphStyle(
        'nf_pdf_header',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#0f172a'),
    )
    logo = _empresa_logo_flowable(empresa) or ''
    table = Table(
        [[logo, Paragraph('<br/>'.join(linhas), hdr_style)]],
        colWidths=[34 * mm, 156 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
        )
    )
    return [table, Spacer(1, 6 * mm)]


def _nf_pdf_section_title(text, styles):
    style = ParagraphStyle(
        f'nf_pdf_title_{re.sub(r"[^a-z0-9]+", "_", text.lower())}',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#0f172a'),
        spaceBefore=2 * mm,
        spaceAfter=2 * mm,
    )
    return Paragraph(xml_escape(text), style)


def _nf_pdf_key_value_card(rows, styles):
    label_style = ParagraphStyle(
        'nf_pdf_label',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#475569'),
    )
    value_style = ParagraphStyle(
        'nf_pdf_value',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#0f172a'),
    )
    data = [
        [
            Paragraph(xml_escape(label), label_style),
            Paragraph(xml_escape(str(value if value not in (None, '') else '-')), value_style),
        ]
        for label, value in rows
    ]
    table = Table(data, colWidths=[42 * mm, 148 * mm], hAlign='LEFT')
    table.setStyle(
        TableStyle(
            [
                ('BOX', (0, 0), (-1, -1), 0.35, colors.HexColor('#cbd5e1')),
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


@login_required
def pagamento_nf_pdf(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    nf = get_object_or_404(
        PagamentoNotaFiscal.objects.filter(empresa=empresa).select_related(
            'fornecedor',
            'caixa',
        ),
        pk=pk,
    )
    itens = list(nf.itens.select_related('categoria', 'caixa').order_by('pk'))
    pagamentos = list(
        nf.pagamentos.select_related('conta_bancaria').order_by('data', 'pk')
    )
    hoje = timezone.localdate()
    boletos = [
        _aplicar_situacao_boleto(boleto, hoje)
        for boleto in nf.boletos.select_related('conta_bancaria').order_by('vencimento', 'parcela', 'pk')
    ]
    total_itens = nf.total_itens()
    total_boletos = sum((b.valor for b in boletos), Decimal('0'))
    resumo_pagamento = _resumo_pagamento_nf(pagamentos, boletos, total_itens, hoje)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f'Nota Fiscal nº {nf.numero_nf}',
    )
    styles = getSampleStyleSheet()
    story = []
    story.extend(_nf_pdf_header(empresa, nf, styles))
    story.extend(_flowables_titulo_pdf_centro(f'Detalhes da NF nº {nf.numero_nf}', styles))

    story.append(_nf_pdf_section_title('Descrição', styles))
    story.append(
        _nf_pdf_key_value_card(
            [
                ('Fornecedor', nf.fornecedor.nome),
                ('Nº NF', nf.numero_nf),
                ('Data de emissão', _format_data_pdf(nf.data_emissao)),
                ('Caixa', nf.caixa.nome if nf.caixa_id else '-'),
                ('Descrição', nf.descricao or '-'),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 4 * mm))

    story.append(_nf_pdf_section_title('Itens', styles))
    cell = ParagraphStyle('nf_pdf_cell', parent=styles['Normal'], fontSize=6.6, leading=8)
    cell_right = ParagraphStyle('nf_pdf_cell_right', parent=cell, alignment=2)
    cell_center = ParagraphStyle('nf_pdf_cell_center', parent=cell, alignment=1)
    item_rows = [['#', 'Descrição', 'Categoria', 'Qtd', 'Valor Unit', 'Und', 'Valor Total']]
    for index, item in enumerate(itens, start=1):
        item_rows.append(
            [
                str(index),
                _pdf_cell(item.descricao, cell),
                _pdf_cell(item.categoria.nome if item.categoria_id else '-', cell),
                _pdf_cell(_format_quantidade_nf_pdf(item.quantidade), cell_center),
                _pdf_cell(_format_moeda_pdf(item.valor_unitario), cell_right),
                _pdf_cell(item.unidade or '-', cell_center),
                _pdf_cell(_format_moeda_pdf(item.valor_total), cell_right),
            ]
        )
    if len(item_rows) == 1:
        item_rows.append(['-', 'Sem itens.', '-', '-', '-', '-', '-'])
    itens_table = Table(
        item_rows,
        colWidths=[8 * mm, 54 * mm, 34 * mm, 18 * mm, 24 * mm, 16 * mm, 28 * mm],
        repeatRows=1,
    )
    itens_table.setStyle(_pdf_table_style())
    itens_table.setStyle(
        TableStyle(
            [
                ('ALIGN', (3, 1), (6, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
            ]
        )
    )
    story.append(itens_table)
    story.append(Spacer(1, 4 * mm))

    story.append(_nf_pdf_section_title('Resumo financeiro', styles))
    story.append(
        _nf_pdf_key_value_card(
            [
                ('Valor total da Nota', _format_moeda_pdf(total_itens)),
                ('Pagamento', resumo_pagamento['pagamento_label']),
                ('Situação', resumo_pagamento['situacao_label']),
                ('Valor pago', _format_moeda_pdf(resumo_pagamento['valor_pago'])),
                ('Valor em aberto', _format_moeda_pdf(resumo_pagamento['valor_em_aberto'])),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 4 * mm))

    story.append(_nf_pdf_section_title('Pagamento', styles))
    if pagamentos:
        pagamento_rows = [['Tipo', 'Data', 'Valor', 'Acréscimos', 'Descontos', 'Pago em', 'Observação']]
        for pagamento in pagamentos:
            conta = (
                f'{pagamento.conta_bancaria.nome} - {pagamento.conta_bancaria.banco}'
                if pagamento.conta_bancaria_id
                else 'DINHEIRO'
            )
            pagamento_rows.append(
                [
                    pagamento.get_tipo_display(),
                    _format_data_pdf(pagamento.data),
                    _pdf_cell(_format_moeda_pdf(pagamento.valor), cell_right),
                    _pdf_cell(_format_moeda_pdf(pagamento.acrescimos), cell_right),
                    _pdf_cell(_format_moeda_pdf(pagamento.descontos), cell_right),
                    _pdf_cell(conta, cell),
                    _pdf_cell(pagamento.observacao or '-', cell),
                ]
            )
        pagamento_table = Table(
            pagamento_rows,
            colWidths=[22 * mm, 20 * mm, 24 * mm, 24 * mm, 24 * mm, 42 * mm, 26 * mm],
            repeatRows=1,
        )
        pagamento_table.setStyle(_pdf_table_style())
        story.append(pagamento_table)
        story.append(Spacer(1, 3 * mm))

    if boletos:
        boleto_rows = [['Doc', 'Vencimento', 'Situação', 'Valor', 'Valor pago', 'Pago em']]
        for boleto in boletos:
            conta = (
                f'{boleto.conta_bancaria.nome} - {boleto.conta_bancaria.banco}'
                if boleto.valor_pago and boleto.conta_bancaria_id
                else ('DINHEIRO' if boleto.valor_pago else '-')
            )
            boleto_rows.append(
                [
                    boleto.numero_doc,
                    _format_data_pdf(boleto.vencimento),
                    boleto.situacao_label,
                    _pdf_cell(_format_moeda_pdf(boleto.valor), cell_right),
                    _pdf_cell(_format_moeda_pdf(boleto.valor_pago), cell_right) if boleto.valor_pago else '-',
                    _pdf_cell(conta, cell),
                ]
            )
        boleto_table = Table(
            boleto_rows,
            colWidths=[34 * mm, 24 * mm, 24 * mm, 28 * mm, 28 * mm, 44 * mm],
            repeatRows=1,
        )
        boleto_table.setStyle(_pdf_table_style())
        story.append(boleto_table)
        story.append(Spacer(1, 2 * mm))
        story.append(
            _nf_pdf_key_value_card([('Total boletos', _format_moeda_pdf(total_boletos))], styles)
        )

    if not pagamentos and not boletos:
        story.append(_nf_pdf_key_value_card([('Pagamento', 'Sem pagamentos registrados.')], styles))

    story.extend(_flowables_rodape_impressao(request, styles, space_before_mm=4))
    doc.build(story)

    filename = f'NF_{_safe_filename_part(str(nf.numero_nf))}_{_safe_filename_part(nf.fornecedor.nome)}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def pagamento_nf_fornecedor_info(request):
    empresa = _empresa(request)
    if not empresa:
        return JsonResponse({'ok': False, 'error': 'Empresa não selecionada.'}, status=400)

    fornecedor_id = _int_param(request, 'fornecedor')
    numero_nf = (request.GET.get('numero_nf') or '').strip()
    exclude_pk = _int_param(request, 'exclude')

    ultimos = []
    duplicate = None

    if fornecedor_id:
        nfs_qs = (
            PagamentoNotaFiscal.objects.filter(
                empresa=empresa,
                fornecedor_id=fornecedor_id,
            )
            .select_related('fornecedor')
            .annotate(
                total_itens_info=Coalesce(
                    Subquery(
                        _subquery_total_itens_nf(),
                        output_field=DecimalField(max_digits=16, decimal_places=2),
                    ),
                    Value(Decimal('0')),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                )
            )
            .order_by('-criado_em', '-pk')
        )
        ultimos_qs = nfs_qs
        if exclude_pk:
            ultimos_qs = ultimos_qs.exclude(pk=exclude_pk)
        for nf in ultimos_qs[:5]:
            ultimos.append(
                {
                    'pk': nf.pk,
                    'data': nf.data_emissao.strftime('%d/%m/%Y') if nf.data_emissao else '-',
                    'numero_nf': nf.numero_nf or '-',
                    'valor': format_decimal_br_moeda(nf.total_itens_info or Decimal('0')),
                    'url': reverse_empresa(
                        request,
                        'financeiro:pagamento_nf_detalhe',
                        kwargs={'pk': nf.pk},
                    ),
                }
            )

        if numero_nf and nf_numero_exige_unicidade_para_fornecedor(numero_nf):
            dup_qs = nfs_qs.filter(numero_nf=numero_nf)
            if exclude_pk:
                dup_qs = dup_qs.exclude(pk=exclude_pk)
            dup_nf = dup_qs.first()
            if dup_nf:
                duplicate = {
                    'pk': dup_nf.pk,
                    'fornecedor': dup_nf.fornecedor.nome,
                    'numero_nf': dup_nf.numero_nf,
                    'data': dup_nf.data_emissao.strftime('%d/%m/%Y') if dup_nf.data_emissao else '-',
                    'url': reverse_empresa(
                        request,
                        'financeiro:pagamento_nf_detalhe',
                        kwargs={'pk': dup_nf.pk},
                    ),
                }

    return JsonResponse({'ok': True, 'ultimos': ultimos, 'duplicate': duplicate})


@login_required
def pagamento_nf_pagar_boleto(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    nf = get_object_or_404(
        PagamentoNotaFiscal.objects.filter(empresa=empresa).select_related('fornecedor'),
        pk=pk,
    )
    detalhe_url = reverse_empresa(
        request,
        'financeiro:pagamento_nf_detalhe',
        kwargs={'pk': nf.pk},
    )
    boletos_qs = nf.boletos.exclude(status=BoletoPagamento.Status.CANCELADO).order_by(
        'vencimento', 'parcela', 'pk'
    )
    boletos_para_pagamento = boletos_qs.exclude(status=BoletoPagamento.Status.PAGO)
    if request.GET.get('multi') == '1' and request.method == 'GET':
        boletos_abertos = [
            _aplicar_situacao_boleto(boleto)
            for boleto in boletos_para_pagamento
        ]
        if not boletos_abertos:
            messages.info(request, 'NÃ£o hÃ¡ boletos em aberto para pagamento.')
            if _is_htmx(request):
                response = HttpResponse(status=200)
                response['HX-Redirect'] = detalhe_url
                return response
            return redirect(detalhe_url)
        for boleto in boletos_abertos:
            boleto.data_pagamento_valor = timezone.localdate().isoformat()
        return render(
            request,
            'financeiro/partials/pagamento_nf_pagar_boletos_modal.html',
            {
                'nf': nf,
                'boletos': boletos_abertos,
                'hoje': timezone.localdate(),
                'contas_bancarias_pagamento': ContaBancaria.objects.filter(
                    empresa=empresa,
                    ativo=True,
                ).order_by('banco', 'nome'),
            },
        )

    if request.method == 'POST' and request.POST.get('action') == 'salvar_multiplos':
        boleto_ids = request.POST.getlist('boletos')
        boletos_selecionados = list(boletos_para_pagamento.filter(pk__in=boleto_ids))
        erros = []
        if not boletos_selecionados:
            erros.append('Selecione pelo menos um boleto para pagar.')

        datas_pagamento = {}
        conta_bancaria = None
        try:
            conta_bancaria_id = int(request.POST.get('conta_bancaria') or '')
        except (TypeError, ValueError):
            conta_bancaria_id = None
        if conta_bancaria_id:
            conta_bancaria = ContaBancaria.objects.filter(
                empresa=empresa,
                ativo=True,
                pk=conta_bancaria_id,
            ).first()
            if not conta_bancaria:
                erros.append('Conta bancária inválida para esta empresa.')
        for boleto in boletos_selecionados:
            data_raw = request.POST.get(f'data_pagamento_{boleto.pk}')
            data_pagamento = parse_date(data_raw or '')
            if not data_pagamento:
                erros.append(f'Informe a data de pagamento do boleto {boleto.numero_doc or boleto.parcela}.')
            else:
                datas_pagamento[boleto.pk] = data_pagamento

        if not erros:
            with transaction.atomic():
                for boleto in boletos_selecionados:
                    boleto.data_pagamento = datas_pagamento[boleto.pk]
                    boleto.acrescimos = Decimal('0')
                    boleto.descontos = Decimal('0')
                    boleto.valor_pago = boleto.valor
                    boleto.conta_bancaria = conta_bancaria
                    boleto.observacao = ''
                    boleto.status = BoletoPagamento.Status.PAGO
                    boleto.full_clean()
                    boleto.save(
                        update_fields=[
                            'data_pagamento',
                            'acrescimos',
                            'descontos',
                            'valor_pago',
                            'conta_bancaria',
                            'observacao',
                            'status',
                            'atualizado_em',
                        ]
                    )
            messages.success(request, f'{len(boletos_selecionados)} boleto(s) pago(s) com sucesso.')
            _audit_nf(
                request,
                nf,
                acao='update',
                resumo=f'Registrou pagamento de {len(boletos_selecionados)} boleto(s) da NF {nf.numero_nf}',
                alteracoes=[
                    f'Boleto {boleto.numero_doc or boleto.parcela}: pago em {datas_pagamento[boleto.pk].strftime("%d/%m/%Y")}'
                    for boleto in boletos_selecionados
                ],
            )
            if _is_htmx(request):
                response = HttpResponse(status=200)
                response['HX-Redirect'] = detalhe_url
                return response
            return redirect(detalhe_url)

        boletos_abertos = []
        selected_ids = set(str(pk) for pk in boleto_ids)
        for boleto in boletos_para_pagamento:
            boleto = _aplicar_situacao_boleto(boleto)
            boleto.is_selected = str(boleto.pk) in selected_ids
            boleto.data_pagamento_valor = request.POST.get(
                f'data_pagamento_{boleto.pk}',
                timezone.localdate().isoformat(),
            )
            boletos_abertos.append(boleto)
        return render(
            request,
            'financeiro/partials/pagamento_nf_pagar_boletos_modal.html',
            {
                'nf': nf,
                'boletos': boletos_abertos,
                'hoje': timezone.localdate(),
                'erros': erros,
                'contas_bancarias_pagamento': ContaBancaria.objects.filter(
                    empresa=empresa,
                    ativo=True,
                ).order_by('banco', 'nome'),
            },
        )

    selected_boleto = None
    selected_boleto_id = (
        request.POST.get('boleto')
        or request.GET.get('boleto')
    )
    if selected_boleto_id:
        selected_boleto = boletos_qs.filter(pk=selected_boleto_id).first()
    if not selected_boleto:
        selected_boleto = (
            boletos_qs.exclude(status=BoletoPagamento.Status.PAGO).first()
            or boletos_qs.first()
        )
        if selected_boleto:
            selected_boleto_id = str(selected_boleto.pk)
    modo = 'editar' if selected_boleto and selected_boleto.status == BoletoPagamento.Status.PAGO else 'pagar'

    if request.method == 'POST':
        action = request.POST.get('action') or 'salvar'
        form = BoletoPagamentoForm(
            request.POST,
            boletos=boletos_qs,
            selected_boleto=selected_boleto,
            empresa=empresa,
        )
        if action == 'excluir_pagamento':
            boleto = get_object_or_404(boletos_qs, pk=request.POST.get('boleto'))
            with transaction.atomic():
                boleto.status = BoletoPagamento.Status.RASCUNHO
                boleto.data_pagamento = None
                boleto.acrescimos = Decimal('0')
                boleto.descontos = Decimal('0')
                boleto.valor_pago = None
                boleto.conta_bancaria = None
                boleto.observacao = ''
                boleto.full_clean()
                boleto.save(
                    update_fields=[
                        'status',
                        'data_pagamento',
                        'acrescimos',
                        'descontos',
                        'valor_pago',
                        'conta_bancaria',
                        'observacao',
                        'atualizado_em',
                    ]
                )
            messages.success(request, f'Pagamento do boleto {boleto.numero_doc} excluído.')
            _audit_nf(
                request,
                nf,
                acao='update',
                resumo=f'Excluiu pagamento do boleto {boleto.numero_doc} da NF {nf.numero_nf}',
                alteracoes=[
                    f'Boleto {boleto.numero_doc or boleto.parcela}: voltou para não pago',
                ],
            )
            if _is_htmx(request):
                response = HttpResponse(status=200)
                response['HX-Redirect'] = detalhe_url
                return response
            return redirect(detalhe_url)
        if form.is_valid():
            boleto = form.cleaned_data['boleto']
            with transaction.atomic():
                boleto.data_pagamento = form.cleaned_data['data_pagamento']
                boleto.acrescimos = form.cleaned_data['acrescimos']
                boleto.descontos = form.cleaned_data['descontos']
                boleto.valor_pago = form.cleaned_data['valor_pago']
                boleto.conta_bancaria = form.cleaned_data.get('conta_bancaria')
                boleto.observacao = form.cleaned_data['observacao']
                boleto.status = BoletoPagamento.Status.PAGO
                boleto.full_clean()
                boleto.save(
                    update_fields=[
                        'data_pagamento',
                        'acrescimos',
                        'descontos',
                        'valor_pago',
                        'conta_bancaria',
                        'observacao',
                        'status',
                        'atualizado_em',
                    ]
                )
            messages.success(request, f'Boleto {boleto.numero_doc} pago com sucesso.')
            _audit_nf(
                request,
                nf,
                acao='update',
                resumo=f'Registrou pagamento do boleto {boleto.numero_doc} da NF {nf.numero_nf}',
                alteracoes=[
                    f'Data de pagamento: {boleto.data_pagamento.strftime("%d/%m/%Y") if boleto.data_pagamento else "—"}',
                    f'Valor pago: {_decimal_audit(boleto.valor_pago or Decimal("0"))}',
                    f'Acréscimos: {_decimal_audit(boleto.acrescimos)}',
                    f'Descontos: {_decimal_audit(boleto.descontos)}',
                ],
            )
            if _is_htmx(request):
                response = HttpResponse(status=200)
                response['HX-Redirect'] = detalhe_url
                return response
            return redirect(detalhe_url)
    else:
        form = BoletoPagamentoForm(
            boletos=boletos_qs,
            selected_boleto=selected_boleto,
            empresa=empresa,
        )
        if not _is_htmx(request):
            return redirect(detalhe_url)

    boletos_abertos = [
        _aplicar_situacao_boleto(boleto)
        for boleto in boletos_qs
    ]
    if not boletos_abertos:
        messages.info(request, 'Não há boletos disponíveis para pagamento.')
        if _is_htmx(request):
            response = HttpResponse(status=200)
            response['HX-Redirect'] = detalhe_url
            return response
        return redirect(detalhe_url)

    if not selected_boleto_id and boletos_abertos:
        selected_boleto_id = str(boletos_abertos[0].pk)
    for boleto in boletos_abertos:
        boleto.is_selected = str(boleto.pk) == str(selected_boleto_id)

    return render(
        request,
        'financeiro/partials/pagamento_nf_pagar_boleto_modal.html',
        {
            'nf': nf,
            'form': form,
            'boletos': boletos_abertos,
            'selected_boleto_id': selected_boleto_id,
            'selected_boleto': selected_boleto,
            'modo': modo,
        },
    )


@login_required
def pagamento_nf_excluir(request, pk: int):
    empresa = _empresa(request)
    if not empresa:
        return redirect('selecionar_empresa')

    nf = get_object_or_404(
        PagamentoNotaFiscal.objects.filter(empresa=empresa).select_related(
            'fornecedor',
        ),
        pk=pk,
    )
    numero_nf = nf.numero_nf

    if request.method != 'POST':
        if not _is_htmx(request):
            return redirect_empresa(
                request,
                'financeiro:pagamento_nf_detalhe',
                kwargs={'pk': nf.pk},
            )
        boletos = list(nf.boletos.order_by('vencimento', 'parcela', 'pk'))
        pagamentos = list(nf.pagamentos.order_by('data', 'pk'))
        total_boletos = sum((boleto.valor for boleto in boletos), Decimal('0'))
        total_pago_boletos = sum(
            (boleto.valor_pago or Decimal('0') for boleto in boletos),
            Decimal('0'),
        )
        return render(
            request,
            'financeiro/partials/pagamento_nf_excluir_modal.html',
            {
                'nf': nf,
                'pagamentos': pagamentos,
                'boletos': boletos,
                'total_boletos': total_boletos,
                'total_pago_boletos': total_pago_boletos,
            },
        )

    fornecedor_nome = nf.fornecedor.nome
    total_itens = nf.total_itens()
    with transaction.atomic():
        _audit_nf(
            request,
            nf,
            acao='delete',
            resumo=f'Excluiu Nota Fiscal {numero_nf}',
            alteracoes=[
                f'Fornecedor: {fornecedor_nome}',
                f'Valor total: {_decimal_audit(total_itens)}',
            ],
        )
        nf.boletos.all().delete()
        PagamentoNotaFiscalPagamento.objects.filter(pagamento_nf=nf).delete()
        nf.delete()

    messages.success(
        request,
        f'Nota Fiscal nº {numero_nf} excluída com pagamento e boletos vinculados.',
    )
    if _is_htmx(request):
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse_empresa(request, 'financeiro:dashboard')
        return response
    return redirect_empresa(request, 'financeiro:dashboard')
