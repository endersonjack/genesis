import re
from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape as xml_escape

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from empresas.models import Empresa

from controles_rh.views.cesta_export import (
    _flowables_header_compact,
    _flowables_titulo_pdf_centro,
    _header_text_html_vt,
    _mes_nome_pt,
    _unlink_temp_logo_paths,
)
from controles_rh.views.pdf_rodape import flowables_rodape_impressao
from controles_rh.views.vale_transporte import _get_tabela_vt_empresa, _total_valor_pago_tabela


def _safe_filename_part(text):
    s = re.sub(r'[^\w\s\-]', '_', str(text), flags=re.UNICODE)
    s = re.sub(r'\s+', '_', s.strip())
    return (s[:60] or 'vt').strip('_')


def _itens_export(tabela):
    return tabela.itens.select_related('funcionario').order_by('ordem', 'nome', 'id')


def _resumo_tabela_export(tabela):
    """
    Mesmos totais e metadados do cabeçalho da tela de detalhe (controle VT).
    """
    comp = tabela.competencia
    total_itens = tabela.itens.count()
    total_valor_pago = _total_valor_pago_tabela(tabela)
    total_a_pagar = tabela.total_valor
    saldo_a_pagar = total_a_pagar - total_valor_pago
    if saldo_a_pagar < 0:
        saldo_a_pagar = Decimal('0.00')
    return {
        'competencia': comp,
        'total_itens': total_itens,
        'total_a_pagar': total_a_pagar,
        'total_valor_pago': total_valor_pago,
        'saldo_a_pagar': saldo_a_pagar,
        'status_vt_label': tabela.vt_status_efetivo_label,
    }


def _row_data_pagamento(item):
    if item.data_pagamento:
        return item.data_pagamento.strftime('%d/%m/%Y')
    return '—'


def _fmt_br_decimal(val):
    """Exibe Decimal/float como string pt-BR para PDF."""
    v = float(val or 0)
    return f'{v:.2f}'.replace('.', ',')


# Largura útil da tabela no PDF VT (paisagem A4, margens 12mm): 297 − 24
VT_PDF_TABLE_WIDTH_MM = 273


def _col_widths_vt_pdf_mm(tabela):
    """
    # | NOME (nome-função) | valores | PIX | DT pag.

    Largura de NOME ≈ maior texto (nome + função) medido em Helvetica-Bold 7 pt,
    com margem mínima — evita faixa vazia à direita do nome.
    O que sobra na linha vai para PIX (e colunas numéricas já estreitas).
    """
    total = float(VT_PDF_TABLE_WIDTH_MM)
    w_n = 5.0
    w_vp = 15.0
    w_vpg = 15.0
    w_saldo = 13.0
    w_dt = 16.0
    fixed_rest = w_n + w_vp + w_vpg + w_saldo + w_dt
    min_pix = 50.0
    min_nome = 40.0

    font = 'Helvetica-Bold'
    size = 7
    # Padding lateral da célula (~6 pt) + pequena folga
    pad_pt = 10.0

    itens = list(_itens_export(tabela))
    if not itens:
        w_nome = 70.0
        w_pix = total - fixed_rest - w_nome
        return (w_n, w_nome, w_vp, w_vpg, w_saldo, max(min_pix, w_pix), w_dt)

    max_w_pt = 0.0
    for item in itens:
        text = _vt_pdf_nome_funcao_texto(item)
        max_w_pt = max(max_w_pt, stringWidth(text, font, size))
    nome_pt = max_w_pt + pad_pt
    w_nome_mm = nome_pt * 0.352778

    w_nome_mm = max(min_nome, w_nome_mm)
    max_nome_allowed = total - fixed_rest - min_pix
    w_nome_mm = min(w_nome_mm, max_nome_allowed)

    w_pix = total - fixed_rest - w_nome_mm
    if w_pix < min_pix:
        w_pix = min_pix
        w_nome_mm = total - fixed_rest - w_pix
        w_nome_mm = max(min_nome, w_nome_mm)

    return (w_n, w_nome_mm, w_vp, w_vpg, w_saldo, w_pix, w_dt)


def _vt_pdf_nome_funcao_texto(item) -> str:
    nome_p = ((item.nome_exibicao or '').strip() or '—').upper()
    func_p = (item.funcao or '').strip().upper() or '—'
    return f'{nome_p} - {func_p}'


def _vt_pdf_pix_tipo_banco_texto(item) -> str:
    pix = (item.pix or '').strip() or '—'
    tipo = (item.get_tipo_pix_display() or '—').upper()
    banco = (item.banco or '').strip().upper() or '—'
    return f'{pix.upper()} - {tipo} - {banco}'


@login_required
def exportar_tabela_vt_xlsx(request, pk):
    tabela = _get_tabela_vt_empresa(request, pk)
    comp = tabela.competencia
    empresa = comp.empresa
    resumo = _resumo_tabela_export(tabela)

    headers = [
        '#',
        'Nome',
        'Função',
        'Endereço',
        'Valor a pagar',
        'Valor pago',
        'Saldo',
        'Data pagamento',
        'Pix',
        'Tipo Pix',
        'Banco',
    ]
    ncols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = 'VT'

    row = 1
    if getattr(empresa, 'logo', None) and empresa.logo:
        try:
            with empresa.logo.open('rb') as f:
                xl_img = XLImage(BytesIO(f.read()))
            max_w = 220
            if xl_img.width > max_w:
                ratio = max_w / xl_img.width
                xl_img.width = max_w
                xl_img.height = int(xl_img.height * ratio)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            ws.add_image(xl_img, 'A1')
            ws.row_dimensions[row].height = max(50, min(xl_img.height * 0.78, 130))
            row += 1
        except Exception:
            pass
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f'{tabela.nome} — Competência {comp.referencia} — {comp.empresa}')
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

    desc_txt = (tabela.descricao or '').strip() or '—'
    left_rows = [
        ('Competência', comp.referencia),
        ('Empresa', str(comp.empresa)),
        ('Descrição', desc_txt),
    ]
    right_rows = [
        ('Total de Funcionários', resumo['total_itens']),
        ('Total a pagar (R$)', float(resumo['total_a_pagar'])),
        ('Total pago (R$)', float(resumo['total_valor_pago'])),
        ('Saldo à pagar (R$)', float(resumo['saldo_a_pagar'])),
        ('Status pagamento (VT)', resumo['status_vt_label']),
    ]
    block_start = row
    max_lines = max(len(left_rows), len(right_rows))
    for i in range(max_lines):
        r = block_start + i
        if i < len(left_rows):
            la = ws.cell(row=r, column=1, value=left_rows[i][0])
            la.font = Font(bold=True)
            la.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            lv = ws.cell(row=r, column=2, value=left_rows[i][1])
            lv.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if i < len(right_rows):
            ra = ws.cell(row=r, column=4, value=right_rows[i][0])
            ra.font = Font(bold=True)
            ra.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            rv = ws.cell(row=r, column=5, value=right_rows[i][1])
            rv.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    row = block_start + max_lines + 1

    for col, h in enumerate(headers, 1):
        hc = ws.cell(row=row, column=col, value=h)
        hc.font = Font(bold=True)
        hc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row += 1

    for n, item in enumerate(_itens_export(tabela), start=1):
        if not item.ativo or not item.valor_pagar or item.valor_pagar <= 0:
            saldo_val = '—'
        else:
            saldo_val = float(item.saldo)
        valores = [
            n,
            item.nome_exibicao,
            item.funcao or '',
            item.endereco or '',
            float(item.valor_pagar or 0),
            float(item.valor_pago or 0),
            saldo_val,
            _row_data_pagamento(item),
            item.pix or '',
            item.get_tipo_pix_display() or '',
            item.banco or '',
        ]
        for col, val in enumerate(valores, 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    ws.cell(row=row, column=1, value='TOTAIS').font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(resumo['total_a_pagar'])).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(resumo['total_valor_pago'])).font = Font(bold=True)
    ws.cell(row=row, column=7, value=float(resumo['saldo_a_pagar'])).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    name = _safe_filename_part(tabela.nome)
    ref = f'{comp.mes:02d}_{comp.ano}'
    filename = f'VT_{name}_{ref}.xlsx'

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def exportar_tabela_vt_pdf(request, pk):
    tabela = _get_tabela_vt_empresa(request, pk)
    comp = tabela.competencia
    empresa = Empresa.objects.get(pk=comp.empresa_id)
    resumo = _resumo_tabela_export(tabela)

    styles = getSampleStyleSheet()
    _cw = _col_widths_vt_pdf_mm(tabela)
    # Nome/função: tamanho fixo para todos (sem reduzir fonte em nomes longos).
    nome_vt_style = ParagraphStyle(
        'vt_nome_fix',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
        spaceBefore=0,
        spaceAfter=0,
    )
    # Mesma fonte do corpo; largura da coluna em _col_widths_vt_pdf_mm.
    data_dt_style = ParagraphStyle(
        'vt_dt_pag',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    # Uma linha por item: NOME inclui função; PIX inclui tipo e banco; coluna de data.
    headers = [
        '#',
        'NOME',
        'VLR PAGAR',
        'VLR PAGO',
        'SALDO',
        'PIX',
        'DT. PAG.',
    ]

    # Resumo só com totais/status (competência/empresa já estão no cabeçalho)
    meta_cell_style = ParagraphStyle(
        'vt_meta_cell',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.HexColor('#0f172a'),
    )
    w_meta = (VT_PDF_TABLE_WIDTH_MM / 5.0) * mm
    meta_rows = [
        [
            Paragraph(
                f'<b>FUNC.:</b> {xml_escape(str(resumo["total_itens"]))}',
                meta_cell_style,
            ),
            Paragraph(
                f'<b>A PAGAR (R$):</b> {xml_escape(_fmt_br_decimal(resumo["total_a_pagar"]))}',
                meta_cell_style,
            ),
            Paragraph(
                f'<b>PAGO (R$):</b> {xml_escape(_fmt_br_decimal(resumo["total_valor_pago"]))}',
                meta_cell_style,
            ),
            Paragraph(
                f'<b>SALDO (R$):</b> {xml_escape(_fmt_br_decimal(resumo["saldo_a_pagar"]))}',
                meta_cell_style,
            ),
            Paragraph(
                f'<b>STATUS:</b> {xml_escape(str(resumo["status_vt_label"]).upper())}',
                meta_cell_style,
            ),
        ]
    ]
    pix_para_style = ParagraphStyle(
        'pixcell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        spaceBefore=0,
        spaceAfter=0,
    )
    data_rows = [headers]

    for n, item in enumerate(_itens_export(tabela), start=1):
        if not item.ativo or not item.valor_pagar or item.valor_pagar <= 0:
            saldo_s = '—'
        else:
            saldo_s = f'{item.saldo:.2f}'.replace('.', ',')
        pix_cell = Paragraph(
            xml_escape(_vt_pdf_pix_tipo_banco_texto(item)),
            pix_para_style,
        )
        nome_txt = _vt_pdf_nome_funcao_texto(item)
        nome_cell = Paragraph(xml_escape(nome_txt), nome_vt_style)
        data_txt = _row_data_pagamento(item).upper()
        data_cell = Paragraph(xml_escape(data_txt), data_dt_style)

        data_rows.append(
            [
                str(n),
                nome_cell,
                f'{item.valor_pagar or 0:.2f}'.replace('.', ','),
                f'{item.valor_pago or 0:.2f}'.replace('.', ','),
                saldo_s,
                pix_cell,
                data_cell,
            ]
        )

    data_rows.append(
        [
            'TOTAIS',
            '',
            _fmt_br_decimal(resumo['total_a_pagar']),
            _fmt_br_decimal(resumo['total_valor_pago']),
            _fmt_br_decimal(resumo['saldo_a_pagar']),
            '',
            '',
        ]
    )

    mes_titulo = f'{_mes_nome_pt(comp.mes).upper()} {comp.ano}'
    titulo_doc = f'VALE TRANSPORTE — {mes_titulo}'

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f'VT {tabela.nome}',
    )
    story = []
    temp_logo_paths = []
    story.extend(
        _flowables_header_compact(
            empresa,
            comp,
            tabela,
            styles,
            _header_text_html_vt,
            temp_logo_paths,
        )
    )
    story.extend(_flowables_titulo_pdf_centro(titulo_doc, styles))

    meta_tbl = Table(meta_rows, colWidths=[w_meta, w_meta, w_meta, w_meta, w_meta])
    meta_tbl.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.2, colors.HexColor('#cbd5e1')),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ]
        )
    )
    story.append(meta_tbl)
    story.append(Spacer(1, 3 * mm))

    last_idx = len(data_rows) - 1
    table = Table(data_rows, colWidths=[w * mm for w in _cw], repeatRows=1)
    pdf_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (6, 0), (6, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, last_idx - 1), 'CENTER'),
        ('ALIGN', (2, 1), (4, last_idx - 1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, last_idx - 1), 'LEFT'),
        ('ALIGN', (6, 1), (6, last_idx - 1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, last_idx), (-1, last_idx), colors.HexColor('#e0f2fe')),
        ('FONTNAME', (0, last_idx), (-1, last_idx), 'Helvetica-Bold'),
        ('ALIGN', (2, last_idx), (4, last_idx), 'RIGHT'),
    ]
    if last_idx > 1:
        pdf_style.append(
            (
                'ROWBACKGROUNDS',
                (0, 1),
                (-1, last_idx - 1),
                [colors.white, colors.HexColor('#f8fafc')],
            )
        )
    table.setStyle(TableStyle(pdf_style))
    story.append(table)

    story.extend(flowables_rodape_impressao(request, styles, space_before_mm=3))

    doc.build(story)
    _unlink_temp_logo_paths(temp_logo_paths)
    buf.seek(0)

    name = _safe_filename_part(tabela.nome)
    ref = f'{comp.mes:02d}_{comp.ano}'
    filename = f'VT_{name}_{ref}.pdf'

    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response
